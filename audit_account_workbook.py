#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook

SOURCE = Path('/home/ubuntu/upload/pasted_file_06GgiQ_LCJ経営管理表_经营用账户.xlsx')
OUTPUT = Path('/home/ubuntu/lcjgent_restore/account_import_workbook_audit.json')

EMAIL_RE = re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]+$')
PHONE_RE = re.compile(r'^\+?[0-9][0-9\s()-]{7,}$')


def text(value):
    if value is None:
        return ''
    return str(value).strip()


def secret_fingerprint(value: str) -> str | None:
    if not value:
        return None
    return hashlib.sha256(value.encode('utf-8')).hexdigest()[:12]


def mask_identifier(value: str) -> str:
    if not value:
        return ''
    if EMAIL_RE.match(value):
        local, domain = value.split('@', 1)
        return f"{local[:2]}***@{domain}"
    digits = re.sub(r'\D', '', value)
    if len(digits) >= 8:
        return f"***{digits[-4:]}"
    if len(value) <= 4:
        return '*' * len(value)
    return value[:2] + '***' + value[-2:]


def identifier_kind(value: str) -> str:
    if not value:
        return 'none'
    if EMAIL_RE.match(value):
        return 'email'
    if PHONE_RE.match(value):
        return 'phone'
    if '\n' in value or len(value) > 100:
        return 'long_text'
    return 'username'


def valid_url(value: str) -> bool:
    if not value:
        return False
    try:
        parsed = urlparse(value)
        return parsed.scheme in {'https', 'http'} and bool(parsed.netloc)
    except Exception:
        return False


wb = load_workbook(SOURCE, data_only=False, read_only=False)
report = {
    'sourceName': SOURCE.name,
    'sha256': hashlib.sha256(SOURCE.read_bytes()).hexdigest(),
    'sheetCount': len(wb.sheetnames),
    'sheets': [],
}

for ws in wb.worksheets:
    rows = []
    current_section = ''
    for row_number, cells in enumerate(ws.iter_rows(values_only=False), start=1):
        values = [text(cell.value) for cell in cells]
        purpose = values[0] if len(values) > 0 else ''
        url = values[1] if len(values) > 1 else ''
        account = values[2] if len(values) > 2 else ''
        password = values[3] if len(values) > 3 else ''
        notes = values[4] if len(values) > 4 else ''
        parent = values[7] if len(values) > 7 else ''
        non_empty = [value for value in values if value]
        if row_number == 1:
            row_type = 'header'
        elif len(non_empty) == 1 and purpose:
            current_section = purpose
            row_type = 'section'
        elif not non_empty:
            row_type = 'blank'
        else:
            row_type = 'record'
        row_report = {
            'row': row_number,
            'rowType': row_type,
            'section': current_section,
            'purpose': purpose,
            'parent': parent,
            'hasUrl': bool(url),
            'validUrl': valid_url(url),
            'urlHost': urlparse(url).netloc.lower() if valid_url(url) else '',
            'identifierKind': identifier_kind(account),
            'identifierMasked': mask_identifier(account),
            'hasPassword': bool(password),
            'passwordFingerprint': secret_fingerprint(password),
            'passwordLength': len(password),
            'hasNotes': bool(notes),
            'nonEmptyCells': len(non_empty),
            'hasFormula': any(cell.data_type == 'f' for cell in cells),
            'hidden': ws.row_dimensions[row_number].hidden is True,
        }
        if row_type != 'blank':
            rows.append(row_report)
    report['sheets'].append({
        'title': ws.title,
        'state': ws.sheet_state,
        'maxRow': ws.max_row,
        'maxColumn': ws.max_column,
        'mergedRanges': [str(rng) for rng in ws.merged_cells.ranges],
        'hiddenColumns': [key for key, dim in ws.column_dimensions.items() if dim.hidden],
        'rows': rows,
    })

OUTPUT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
print(json.dumps({
    'output': str(OUTPUT),
    'sha256': report['sha256'],
    'sheetCount': report['sheetCount'],
    'recordRows': sum(1 for sheet in report['sheets'] for row in sheet['rows'] if row['rowType'] == 'record'),
    'sectionRows': sum(1 for sheet in report['sheets'] for row in sheet['rows'] if row['rowType'] == 'section'),
    'formulaRows': sum(1 for sheet in report['sheets'] for row in sheet['rows'] if row['hasFormula']),
}, ensure_ascii=False, indent=2))
