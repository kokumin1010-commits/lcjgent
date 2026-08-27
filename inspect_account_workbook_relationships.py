#!/usr/bin/env python3
from __future__ import annotations

import json
from pathlib import Path
from openpyxl import load_workbook

SOURCE = Path('/home/ubuntu/upload/pasted_file_06GgiQ_LCJ経営管理表_经营用账户.xlsx')
wb_formula = load_workbook(SOURCE, data_only=False, read_only=False)
wb_values = load_workbook(SOURCE, data_only=True, read_only=False)
wsf = wb_formula['经营用账户']
wsv = wb_values['经营用账户']


def text(value):
    return '' if value is None else str(value).strip()


def mask(value):
    value = text(value)
    if not value:
        return ''
    if '@' in value:
        local, domain = value.split('@', 1)
        return f'{local[:2]}***@{domain}'
    digits = ''.join(ch for ch in value if ch.isdigit())
    if len(digits) >= 8:
        return f'***{digits[-4:]}'
    return f'{value[:2]}***{value[-2:]}' if len(value) > 4 else '*' * len(value)

rows = []
for row in list(range(20, 30)) + [41, 43, 44]:
    rows.append({
        'row': row,
        'A_purpose': text(wsf.cell(row, 1).value),
        'B_display': text(wsv.cell(row, 2).value),
        'C_identifierMasked': mask(wsf.cell(row, 3).value),
        'D_hasPassword': bool(text(wsf.cell(row, 4).value)),
        'E_notes': text(wsf.cell(row, 5).value),
        'F': text(wsf.cell(row, 6).value),
        'G': text(wsf.cell(row, 7).value),
        'H_parent': text(wsf.cell(row, 8).value),
        'I_relationType': text(wsf.cell(row, 9).value),
        'J': text(wsf.cell(row, 10).value),
        'K': text(wsf.cell(row, 11).value),
        'L': text(wsf.cell(row, 12).value),
        'M': text(wsf.cell(row, 13).value),
        'N': text(wsf.cell(row, 14).value),
    })
print(json.dumps(rows, ensure_ascii=False, indent=2))
