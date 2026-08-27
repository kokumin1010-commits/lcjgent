#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import quote

import requests
from openpyxl import load_workbook

BASE = "https://lcjmall.com"
SOURCE = Path("/home/ubuntu/upload/pasted_file_06GgiQ_LCJ経営管理表_经营用账户.xlsx")
OUTPUT = Path("/home/ubuntu/lcjgent_hr_persistence_fix/hr_report_store_production_recovery.json")
CONFIRMATION = "RECOVER_MANUAL_HR_REPORT_STORE_2026_08_27"
ALLOWED_FIELDS = {
    "staff": {"name", "email", "employmentType", "employmentTypeEvidence", "emailEvidenceStatus", "isActive"},
    "report_staff": {"name"},
    "managed_stores": {"operatorId", "operatorName", "operator2Id", "operator2Name", "notes"},
}


def trpc_json(response: requests.Response):
    response.raise_for_status()
    payload = response.json()
    if isinstance(payload, list):
        payload = payload[0]
    if "error" in payload:
        data = payload["error"].get("json", {})
        raise RuntimeError(f"{data.get('code', 'TRPC_ERROR')}: {data.get('message') or 'tRPC error'}")
    return payload["result"]["data"]["json"]


def query(session: requests.Session, procedure: str, payload: dict | None = None):
    encoded = quote(json.dumps({"json": payload or {}}, separators=(",", ":")))
    return trpc_json(session.get(f"{BASE}/api/trpc/{procedure}?input={encoded}", timeout=120))


def mutate(session: requests.Session, procedure: str, payload: dict):
    return trpc_json(session.post(f"{BASE}/api/trpc/{procedure}", json={"json": payload}, timeout=300))


def workbook_login() -> tuple[str, str]:
    workbook = load_workbook(SOURCE, data_only=False, read_only=False)
    sheet = workbook["经营用账户"]
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if "lcj系统登录网站" in str(values[0] or "").strip().lower():
            return str(values[2] or "").strip(), str(values[3] or "").strip()
    raise RuntimeError("LCJ admin credential row not found")


def safe_candidates(preview: dict) -> list[dict]:
    return [{
        "table": row.get("table"),
        "id": row.get("id"),
        "displayName": row.get("displayName"),
        "fields": row.get("fields") or [],
        "reason": row.get("reason"),
    } for row in preview.get("safeCandidates") or []]


def validate_preview(preview: dict) -> list[dict]:
    candidates = safe_candidates(preview)
    if len(candidates) > 20:
        raise RuntimeError(f"recovery candidate count exceeds safety limit: {len(candidates)}")
    for row in candidates:
        table = str(row.get("table"))
        fields = set(str(value) for value in row.get("fields") or [])
        if table not in ALLOWED_FIELDS or not fields or not fields.issubset(ALLOWED_FIELDS[table]):
            raise RuntimeError(f"unsafe candidate shape: {table}:{row.get('id')} fields={sorted(fields)}")
    return candidates


def index_by_id(rows: list[dict]) -> dict[int, dict]:
    return {int(row["id"]): row for row in rows if row.get("id") is not None}


parser = argparse.ArgumentParser()
parser.add_argument("--execute", action="store_true", help="execute the verified recovery after preview")
args = parser.parse_args()

email, password = workbook_login()
session = requests.Session()
login = mutate(session, "auth.login", {"email": email, "password": password})
if not login.get("success") or login.get("user", {}).get("role") != "admin":
    raise RuntimeError("Admin login failed")

preview = query(session, "staff.manualLossRecoveryPreview")
candidates = validate_preview(preview)
result: dict = {
    "checkedAt": datetime.now(timezone.utc).isoformat(),
    "baseUrl": BASE,
    "mode": "execute" if args.execute else "preview-only",
    "authenticatedRole": "admin",
    "preview": {
        "recoveryKey": preview.get("recoveryKey"),
        "context": preview.get("context"),
        "counts": preview.get("counts"),
        "safeCandidates": candidates,
        "ambiguousDifferences": preview.get("ambiguousDifferences") or [],
        "evidencePolicy": preview.get("evidencePolicy") or [],
    },
    "credentialValuesLogged": 0,
    "oldTiDbConnections": 0,
}

if args.execute:
    if not candidates:
        result["execution"] = {"skipped": True, "reason": "no safe candidates"}
    else:
        execution = mutate(session, "staff.manualLossRecoveryExecute", {"confirmation": CONFIRMATION})
        result["execution"] = execution

staff_rows = query(session, "staff.list")
report_rows = query(session, "reportStaff.list")
stores = query(session, "storeManagement.list")
backup = query(session, "databaseBackup.health")
by_table = {
    "staff": index_by_id(staff_rows),
    "report_staff": index_by_id(report_rows),
    "managed_stores": index_by_id(stores),
}
verification = []
for candidate in candidates:
    current = by_table[candidate["table"]].get(int(candidate["id"]))
    verification.append({
        "table": candidate["table"],
        "id": candidate["id"],
        "displayName": candidate["displayName"],
        "fields": candidate["fields"],
        "rowPresent": current is not None,
        "manualRevisionAtPresent": bool(current and current.get("manualRevisionAt")),
        "manualRevisionByPresent": bool(current and current.get("manualRevisionBy") is not None),
    })
result["postState"] = {
    "counts": {"staff": len(staff_rows), "reportStaff": len(report_rows), "stores": len(stores)},
    "candidateRows": verification,
    "backupHealthy": backup.get("healthy"),
    "schedulerStarted": backup.get("schedulerStarted"),
    "latestSuccess": backup.get("latestSuccess"),
}
result["productionWrites"] = len(candidates) if args.execute and not result.get("execution", {}).get("alreadyRecovered") else 0
result["reportSha256"] = hashlib.sha256(json.dumps(result, ensure_ascii=False, sort_keys=True).encode("utf-8")).hexdigest()
OUTPUT.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
print(json.dumps({
    "mode": result["mode"],
    "candidateCounts": result["preview"]["counts"],
    "candidates": [{"table": row["table"], "id": row["id"], "displayName": row["displayName"], "fields": row["fields"]} for row in candidates],
    "ambiguous": len(result["preview"]["ambiguousDifferences"]),
    "execution": result.get("execution"),
    "postState": result["postState"],
    "reportSha256": result["reportSha256"],
}, ensure_ascii=False, indent=2))

if args.execute:
    if any(not row["rowPresent"] or not row["manualRevisionAtPresent"] for row in verification):
        raise SystemExit("recovered row verification failed")
    if backup.get("healthy") is not True:
        raise SystemExit("backup health is not healthy")
