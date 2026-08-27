#!/usr/bin/env python3
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import quote

import requests
from openpyxl import load_workbook

BASE = 'https://lcjmall.com'
SOURCE = Path('/home/ubuntu/upload/pasted_file_06GgiQ_LCJ経営管理表_经营用账户.xlsx')
OUTPUT = Path('/home/ubuntu/lcjgent_hr_persistence_fix/store_manager_production_audit.json')


def trpc_json(response: requests.Response):
    response.raise_for_status()
    payload = response.json()
    if isinstance(payload, list):
        payload = payload[0]
    if 'error' in payload:
        data = payload['error'].get('json', {})
        raise RuntimeError(f"{data.get('code', 'TRPC_ERROR')}: {data.get('message') or 'tRPC error'}")
    return payload['result']['data']['json']


def query(session: requests.Session, procedure: str, payload: dict | None = None):
    encoded = quote(json.dumps({'json': payload or {}}, separators=(',', ':')))
    return trpc_json(session.get(f'{BASE}/api/trpc/{procedure}?input={encoded}', timeout=60))


def mutate(session: requests.Session, procedure: str, payload: dict):
    return trpc_json(session.post(f'{BASE}/api/trpc/{procedure}', json={'json': payload}, timeout=30))


def workbook_login() -> tuple[str, str]:
    workbook = load_workbook(SOURCE, data_only=False, read_only=False)
    sheet = workbook['经营用账户']
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if 'lcj系统登录网站' in str(values[0] or '').strip().lower():
            return str(values[2] or '').strip(), str(values[3] or '').strip()
    raise RuntimeError('LCJ admin credential row not found')


def parse_json(value):
    if value in (None, ''):
        return None
    if isinstance(value, (dict, list)):
        return value
    try:
        return json.loads(value)
    except Exception:
        return {'unparsed': True}


email, password = workbook_login()
session = requests.Session()
login = mutate(session, 'auth.login', {'email': email, 'password': password})
if not login.get('success') or login.get('user', {}).get('role') != 'admin':
    raise RuntimeError('Admin login failed')

stores = query(session, 'storeManagement.list')
health = query(session, 'storeManagement.managementUpgradeHealth')
records = []
for store in stores:
    audit_rows = query(session, 'storeManagement.profileAudit', {'storeId': int(store['id']), 'limit': 200})
    normalized_audit = []
    for row in audit_rows:
        normalized_audit.append({
            'id': row.get('id'),
            'storeId': row.get('storeId'),
            'action': row.get('action'),
            'changedFields': parse_json(row.get('changedFields')),
            'before': parse_json(row.get('beforeJson')),
            'after': parse_json(row.get('afterJson')),
            'actorId': row.get('actorId'),
            'actorName': row.get('actorName'),
            'source': row.get('source'),
            'createdAt': str(row.get('createdAt')) if row.get('createdAt') is not None else None,
        })
    records.append({
        'store': {
            'id': store.get('id'),
            'name': store.get('name'),
            'operatorId': store.get('operatorId'),
            'operatorName': store.get('operatorName'),
            'operator2Id': store.get('operator2Id'),
            'operator2Name': store.get('operator2Name'),
            'contactEmailPresent': bool(store.get('contactEmail')),
            'contactPhonePresent': bool(store.get('contactPhone')),
            'avatarPresent': bool(store.get('avatarUrl') or store.get('avatarKey')),
            'notesPresent': bool(store.get('notes')),
            'updatedAt': str(store.get('updatedAt')) if store.get('updatedAt') is not None else None,
        },
        'audit': normalized_audit,
    })

report = {
    'checkedAt': datetime.now(timezone.utc).isoformat(),
    'baseUrl': BASE,
    'counts': {
        'stores': len(stores),
        'storesWithPrimaryManager': sum(1 for row in stores if row.get('operatorId') or row.get('operatorName')),
        'storesWithSecondaryManager': sum(1 for row in stores if row.get('operator2Id') or row.get('operator2Name')),
        'profileAuditRows': sum(len(item['audit']) for item in records),
    },
    'managementHealth': health,
    'records': records,
    'credentialValuesLogged': 0,
    'productionWrites': 0,
}
OUTPUT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
print(json.dumps({
    'counts': report['counts'],
    'stores': [item['store'] for item in records],
    'auditLatest': [item['audit'][0]['createdAt'] if item['audit'] else None for item in records],
    'healthy': health.get('healthy'),
    'productionWrites': 0,
}, ensure_ascii=False, indent=2))
