#!/usr/bin/env python3
from __future__ import annotations

import json
from collections import Counter
from datetime import datetime, timedelta, timezone
from pathlib import Path

import requests
from openpyxl import load_workbook

BASE_URL = "https://lcjmall.com"
SOURCE = Path("/home/ubuntu/upload/pasted_file_06GgiQ_LCJ経営管理表_经营用账户.xlsx")
OUTPUT = Path(__file__).resolve().parent / "staff_schedule_production_readonly_audit.json"
JST = timezone(timedelta(hours=9))
TODAY = datetime.now(JST).date().isoformat()


def trpc_json(response: requests.Response):
    response.raise_for_status()
    payload = response.json()
    if isinstance(payload, list):
        payload = payload[0]
    if "error" in payload:
        raise RuntimeError(payload["error"].get("json", {}).get("message") or "tRPC error")
    return payload["result"]["data"]["json"]


def workbook_login() -> tuple[str, str]:
    workbook = load_workbook(SOURCE, data_only=False, read_only=False)
    sheet = workbook["经营用账户"]
    for values in sheet.iter_rows(min_row=2, values_only=True):
        purpose = str(values[0] or "").strip().lower()
        if "lcj系统登录网站" in purpose:
            email = str(values[2] or "").strip()
            password = str(values[3] or "").strip()
            if email and password:
                return email, password
    raise RuntimeError("LCJ login credential row not found")


def query(session: requests.Session, procedure: str, value: dict | None = None):
    params = {}
    if value is not None:
        params["input"] = json.dumps({"json": value}, ensure_ascii=False)
    response = session.get(f"{BASE_URL}/api/trpc/{procedure}", params=params, timeout=30)
    return trpc_json(response)


email, password = workbook_login()
session = requests.Session()
login = trpc_json(
    session.post(
        f"{BASE_URL}/api/trpc/auth.login",
        json={"json": {"email": email, "password": password}},
        timeout=30,
    )
)
if not login.get("success"):
    raise RuntimeError("Production login failed")

staff = query(session, "staff.listActive")
schedules = query(
    session,
    "staffSchedule.getByDateRange",
    {"startDate": TODAY, "endDate": f"{TODAY} 23:59:59"},
)

active_ids = {int(row["id"]) for row in staff}
scheduled_active_ids = {int(row["staffId"]) for row in schedules if int(row["staffId"]) in active_ids}
leave_ids = {
    int(row["staffId"])
    for row in schedules
    if int(row["staffId"]) in active_ids and "[请假]" in str(row.get("notes") or "")
}
working_ids = scheduled_active_ids - leave_ids
rest_ids = active_ids - scheduled_active_ids

country_by_id = {int(row["id"]): str(row.get("country") or "未設定") for row in staff}
rest_country_counts = Counter(country_by_id[staff_id] for staff_id in rest_ids)

report = {
    "checkedAt": datetime.now(timezone.utc).isoformat(),
    "baseUrl": BASE_URL,
    "dateJst": TODAY,
    "authenticatedRole": login.get("user", {}).get("role"),
    "activeStaffCount": len(active_ids),
    "savedScheduleRowCount": len(schedules),
    "scheduledActiveStaffCount": len(scheduled_active_ids),
    "workingActiveStaffCount": len(working_ids),
    "leaveActiveStaffCount": len(leave_ids),
    "derivedRestActiveStaffCount": len(rest_ids),
    "derivedRestCountryCounts": dict(sorted(rest_country_counts.items())),
    "syntheticRowsStoredInDatabase": False,
    "productionWrites": 0,
}
report["passed"] = (
    len(active_ids) == len(working_ids) + len(leave_ids) + len(rest_ids)
    and all(int(row["id"]) > 0 for row in schedules)
)
OUTPUT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
print(json.dumps(report, ensure_ascii=False, indent=2))
raise SystemExit(0 if report["passed"] else 1)
