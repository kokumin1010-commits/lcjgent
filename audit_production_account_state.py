#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
from pathlib import Path
from urllib.parse import quote

import requests
from openpyxl import load_workbook

SOURCE = Path('/home/ubuntu/upload/pasted_file_06GgiQ_LCJ経営管理表_经营用账户.xlsx')
OUTPUT = Path('/home/ubuntu/lcjgent_restore/account_import_production_state.json')
BASE_URL = 'https://lcjmall.com'


def fingerprint(value: object) -> str | None:
    text = str(value or '').strip().lower()
    if not text:
        return None
    return hashlib.sha256(text.encode('utf-8')).hexdigest()[:16]


def trpc_json(response: requests.Response):
    response.raise_for_status()
    payload = response.json()
    if isinstance(payload, list):
        payload = payload[0]
    if 'error' in payload:
        raise RuntimeError(payload['error'].get('json', {}).get('message') or 'tRPC error')
    return payload['result']['data']['json']


wb = load_workbook(SOURCE, data_only=False, read_only=False)
ws = wb['经营用账户']
login_email = ''
login_password = ''
for values in ws.iter_rows(min_row=2, values_only=True):
    purpose = str(values[0] or '').strip()
    if 'lcj系统登录网站' in purpose.lower():
        login_email = str(values[2] or '').strip()
        login_password = str(values[3] or '').strip()
        break
if not login_email or not login_password:
    raise RuntimeError('LCJ system login row not found in workbook')

session = requests.Session()
login = session.post(
    f'{BASE_URL}/api/trpc/auth.login',
    json={'json': {'email': login_email, 'password': login_password}},
    timeout=30,
)
login_data = trpc_json(login)

query_input = quote(json.dumps({'json': {}}, separators=(',', ':')))
accounts_response = session.get(
    f'{BASE_URL}/api/trpc/account.listAccounts?input={query_input}', timeout=30
)
contacts_response = session.get(
    f'{BASE_URL}/api/trpc/account.listContacts?input={query_input}', timeout=30
)
permissions_response = session.get(
    f'{BASE_URL}/api/trpc/rbac.myPermissions?input={query_input}', timeout=30
)
accounts = trpc_json(accounts_response)
contacts = trpc_json(contacts_response)
permissions = trpc_json(permissions_response)

contact_candidates = [
    ('Live Commerce Japan株式会社', '日本办公室'),
    ('Hangzhou Shiyao Yuanyu Technology Co., Ltd.', '杭州办公室'),
    ('PJ サプライチェーン株式会社', 'pj仓库'),
    ('株式会社Kyogoku', 'KG仓库'),
]

def normalized(value: object) -> str:
    return ''.join(str(value or '').lower().split())

contact_matches = []
for company_name, contact_name in contact_candidates:
    matches = [
        row for row in contacts
        if normalized(row.get('companyName')) == normalized(company_name)
        or normalized(row.get('contactName')) == normalized(contact_name)
    ]
    contact_matches.append({
        'companyName': company_name,
        'contactName': contact_name,
        'matchCount': len(matches),
        'matchedIds': [row.get('id') for row in matches],
    })

report = {
    'baseUrl': BASE_URL,
    'loginSucceeded': bool(login_data.get('success')),
    'authenticatedRole': login_data.get('user', {}).get('role'),
    'accountCount': len(accounts),
    'contactCount': len(contacts),
    'rbacIsAdmin': bool(permissions.get('isAdmin')),
    'accountManagementPermission': next((item for item in (permissions.get('permissions') or []) if item.get('pageKey') == '/master/account-management'), None),
    'accounts': [
        {
            'id': row.get('id'),
            'platform': row.get('platform'),
            'accountName': row.get('accountName'),
            'accountIdentifierFingerprint': fingerprint(row.get('accountId') or row.get('email') or row.get('phone')),
            'hasPassword': bool(row.get('password')),
            'passwordEnvelope': str(row.get('password') or '').startswith('enc:v1:'),
            'status': row.get('status'),
            'hasRecoveryMarker': 'recovery_source=' in str(row.get('notes') or ''),
        }
        for row in accounts
    ],
    'contactCategories': {
        category: sum(1 for row in contacts if row.get('category') == category)
        for category in ['brand', 'client', 'partner', 'supplier', 'other']
    },
    'candidateContactMatches': contact_matches,
}
OUTPUT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
print(json.dumps({
    'loginSucceeded': report['loginSucceeded'],
    'authenticatedRole': report['authenticatedRole'],
    'accountCount': report['accountCount'],
    'contactCount': report['contactCount'],
    'rbacIsAdmin': report['rbacIsAdmin'],
    'accountManagementPermission': report['accountManagementPermission'],
    'encryptedAccountPasswords': sum(1 for row in report['accounts'] if row['passwordEnvelope']),
    'candidateContactMatches': report['candidateContactMatches'],
    'output': str(OUTPUT),
}, ensure_ascii=False, indent=2))
