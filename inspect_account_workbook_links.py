#!/usr/bin/env python3
from __future__ import annotations

import json
import re
from pathlib import Path
from openpyxl import load_workbook

SOURCE = Path('/home/ubuntu/upload/pasted_file_06GgiQ_LCJ経営管理表_经营用账户.xlsx')
OUTPUT = Path('/home/ubuntu/lcjgent_restore/account_import_link_audit.json')
HYPERLINK_FORMULA = re.compile(r'^=+HYPERLINK\("([^"]+)"(?:,\s*"([^"]*)")?\)$', re.IGNORECASE)

wb_formula = load_workbook(SOURCE, data_only=False, read_only=False)
wb_values = load_workbook(SOURCE, data_only=True, read_only=False)
rows = []
for ws_formula, ws_values in zip(wb_formula.worksheets, wb_values.worksheets):
    for row in range(2, ws_formula.max_row + 1):
        formula_cell = ws_formula.cell(row=row, column=2)
        value_cell = ws_values.cell(row=row, column=2)
        raw = str(formula_cell.value or '').strip()
        cached = str(value_cell.value or '').strip()
        target = ''
        display = cached
        source = 'none'
        if formula_cell.hyperlink and formula_cell.hyperlink.target:
            target = str(formula_cell.hyperlink.target).strip()
            source = 'cell-hyperlink'
        elif raw.lstrip('=').upper().startswith('HYPERLINK('):
            match = HYPERLINK_FORMULA.match(raw)
            if match:
                target = match.group(1).strip()
                display = (match.group(2) or cached or target).strip()
                source = 'formula'
        elif raw.startswith('http://') or raw.startswith('https://'):
            target = raw
            display = raw
            source = 'literal'
        elif cached.startswith('http://') or cached.startswith('https://'):
            target = cached
            display = cached
            source = 'cached-value'
        if raw or cached or target:
            rows.append({
                'row': row,
                'purpose': str(ws_formula.cell(row=row, column=1).value or '').strip(),
                'parent': str(ws_formula.cell(row=row, column=8).value or '').strip(),
                'source': source,
                'target': target,
                'display': display,
                'rawFormula': raw if raw.startswith('=') else '',
            })
OUTPUT.write_text(json.dumps({'source': SOURCE.name, 'rows': rows}, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
print(json.dumps({'rowsWithLinkCells': len(rows), 'resolvedTargets': sum(1 for row in rows if row['target']), 'output': str(OUTPUT)}, ensure_ascii=False))
