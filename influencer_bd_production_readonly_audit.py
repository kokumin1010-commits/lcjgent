#!/usr/bin/env python3
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import quote, urlparse

import requests
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

BASE_URL = "https://lcjmall.com"
SOURCE = Path("/home/ubuntu/upload/pasted_file_06GgiQ_LCJ経営管理表_经营用账户.xlsx")
OUTPUT = Path("/home/ubuntu/lcjgent_influencer_bd/influencer_bd_production_readonly_audit.json")
SCREENSHOT = Path("/home/ubuntu/lcjgent_influencer_bd/influencer_bd_visual_artifacts/influencer_bd_production_empty_state.png")
COMMIT = "19a41c9a"


def trpc_payload(response: requests.Response):
    payload = response.json()
    if isinstance(payload, list):
        payload = payload[0]
    return payload


def trpc_json(response: requests.Response):
    response.raise_for_status()
    payload = trpc_payload(response)
    if "error" in payload:
        data = payload["error"].get("json", {})
        raise RuntimeError(f"{data.get('code', 'TRPC_ERROR')}: {data.get('message') or 'tRPC error'}")
    return payload["result"]["data"]["json"]


def query(session: requests.Session, procedure: str, payload: dict | None = None, timeout: int = 90):
    encoded = quote(json.dumps({"json": payload or {}}, separators=(",", ":")))
    return trpc_json(session.get(f"{BASE_URL}/api/trpc/{procedure}?input={encoded}", timeout=timeout))


def unauthorized_status(procedure: str, payload: dict | None = None):
    encoded = quote(json.dumps({"json": payload or {}}, separators=(",", ":")))
    response = requests.get(f"{BASE_URL}/api/trpc/{procedure}?input={encoded}", timeout=45)
    data = trpc_payload(response)
    error = data.get("error", {}).get("json", {})
    return {"httpStatus": response.status_code, "code": error.get("code"), "message": error.get("message")}


def workbook_login() -> tuple[str, str]:
    workbook = load_workbook(SOURCE, data_only=False, read_only=False)
    sheet = workbook["经营用账户"]
    for values in sheet.iter_rows(min_row=2, values_only=True):
        purpose = str(values[0] or "").strip().lower()
        if "lcj系统登录网站" in purpose:
            email = str(values[2] or "").strip()
            password = str(values[3] or "").strip()
            if email and password:
                return email, password
    raise RuntimeError("LCJ system login row not found")


def normalized_details(value):
    if isinstance(value, dict):
        return value
    if isinstance(value, str) and value:
        try:
            parsed = json.loads(value)
            return parsed if isinstance(parsed, dict) else {}
        except json.JSONDecodeError:
            return {}
    return {}


email, password = workbook_login()
session = requests.Session()
login = trpc_json(session.post(f"{BASE_URL}/api/trpc/auth.login", json={"json": {"email": email, "password": password}}, timeout=45))
if not login.get("success") or login.get("user", {}).get("role") != "admin":
    raise RuntimeError("Production admin login failed")

health = query(session, "influencerBd.health")
backup = query(session, "databaseBackup.health")
bootstrap = query(session, "influencerBd.bootstrap")
dashboard = query(session, "influencerBd.dashboard", {"periodStart": "2026-08-01", "periodEnd": "2026-08-27"})
campaigns = query(session, "influencerBd.listCampaigns", {"includeArchived": False})
creators = query(session, "influencerBd.listCreators", {"includeArchived": False, "limit": 200})
outreach = query(session, "influencerBd.listOutreach", {"periodStart": "2026-08-01", "periodEnd": "2026-08-27", "limit": 500})
analyses = query(session, "influencerBd.listAnalyses", {"limit": 100})
audit_rows = query(session, "influencerBd.audit", {"limit": 100})

run = health.get("recoveryRun") or {}
details = normalized_details(run.get("details"))
upgrade_backups = health.get("backups") or []
backup_reasons = {str(row.get("reason")): str(row.get("status")) for row in upgrade_backups}
source_keys = ["userCount", "staffCount", "brandCount", "brandProductCount", "reportCount", "managedStoreCount"]
before = details.get("before") or {}
after = details.get("after") or {}
source_counts_preserved = bool(before) and all(before.get(key) == after.get(key) for key in source_keys)

unauthorized = {
    procedure: unauthorized_status(procedure, payload)
    for procedure, payload in [
        ("influencerBd.health", {}),
        ("influencerBd.bootstrap", {}),
        ("influencerBd.dashboard", {"periodStart": "2026-08-01", "periodEnd": "2026-08-27"}),
        ("influencerBd.listCreators", {"includeArchived": False, "limit": 20}),
        ("influencerBd.listOutreach", {"periodStart": "2026-08-01", "periodEnd": "2026-08-27", "limit": 20}),
        ("influencerBd.listAnalyses", {"limit": 20}),
        ("influencerBd.audit", {"limit": 20}),
    ]
}
unauthorized_all_blocked = all(
    item.get("httpStatus") in (401, 403)
    and item.get("code") in (-32001, -32003)
    for item in unauthorized.values()
)

cookies = [{
    "name": cookie.name,
    "value": cookie.value,
    "domain": cookie.domain or "lcjmall.com",
    "path": cookie.path or "/",
    "secure": True,
    "httpOnly": False,
    "sameSite": "Lax",
} for cookie in session.cookies]

SCREENSHOT.parent.mkdir(parents=True, exist_ok=True)
with sync_playwright() as playwright:
    browser = playwright.chromium.launch(headless=True, executable_path="/usr/bin/chromium", args=["--no-sandbox", "--disable-dev-shm-usage"])
    context = browser.new_context(viewport={"width": 1600, "height": 1100})
    context.add_cookies(cookies)
    context.add_init_script("localStorage.setItem('language','zh');")
    page = context.new_page()
    console_errors, page_errors, failed_requests, bd_post_requests = [], [], [], []
    page.on("console", lambda message: console_errors.append(message.text) if message.type == "error" else None)
    page.on("pageerror", lambda error: page_errors.append(str(error)))
    page.on("requestfailed", lambda request: failed_requests.append(f"{request.method} {request.url} :: {request.failure}"))
    page.on("request", lambda request: bd_post_requests.append(f"{request.method} {request.url}") if request.method != "GET" and "/api/trpc/influencerBd." in request.url else None)
    response = page.goto(f"{BASE_URL}/master/influencer-bd?verify={COMMIT}", wait_until="networkidle", timeout=120_000)
    page.get_by_role("heading", name="达人BD增长工作台").wait_for(state="visible", timeout=45_000)
    empty_progress_visible = page.get_by_text("所选期间还没有真实BD进度", exact=True).is_visible() if not outreach else True
    empty_creator_visible = True
    if not creators:
        page.get_by_role("tab", name="达人库").click()
        empty_creator_visible = page.get_by_text("尚未登记达人", exact=True).is_visible()
    page.screenshot(path=str(SCREENSHOT), full_page=True)
    final_url = page.url
    http_status = response.status if response else None
    browser.close()

report = {
    "checkedAt": datetime.now(timezone.utc).isoformat(),
    "commit": COMMIT,
    "baseUrl": BASE_URL,
    "authenticatedRole": login.get("user", {}).get("role"),
    "upgrade": {
        "healthy": health.get("healthy"),
        "missingTables": health.get("missingTables"),
        "runStatus": run.get("status"),
        "completedAt": run.get("completedAt"),
        "requiredTableCount": 8,
        "existingBusinessRowsModified": details.get("existingBusinessRowsModified"),
        "sourceCountsPreserved": source_counts_preserved,
        "preBackup": backup_reasons.get("pre-influencer-bd-v1"),
        "postBackup": backup_reasons.get("post-influencer-bd-v1"),
    },
    "backupHealth": {
        "healthy": backup.get("healthy"),
        "schedulerStarted": backup.get("schedulerStarted"),
        "backupRunning": backup.get("backupRunning"),
        "latestFailure": backup.get("latestFailure"),
        "latestSuccessReason": (backup.get("latestSuccess") or {}).get("reason"),
        "ageHours": backup.get("ageHours"),
    },
    "dataCounts": {
        "campaigns": len(campaigns),
        "creators": len(creators),
        "outreach": len(outreach),
        "analyses": len(analyses),
        "audit": len(audit_rows),
        "contactedCreators": (dashboard.get("total") or {}).get("contactedCreators"),
        "contactAttempts": (dashboard.get("total") or {}).get("contactAttempts"),
    },
    "defaultSettings": {
        "autoAnalysisEnabled": (bootstrap.get("settings") or {}).get("autoAnalysisEnabled"),
        "lowReplyRatePercent": (bootstrap.get("settings") or {}).get("lowReplyRatePercent"),
        "stagnationDays": (bootstrap.get("settings") or {}).get("stagnationDays"),
        "minimumContactedCreators": (bootstrap.get("settings") or {}).get("minimumContactedCreators"),
    },
    "unauthorized": unauthorized,
    "unauthorizedAllBlocked": unauthorized_all_blocked,
    "browser": {
        "httpStatus": http_status,
        "finalUrl": final_url,
        "emptyProgressVisible": empty_progress_visible,
        "emptyCreatorVisible": empty_creator_visible,
        "consoleErrors": console_errors,
        "pageErrors": page_errors,
        "failedRequests": failed_requests,
        "bdPostRequests": bd_post_requests,
        "screenshot": str(SCREENSHOT),
    },
    "productionBusinessWrites": 0,
}
report["passed"] = all([
    report["upgrade"]["healthy"] is True,
    report["upgrade"]["missingTables"] == [],
    report["upgrade"]["runStatus"] == "success",
    report["upgrade"]["existingBusinessRowsModified"] == 0,
    report["upgrade"]["sourceCountsPreserved"],
    report["upgrade"]["preBackup"] == "success",
    report["upgrade"]["postBackup"] == "success",
    report["backupHealth"]["healthy"] is True,
    report["backupHealth"]["schedulerStarted"] is True,
    report["backupHealth"]["latestFailure"] is None,
    report["defaultSettings"]["autoAnalysisEnabled"] in (False, 0),
    unauthorized_all_blocked,
    http_status == 200,
    "/login" not in final_url,
    empty_progress_visible,
    empty_creator_visible,
    not console_errors,
    not page_errors,
    not failed_requests,
    not bd_post_requests,
])
OUTPUT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
print(json.dumps(report, ensure_ascii=False, indent=2))
raise SystemExit(0 if report["passed"] else 1)
