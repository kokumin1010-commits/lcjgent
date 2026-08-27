#!/usr/bin/env python3
from __future__ import annotations

import argparse
import base64
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import quote

import requests
from openpyxl import load_workbook

SOURCE = Path('/home/ubuntu/upload/pasted_file_06GgiQ_LCJ経営管理表_经营用账户.xlsx')
OUTPUT = Path('/home/ubuntu/lcjgent_restore/account_workbook_production_import.json')
BASE_URL = 'https://lcjmall.com'
EXPECTED_SHA = '78c837ae232f76fee8061257906b86af3a36afb19a586f3311065c2bfacecb18'
EXPECTED_COUNTS = {'accounts': 22, 'contacts': 4, 'references': 4, 'excluded': 11}


def trpc_json(response: requests.Response):
    response.raise_for_status()
    payload = response.json()
    if isinstance(payload, list):
        payload = payload[0]
    if 'error' in payload:
        data = payload['error'].get('json', {})
        raise RuntimeError(f"{data.get('code', 'TRPC_ERROR')}: {data.get('message') or 'tRPC error'}")
    return payload['result']['data']['json']


def query(session: requests.Session, procedure: str, payload: dict | None = None):
    encoded = quote(json.dumps({'json': payload or {}}, separators=(',', ':')))
    return trpc_json(session.get(f'{BASE_URL}/api/trpc/{procedure}?input={encoded}', timeout=60))


def mutate(session: requests.Session, procedure: str, payload: dict, timeout: int = 60):
    return trpc_json(session.post(f'{BASE_URL}/api/trpc/{procedure}', json={'json': payload}, timeout=timeout))


def workbook_login() -> tuple[str, str]:
    workbook = load_workbook(SOURCE, data_only=False, read_only=False)
    sheet = workbook['经营用账户']
    for values in sheet.iter_rows(min_row=2, values_only=True):
        purpose = str(values[0] or '').strip().lower()
        if 'lcj系统登录网站' in purpose:
            email = str(values[2] or '').strip()
            password = str(values[3] or '').strip()
            if email and password:
                return email, password
    raise RuntimeError('LCJ system login row not found')


def sanitize_state(accounts: list[dict], contacts: list[dict], references: list[dict], imports: list[dict]):
    source_accounts = [row for row in accounts if row.get('sourceFileHash') == EXPECTED_SHA]
    source_contacts = [row for row in contacts if row.get('sourceFileHash') == EXPECTED_SHA]
    source_references = [row for row in references if row.get('sourceFileHash') == EXPECTED_SHA]
    return {
        'accountTotal': len(accounts),
        'contactTotal': len(contacts),
        'referenceTotal': len(references),
        'sourceAccountCount': len(source_accounts),
        'sourceContactCount': len(source_contacts),
        'sourceReferenceCount': len(source_references),
        'sourceAccountEncryptedAtRestCount': sum(1 for row in source_accounts if row.get('passwordEncryptedAtRest')),
        'sourceAccountWithPasswordCount': sum(1 for row in source_accounts if bool(row.get('password'))),
        'sourceKeysUnique': len({row.get('sourceKey') for row in source_accounts}) == len(source_accounts),
        'latestImport': ({
            'fileSha256': imports[0].get('fileSha256'),
            'status': imports[0].get('status'),
            'counts': imports[0].get('counts'),
            'startedAt': imports[0].get('startedAt'),
            'completedAt': imports[0].get('completedAt'),
        } if imports else None),
    }


def fetch_state(session: requests.Session):
    accounts = query(session, 'account.listAccounts')
    contacts = query(session, 'account.listContacts')
    references = query(session, 'account.listReferences')
    imports = query(session, 'account.listWorkbookImports')
    return sanitize_state(accounts, contacts, references, imports)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--execute', action='store_true', help='Perform the production import')
    args = parser.parse_args()

    file_bytes = SOURCE.read_bytes()
    actual_sha = hashlib.sha256(file_bytes).hexdigest()
    if actual_sha != EXPECTED_SHA:
        raise RuntimeError(f'Workbook SHA mismatch: {actual_sha}')
    email, password = workbook_login()
    session = requests.Session()
    login_data = mutate(session, 'auth.login', {'email': email, 'password': password}, timeout=30)
    if not login_data.get('success') or login_data.get('user', {}).get('role') != 'admin':
        raise RuntimeError('Admin login failed')

    file_payload = {
        'fileName': SOURCE.name,
        'fileBase64': base64.b64encode(file_bytes).decode('ascii'),
    }
    preview = mutate(session, 'account.previewWorkbook', file_payload, timeout=60)
    if preview.get('fileSha256') != EXPECTED_SHA or preview.get('counts') != EXPECTED_COUNTS:
        raise RuntimeError(f'Preview mismatch: {preview.get("counts")}')

    pre_state = fetch_state(session)
    report = {
        'checkedAt': datetime.now(timezone.utc).isoformat(),
        'baseUrl': BASE_URL,
        'sourceSha256': actual_sha,
        'previewCounts': preview.get('counts'),
        'previewContainsPasswords': any('password' in row for row in preview.get('accounts', [])),
        'preState': pre_state,
        'executed': bool(args.execute),
        'credentialValuesLogged': 0,
    }
    if not args.execute:
        OUTPUT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
        print(json.dumps(report, ensure_ascii=False, indent=2))
        return

    first = mutate(session, 'account.importWorkbook', {**file_payload, 'confirmSha256': EXPECTED_SHA}, timeout=900)
    post_first = fetch_state(session)
    second = mutate(session, 'account.importWorkbook', {**file_payload, 'confirmSha256': EXPECTED_SHA}, timeout=180)
    post_second = fetch_state(session)
    backup_health = query(session, 'databaseBackup.health')

    report.update({
        'firstImport': {
            'success': first.get('success'),
            'alreadyImported': first.get('alreadyImported'),
            'counts': first.get('counts'),
            'postBackupStatus': first.get('postBackupStatus'),
        },
        'postFirstState': post_first,
        'idempotentSecondImport': {
            'success': second.get('success'),
            'alreadyImported': second.get('alreadyImported'),
            'counts': second.get('counts'),
        },
        'postSecondState': post_second,
        'backupHealth': {
            'healthy': backup_health.get('healthy'),
            'schedulerStarted': backup_health.get('schedulerStarted'),
            'backupRunning': backup_health.get('backupRunning'),
            'retention': backup_health.get('retention'),
            'ageHours': backup_health.get('ageHours'),
            'latestSuccess': backup_health.get('latestSuccess'),
            'latestFailure': backup_health.get('latestFailure'),
        },
    })
    report['passed'] = all([
        first.get('success') is True,
        post_first.get('sourceAccountCount') == 22,
        post_first.get('sourceContactCount') == 4,
        post_first.get('sourceReferenceCount') == 4,
        post_first.get('sourceAccountWithPasswordCount') == 20,
        post_first.get('sourceAccountEncryptedAtRestCount') == post_first.get('sourceAccountWithPasswordCount'),
        post_first.get('sourceKeysUnique') is True,
        second.get('alreadyImported') is True,
        post_second == post_first,
        backup_health.get('healthy') is True,
        backup_health.get('schedulerStarted') is True,
        bool(backup_health.get('latestSuccess')),
        backup_health.get('latestFailure') is None,
    ])
    OUTPUT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    print(json.dumps(report, ensure_ascii=False, indent=2))
    if not report['passed']:
        raise SystemExit(1)


if __name__ == '__main__':
    main()
