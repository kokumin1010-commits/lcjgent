#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path('/home/ubuntu/lcjgent_restore')
SOURCE = Path('/home/ubuntu/upload/pasted_file_06GgiQ_LCJ経営管理表_经营用账户.xlsx')
FILES = [
    ROOT / 'account_import_evidence_2026-08-27.md',
    ROOT / 'account_import_workbook_audit.json',
    ROOT / 'account_workbook_parser_preview.json',
    ROOT / 'account_workbook_import_static_verification.json',
    ROOT / 'account_workbook_runtime_no_db.json',
    ROOT / 'account_management_visual_regression.json',
    ROOT / 'account_management_visual_findings_2026-08-27.md',
    ROOT / 'account_workbook_production_import.json',
    ROOT / 'account_management_production_authenticated_visual.json',
    ROOT / 'account_management_production_visual_findings_2026-08-27.md',
    ROOT / 'account_import_production_http_verification.json',
    ROOT / 'account_import_deployment_status.json',
    ROOT / 'account_import_final_deployment_status.json',
    ROOT / 'account_import_delivery_deployment_status.json',
]

workbook = load_workbook(SOURCE, data_only=False, read_only=False)
sheet = workbook['经营用账户']
preview = json.loads((ROOT / 'account_workbook_parser_preview.json').read_text(encoding='utf-8'))
password_rows = {
    int(row_number)
    for account in preview.get('accounts', [])
    if account.get('hasPassword')
    for row_number in account.get('sourceRows', [])
}
secret_values = {
    str(sheet.cell(row=row_number, column=4).value or '').strip()
    for row_number in password_rows
    if str(sheet.cell(row=row_number, column=4).value or '').strip()
}
secret_metadata = {
    secret: {
        'fingerprint': hashlib.sha256(secret.encode('utf-8')).hexdigest()[:16],
        'length': len(secret),
        'sourceRows': sorted(row_number for row_number in password_rows if str(sheet.cell(row=row_number, column=4).value or '').strip() == secret),
    }
    for secret in secret_values
}
findings = []
for path in FILES:
    if not path.exists():
        continue
    text = path.read_text(encoding='utf-8', errors='ignore')
    matches = [secret_metadata[secret] for secret in secret_values if secret and secret in text]
    if matches:
        findings.append({'file': path.name, 'matchedSecretCount': len(matches), 'matches': matches})

report = {
    'checkedFiles': sum(1 for path in FILES if path.exists()),
    'sourceSecretValueCount': len(secret_values),
    'filesWithSecretValues': findings,
    'passed': not findings,
}
output = ROOT / 'account_evidence_secret_scan.json'
output.write_text(json.dumps(report, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
print(json.dumps(report, ensure_ascii=False, indent=2))
raise SystemExit(0 if report['passed'] else 1)
