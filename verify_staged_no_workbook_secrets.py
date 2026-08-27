from __future__ import annotations

import json
import subprocess
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
SOURCE = Path("/home/ubuntu/upload/pasted_file_06GgiQ_LCJ経営管理表_经营用账户.xlsx")
wb = load_workbook(SOURCE, data_only=False, read_only=False)
ws = wb["经营用账户"]
secrets: list[tuple[int, str]] = []
for row in range(2, ws.max_row + 1):
    account_id = str(ws.cell(row, 3).value or "").strip()
    password = str(ws.cell(row, 4).value or "").strip()
    if account_id and password and len(password) >= 6 and password.lower() not in {"none", "null", "password"}:
        secrets.append((row, password))

staged = subprocess.check_output(["git", "diff", "--cached", "--name-only", "--diff-filter=ACMR"], cwd=ROOT, text=True).splitlines()
findings: list[dict] = []
for relative in staged:
    try:
        content = subprocess.check_output(["git", "show", f":{relative}"], cwd=ROOT)
    except subprocess.CalledProcessError:
        continue
    for row, secret in secrets:
        if secret.encode("utf-8") in content:
            findings.append({"file": relative, "sourceRow": row, "secretLength": len(secret)})

report = {
    "checkedStagedFiles": len(staged),
    "checkedWorkbookSecrets": len(secrets),
    "matches": findings,
    "secretValuesPrinted": 0,
    "passed": not findings,
}
(ROOT / "hr_store_staged_secret_scan.json").write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
print(json.dumps(report, ensure_ascii=False, indent=2))
if findings:
    raise SystemExit(1)
