#!/usr/bin/env python3
from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path

import requests
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

BASE_URL = 'https://lcjmall.com'
SOURCE = Path('/home/ubuntu/upload/pasted_file_06GgiQ_LCJ経営管理表_经营用账户.xlsx')
OUTPUT_DIR = Path('/home/ubuntu/lcjgent_restore/account_management_production_artifacts')
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT = Path('/home/ubuntu/lcjgent_restore/account_management_production_authenticated_visual.json')


def trpc_json(response: requests.Response):
    response.raise_for_status()
    payload = response.json()
    if isinstance(payload, list):
        payload = payload[0]
    if 'error' in payload:
        raise RuntimeError(payload['error'].get('json', {}).get('message') or 'tRPC error')
    return payload['result']['data']['json']


def workbook_login_and_sample_password() -> tuple[str, str, str]:
    workbook = load_workbook(SOURCE, data_only=False, read_only=False)
    sheet = workbook['经营用账户']
    login_email = login_password = sample_password = ''
    for values in sheet.iter_rows(min_row=2, values_only=True):
        purpose = str(values[0] or '').strip().lower()
        value = str(values[3] or '').strip()
        if value and not sample_password:
            sample_password = value
        if 'lcj系统登录网站' in purpose:
            login_email = str(values[2] or '').strip()
            login_password = value
    if not login_email or not login_password or not sample_password:
        raise RuntimeError('Required workbook credential rows not found')
    return login_email, login_password, sample_password


email, password, sample_password = workbook_login_and_sample_password()
session = requests.Session()
login_data = trpc_json(session.post(f'{BASE_URL}/api/trpc/auth.login', json={'json': {'email': email, 'password': password}}, timeout=30))
if not login_data.get('success'):
    raise RuntimeError('Production login failed')

cookies = []
for cookie in session.cookies:
    cookies.append({
        'name': cookie.name,
        'value': cookie.value,
        'domain': cookie.domain or 'lcjmall.com',
        'path': cookie.path or '/',
        'secure': True,
        'httpOnly': False,
        'sameSite': 'Lax',
    })

with sync_playwright() as playwright:
    browser = playwright.chromium.launch(headless=True, executable_path='/usr/bin/chromium', args=['--no-sandbox', '--disable-dev-shm-usage'])
    context = browser.new_context(viewport={'width': 1500, 'height': 1100})
    context.add_cookies(cookies)
    page = context.new_page()
    page.add_init_script("localStorage.setItem('language', 'zh')")
    console_errors, page_errors, failed_requests = [], [], []
    page.on('console', lambda message: console_errors.append(message.text) if message.type == 'error' else None)
    page.on('pageerror', lambda error: page_errors.append(str(error)))
    page.on('requestfailed', lambda request: failed_requests.append(f'{request.method} {request.url} :: {request.failure}'))

    response = page.goto(f'{BASE_URL}/master/account-management?verify-import=4c6b3b0f', wait_until='networkidle', timeout=90_000)
    page.get_by_text('登录凭据与联系人管理', exact=True).wait_for(state='visible', timeout=30_000)
    account_rows = page.locator('table tbody tr')
    account_rows.first.wait_for(state='visible', timeout=30_000)
    account_count = account_rows.count()
    account_text = page.locator('body').inner_text()
    password_not_visible = sample_password not in account_text and password not in account_text
    encrypted_badges = page.get_by_text('数据库已加密', exact=True).count()
    account_screenshot = OUTPUT_DIR / 'account_management_accounts_zh.png'
    page.screenshot(path=str(account_screenshot), full_page=True)

    page.get_by_role('tab', name='品牌・店铺・联系人（CRM）').click()
    page.get_by_text('Live Commerce Japan株式会社', exact=True).wait_for(state='visible', timeout=20_000)
    source_contacts_visible = all(page.get_by_text(name, exact=True).is_visible() for name in [
        'Live Commerce Japan株式会社',
        'Hangzhou Shiyao Yuanyu Technology Co., Ltd.',
        'PJ サプライチェーン株式会社',
        '株式会社Kyogoku',
    ])

    page.get_by_role('tab', name='参考链接').click()
    reference_table = page.get_by_role('table').last
    reference_table.get_by_role('cell', name='LCJ MALL', exact=True).wait_for(state='visible', timeout=20_000)
    reference_count = reference_table.locator('tbody tr').count()
    source_references_visible = all(reference_table.get_by_text(name, exact=True).is_visible() for name in [
        'LCJ MALL', 'オンラインMTG調整リンク', 'LCJシステムユーザー管理', 'Gemini'
    ])
    reference_screenshot = OUTPUT_DIR / 'account_management_references_zh.png'
    page.screenshot(path=str(reference_screenshot), full_page=True)

    report = {
        'checkedAt': datetime.now(timezone.utc).isoformat(),
        'baseUrl': BASE_URL,
        'httpStatus': response.status if response else None,
        'finalUrl': page.url,
        'authenticatedRole': login_data.get('user', {}).get('role'),
        'accountRowCount': account_count,
        'encryptedBadgeCount': encrypted_badges,
        'passwordValuesVisible': not password_not_visible,
        'sourceContactsVisible': source_contacts_visible,
        'referenceRowCount': reference_count,
        'sourceReferencesVisible': source_references_visible,
        'consoleErrors': console_errors,
        'pageErrors': page_errors,
        'failedRequests': failed_requests,
        'productionWrites': 0,
        'screenshots': [str(account_screenshot), str(reference_screenshot)],
    }
    report['passed'] = all([
        response is not None and response.ok,
        '/login' not in page.url,
        account_count == 22,
        encrypted_badges == 20,
        password_not_visible,
        source_contacts_visible,
        reference_count == 4,
        source_references_visible,
        not console_errors,
        not page_errors,
        not failed_requests,
    ])
    OUTPUT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
    print(json.dumps(report, ensure_ascii=False, indent=2))
    browser.close()
    raise SystemExit(0 if report['passed'] else 1)
