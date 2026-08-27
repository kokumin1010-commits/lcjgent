#!/usr/bin/env python3
from __future__ import annotations

import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import quote

import requests
from openpyxl import load_workbook

BASE = 'https://lcjmall.com'
SOURCE = Path('/home/ubuntu/upload/pasted_file_06GgiQ_LCJ経営管理表_经营用账户.xlsx')
OUTPUT = Path('/home/ubuntu/lcjgent_hr_persistence_fix/hr_report_staff_production_audit.json')


def trpc_json(response: requests.Response):
    response.raise_for_status()
    payload = response.json()
    if isinstance(payload, list):
        payload = payload[0]
    if 'error' in payload:
        data = payload['error'].get('json', {})
        raise RuntimeError(f"{data.get('code', 'TRPC_ERROR')}: {data.get('message') or 'tRPC error'}")
    return payload['result']['data']['json']


def query(session: requests.Session, procedure: str, payload: dict | None = None):
    encoded = quote(json.dumps({'json': payload or {}}, separators=(',', ':')))
    return trpc_json(session.get(f'{BASE}/api/trpc/{procedure}?input={encoded}', timeout=60))


def mutate(session: requests.Session, procedure: str, payload: dict):
    return trpc_json(session.post(f'{BASE}/api/trpc/{procedure}', json={'json': payload}, timeout=30))


def workbook_login() -> tuple[str, str]:
    workbook = load_workbook(SOURCE, data_only=False, read_only=False)
    sheet = workbook['经营用账户']
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if 'lcj系统登录网站' in str(values[0] or '').strip().lower():
            return str(values[2] or '').strip(), str(values[3] or '').strip()
    raise RuntimeError('LCJ admin credential row not found')


def timestamp(value):
    return str(value) if value is not None else None


def staff_summary(row: dict):
    return {
        'id': row.get('id'),
        'name': row.get('name'),
        'nameEn': row.get('nameEn'),
        'country': row.get('country'),
        'department': row.get('department'),
        'position': row.get('position'),
        'employmentType': row.get('employmentType'),
        'isActive': row.get('isActive'),
        'archivedAt': timestamp(row.get('archivedAt')),
        'resignDate': timestamp(row.get('resignDate')),
        'evidenceStatus': row.get('evidenceStatus'),
        'evidenceSource': row.get('evidenceSource'),
        'emailEvidenceStatus': row.get('emailEvidenceStatus'),
        'updatedAt': timestamp(row.get('updatedAt')),
    }


def report_summary(row: dict):
    return {
        'id': row.get('id'),
        'name': row.get('name'),
        'nameCn': row.get('nameCn'),
        'nameEn': row.get('nameEn'),
        'country': row.get('country'),
        'linkedStaffId': row.get('linkedStaffId'),
        'isActive': row.get('isActive'),
        'updatedAt': timestamp(row.get('updatedAt')),
    }


email, password = workbook_login()
session = requests.Session()
login = mutate(session, 'auth.login', {'email': email, 'password': password})
if not login.get('success') or login.get('user', {}).get('role') != 'admin':
    raise RuntimeError('Admin login failed')

staff_rows = query(session, 'staff.list')
report_rows = query(session, 'reportStaff.list')
unified_rows = query(session, 'staff.listReportStaffUnified')
archived_rows = query(session, 'staff.listArchivedReportStaffUnified')
directory_health = query(session, 'staff.directoryRecoveryHealth')
archive_health = query(session, 'staff.archiveHealth')
backup_health = query(session, 'databaseBackup.health')

staff_by_id = {int(row['id']): row for row in staff_rows}
report_by_linked: dict[int, list[dict]] = {}
for row in report_rows:
    if row.get('linkedStaffId') is not None:
        report_by_linked.setdefault(int(row['linkedStaffId']), []).append(row)

mismatches = []
for linked_id, rows in report_by_linked.items():
    person = staff_by_id.get(linked_id)
    if not person:
        mismatches.append({'type': 'missing_staff', 'linkedStaffId': linked_id, 'reportStaffIds': [row.get('id') for row in rows]})
        continue
    for row in rows:
        changed = {}
        if str(row.get('name') or '').strip() != str(person.get('name') or '').strip():
            changed['name'] = {'staff': person.get('name'), 'reportStaff': row.get('name')}
        if str(row.get('country') or '').strip() != str(person.get('country') or '').strip():
            changed['country'] = {'staff': person.get('country'), 'reportStaff': row.get('country')}
        if str(row.get('isActive') or '') != str(person.get('isActive') or ''):
            changed['isActive'] = {'staff': person.get('isActive'), 'reportStaff': row.get('isActive')}
        if changed:
            mismatches.append({'type': 'linked_drift', 'staffId': linked_id, 'reportStaffId': row.get('id'), 'fields': changed})

recent_staff = sorted((staff_summary(row) for row in staff_rows), key=lambda row: row.get('updatedAt') or '', reverse=True)[:20]
recent_report = sorted((report_summary(row) for row in report_rows), key=lambda row: row.get('updatedAt') or '', reverse=True)[:20]

report = {
    'checkedAt': datetime.now(timezone.utc).isoformat(),
    'baseUrl': BASE,
    'authenticatedRole': login.get('user', {}).get('role'),
    'counts': {
        'staff': len(staff_rows),
        'staffActive': sum(1 for row in staff_rows if row.get('isActive') == 'active'),
        'staffArchived': sum(1 for row in staff_rows if row.get('archivedAt')),
        'reportStaff': len(report_rows),
        'reportStaffActive': sum(1 for row in report_rows if row.get('isActive') == 'active'),
        'unifiedVisible': len(unified_rows),
        'unifiedArchived': len(archived_rows),
        'linkedStaffIds': len(report_by_linked),
        'linkedDrift': len(mismatches),
    },
    'directoryHealth': directory_health,
    'archiveHealth': archive_health,
    'backupHealth': backup_health,
    'recentStaff': recent_staff,
    'recentReportStaff': recent_report,
    'linkedMismatches': mismatches,
    'credentialValuesLogged': 0,
    'productionWrites': 0,
}
report['snapshotSha256'] = hashlib.sha256(json.dumps(report, ensure_ascii=False, sort_keys=True).encode('utf-8')).hexdigest()
OUTPUT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')
print(json.dumps({
    'counts': report['counts'],
    'directoryHealthy': directory_health.get('healthy'),
    'archiveHealthy': archive_health.get('healthy'),
    'backupHealthy': backup_health.get('healthy'),
    'recentStaffUpdatedAt': [row.get('updatedAt') for row in recent_staff[:5]],
    'recentReportUpdatedAt': [row.get('updatedAt') for row in recent_report[:5]],
    'snapshotSha256': report['snapshotSha256'],
    'productionWrites': 0,
}, ensure_ascii=False, indent=2))
