"""
Admin dashboard API endpoint.
Provides platform-wide statistics for the master dashboard.
Each query is isolated with rollback on failure to prevent cascade errors.
"""
from fastapi import APIRouter, Depends, HTTPException, Header, Query, Request
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from loguru import logger
from typing import Optional, List

from app.core.dependencies import get_db, get_current_user

router = APIRouter(prefix="/admin", tags=["Admin"])

ADMIN_ID = "aither"
ADMIN_PASS = "hub"


def _feedback_order_clause(sort_by: str, sort_order: str) -> str:
    """Build ORDER BY clause for feedback list."""
    allowed = {
        "rated_at": "vp.rated_at DESC NULLS LAST, vp.video_id, vp.phase_index",
        "video_uploaded_at": "v.created_at {dir}, vp.video_id, vp.phase_index",
    }
    direction = "DESC" if sort_order.lower() == "desc" else "ASC"
    template = allowed.get(sort_by, allowed["rated_at"])
    return template.replace("{dir}", direction)


async def _q(db: AsyncSession, sql: str, default=0):
    """Run a scalar query with rollback on failure to keep the session alive."""
    try:
        r = await db.execute(text(sql))
        val = r.scalar()
        return val if val is not None else default
    except Exception as e:
        logger.warning(f"Admin query error: {e}")
        try:
            await db.rollback()
        except Exception as _rb_err:
            logger.debug(f"Rollback cleanup failed: {_rb_err}")
        return default


async def _get_dashboard_data(db: AsyncSession) -> dict:
    """Gather all dashboard statistics."""

    # ── Data Volume ──
    total_videos = await _q(db, "SELECT COUNT(*) FROM videos")
    analyzed_videos = await _q(db, "SELECT COUNT(*) FROM videos WHERE status = 'DONE'")
    pending_videos = total_videos - analyzed_videos

    # time_end is double precision (seconds)
    total_duration_seconds = await _q(db, """
        SELECT COALESCE(SUM(max_sec), 0) FROM (
            SELECT video_id, MAX(COALESCE(time_end, 0)) as max_sec
            FROM video_phases
            WHERE time_end IS NOT NULL
            GROUP BY video_id
        ) sub
    """)
    total_duration_seconds = int(total_duration_seconds)

    # ── Video Types ──
    screen_recording_count = await _q(
        db,
        "SELECT COUNT(*) FROM videos WHERE upload_type = 'screen_recording' OR upload_type IS NULL",
    )
    clean_video_count = await _q(
        db,
        "SELECT COUNT(*) FROM videos WHERE upload_type = 'clean_video'",
    )
    if screen_recording_count == 0 and clean_video_count == 0 and total_videos > 0:
        screen_recording_count = total_videos

    latest_upload_raw = await _q(db, "SELECT MAX(created_at) FROM videos", default=None)
    latest_upload = str(latest_upload_raw) if latest_upload_raw else None

    # ── User Scale ──
    total_users = await _q(db, "SELECT COUNT(*) FROM users WHERE is_active = true")
    if total_users == 0:
        total_users = await _q(db, "SELECT COUNT(*) FROM users")

    total_streamers = await _q(db, "SELECT COUNT(DISTINCT user_id) FROM videos")
    this_month_uploaders = await _q(
        db,
        "SELECT COUNT(DISTINCT user_id) FROM videos "
        "WHERE created_at >= DATE_TRUNC('month', CURRENT_DATE)",
    )

    # Format duration
    total_hours = total_duration_seconds // 3600
    total_minutes = (total_duration_seconds % 3600) // 60

    # ── Daily uploads (past 30 days) ──
    daily_uploads_raw = await db.execute(
        text("""
            SELECT DATE(created_at) as dt, COUNT(*) as cnt,
                   COALESCE(SUM(duration), 0) as total_duration
            FROM videos
            WHERE created_at >= CURRENT_DATE - INTERVAL '30 days'
            GROUP BY DATE(created_at)
            ORDER BY dt ASC
        """)
    )
    daily_uploads_rows = daily_uploads_raw.fetchall()
    daily_uploads = [{"date": str(r.dt), "count": r.cnt, "duration_seconds": int(r.total_duration)} for r in daily_uploads_rows]

    return {
        "data_volume": {
            "total_videos": total_videos,
            "analyzed_videos": analyzed_videos,
            "pending_videos": pending_videos,
            "total_duration_seconds": total_duration_seconds,
            "total_duration_display": f"{total_hours}時間{total_minutes}分",
            "daily_uploads": daily_uploads,
        },
        "video_types": {
            "screen_recording_count": screen_recording_count,
            "clean_video_count": clean_video_count,
            "latest_upload": latest_upload,
        },
        "user_scale": {
            "total_users": total_users,
            "total_streamers": total_streamers,
            "this_month_uploaders": this_month_uploaders,
        },
    }


@router.get("/dashboard")
async def get_dashboard_stats(
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """JWT auth, admin role required."""
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return await _get_dashboard_data(db)


@router.get("/dashboard-public")
async def get_dashboard_stats_public(
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
):
    """Simple ID:password auth via header."""
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")
    return await _get_dashboard_data(db)


@router.get("/feedbacks")
async def get_all_feedbacks(
    include_unrated: bool = False,
    page: int = 1,
    per_page: int = 50,
    filter_rating: int = 0,
    has_clip: Optional[str] = None,
    reviewer_id: Optional[int] = None,
    sort_by: str = "rated_at",
    sort_order: str = "desc",
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
):
    """
    Get phase feedbacks with server-side pagination and filtering.
    - include_unrated: if true, also returns unrated phases
    - page: page number (1-indexed)
    - per_page: items per page (default 50)
    - filter_rating: 0=all, 1-5=specific star, -1=unrated only
    - has_clip: 'yes' = only phases with clips, 'no' = only without clips, None = all
    """
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")

    try:
        # Ensure download log table exists
        await db.execute(text("""
            CREATE TABLE IF NOT EXISTS clip_download_log (
                id UUID PRIMARY KEY,
                video_id UUID NOT NULL,
                phase_index VARCHAR(50),
                time_start FLOAT,
                time_end FLOAT,
                clip_id UUID,
                export_type VARCHAR(20) DEFAULT 'raw',
                created_at TIMESTAMP DEFAULT NOW()
            )
        """))

        # --- Summary query (always counts all) ---
        summary_where = "" if include_unrated else "WHERE vp.user_rating IS NOT NULL"
        summary_sql = text(f"""
            SELECT
                COUNT(*) as total,
                COUNT(vp.user_rating) as rated_count,
                COUNT(*) - COUNT(vp.user_rating) as unrated_count,
                COALESCE(AVG(vp.user_rating), 0) as avg_rating,
                COUNT(CASE WHEN vp.user_rating = 1 THEN 1 END) as r1,
                COUNT(CASE WHEN vp.user_rating = 2 THEN 1 END) as r2,
                COUNT(CASE WHEN vp.user_rating = 3 THEN 1 END) as r3,
                COUNT(CASE WHEN vp.user_rating = 4 THEN 1 END) as r4,
                COUNT(CASE WHEN vp.user_rating = 5 THEN 1 END) as r5,
                COUNT(CASE WHEN vp.user_comment IS NOT NULL AND vp.user_comment != '' THEN 1 END) as with_comments,
                COUNT(CASE WHEN vc_s.id IS NOT NULL THEN 1 END) as with_clip_count
            FROM video_phases vp
            JOIN videos v ON CAST(vp.video_id AS UUID) = v.id
            LEFT JOIN (
                SELECT DISTINCT ON (video_id, phase_index) id, video_id, phase_index
                FROM video_clips
                WHERE clip_url IS NOT NULL
                ORDER BY video_id, phase_index, created_at DESC
            ) vc_s ON CAST(vp.video_id AS VARCHAR) = CAST(vc_s.video_id AS VARCHAR)
                AND vp.phase_index::text = vc_s.phase_index
            {summary_where}
        """)
        summary_result = await db.execute(summary_sql)
        sr = summary_result.fetchone()

        # Download stats (separate lightweight query)
        dl_sql = text("""
            SELECT COUNT(DISTINCT (video_id, phase_index)) as downloaded_clips,
                   COUNT(*) as total_downloads
            FROM clip_download_log
        """)
        dl_result = await db.execute(dl_sql)
        dl = dl_result.fetchone()

        # --- Build WHERE clause for filtered list ---
        conditions = []
        if not include_unrated and filter_rating != -1:
            conditions.append("vp.user_rating IS NOT NULL")
        if filter_rating > 0:
            conditions.append(f"vp.user_rating = {int(filter_rating)}")
        elif filter_rating == -1:
            conditions.append("vp.user_rating IS NULL")

        # reviewer filter
        if reviewer_id is not None:
            conditions.append(f"vp.rated_by_reviewer_id = {int(reviewer_id)}")
        # has_clip filter: requires LEFT JOIN on video_clips
        clip_join_sql = ""
        if has_clip in ("yes", "no"):
            clip_join_sql = """LEFT JOIN (
                SELECT DISTINCT ON (video_id, phase_index) id, video_id, phase_index
                FROM video_clips
                WHERE clip_url IS NOT NULL
                ORDER BY video_id, phase_index, created_at DESC
            ) vc_filter
                ON CAST(vp.video_id AS VARCHAR) = CAST(vc_filter.video_id AS VARCHAR)
                AND vp.phase_index::text = vc_filter.phase_index"""
            if has_clip == "yes":
                conditions.append("vc_filter.id IS NOT NULL")
            else:
                conditions.append("vc_filter.id IS NULL")

        where_clause = ("WHERE " + " AND ".join(conditions)) if conditions else ""

        # Count for pagination
        count_sql = text(f"""
            SELECT COUNT(*) FROM video_phases vp
            JOIN videos v ON CAST(vp.video_id AS UUID) = v.id
            {clip_join_sql}
            {where_clause}
        """)
        count_result = await db.execute(count_sql)
        total_filtered = count_result.scalar()

        # Paginated data query
        offset = (max(1, page) - 1) * per_page
        sql = text(f"""
            SELECT
                vp.video_id,
                vp.phase_index,
                vp.time_start,
                vp.time_end,
                SUBSTRING(vp.phase_description, 1, 200) as phase_description,
                vp.user_rating,
                vp.user_comment,
                vp.rated_at,
                vp.importance_score,
                vp.rated_by_reviewer_id,
                v.original_filename,
                v.user_id,
                v.compressed_blob_url,
                u.email as user_email,
                rv.display_name as reviewer_name,
                COALESCE(dl.download_count, 0) as download_count,
                vc.clip_url as clip_url,
                vc.id as clip_id,
                vc.generation_source as generation_source,
                vc.duration_sec as clip_duration_sec,
                v.created_at as video_uploaded_at
            FROM video_phases vp
            JOIN videos v ON CAST(vp.video_id AS UUID) = v.id
            LEFT JOIN users u ON v.user_id = u.id
            LEFT JOIN users rv ON vp.rated_by_reviewer_id = rv.id
            LEFT JOIN (
                SELECT video_id, phase_index, COUNT(*) AS download_count
                FROM clip_download_log
                GROUP BY video_id, phase_index
            ) dl ON CAST(vp.video_id AS VARCHAR) = CAST(dl.video_id AS VARCHAR)
                AND CAST(vp.phase_index AS VARCHAR) = dl.phase_index
            LEFT JOIN LATERAL (
                SELECT id, clip_url, generation_source, duration_sec
                FROM video_clips
                WHERE CAST(vp.video_id AS VARCHAR) = CAST(video_id AS VARCHAR)
                    AND vp.phase_index::text = phase_index
                    AND clip_url IS NOT NULL
                ORDER BY created_at DESC
                LIMIT 1
            ) vc ON true
            {clip_join_sql}
            {where_clause}
            ORDER BY {_feedback_order_clause(sort_by, sort_order)}
            LIMIT :limit OFFSET :offset
        """)
        result = await db.execute(sql, {"limit": per_page, "offset": offset})
        rows = result.fetchall()

        # Generate SAS URLs for clip playback
        from app.services.storage_service import generate_read_sas_from_url

        # Generate source_url from compressed_blob_url for universal playback
        import os, re
        from azure.storage.blob import generate_blob_sas, BlobSasPermissions
        from datetime import datetime as _dt, timedelta, timezone
        from app.services.storage_service import ACCOUNT_NAME as _ACCT_NAME, CONTAINER_NAME as _CTR_NAME
        _conn_str = os.getenv("AZURE_STORAGE_CONNECTION_STRING", "")
        account_name = _ACCT_NAME or os.getenv("AZURE_STORAGE_ACCOUNT_NAME", "")
        account_key = ""
        for _part in _conn_str.split(";"):
            if _part.startswith("AccountKey="):
                account_key = _part.split("=", 1)[1]
                break
        if not account_key:
            account_key = os.getenv("AZURE_STORAGE_ACCOUNT_KEY", "")
        container_name = _CTR_NAME or os.getenv("AZURE_BLOB_CONTAINER", "videos")
        sas_expiry = _dt.now(timezone.utc) + timedelta(hours=2)

        def _build_source_url(compressed_blob, email, video_id):
            """Build SAS-signed source video URL.
            Priority: compressed_blob_url → original video blob (email/video_id/video_id.mp4)
            """
            if not account_key:
                return None
            try:
                blob_name = None
                if compressed_blob:
                    segments = compressed_blob.split("/")
                    if "@" in segments[0] or len(segments) >= 3:
                        blob_name = compressed_blob
                    else:
                        fname = segments[-1]
                        uuid_match = re.search(r'([0-9A-Fa-f]{8}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{12})', fname)
                        original_case_vid = uuid_match.group(1) if uuid_match else video_id
                        blob_name = f"{email}/{original_case_vid}/{compressed_blob}"
                if not blob_name and email and video_id:
                    # Fallback: use original video blob path
                    blob_name = f"{email}/{video_id}/{video_id}.mp4"
                if not blob_name:
                    return None
                sas = generate_blob_sas(
                    account_name=account_name,
                    container_name=container_name,
                    blob_name=blob_name,
                    account_key=account_key,
                    permission=BlobSasPermissions(read=True),
                    expiry=sas_expiry,
                )
                return f"https://{account_name}.blob.core.windows.net/{container_name}/{blob_name}?{sas}"
            except Exception:
                return None

        # CDN host for faster delivery
        _cdn_host = os.getenv("CDN_HOST", "https://cdn.aitherhub.com")
        _blob_host = f"https://{account_name}.blob.core.windows.net"

        def _to_cdn(url):
            """Replace blob host with CDN host for faster delivery."""
            if url and _blob_host and _cdn_host and _blob_host in url:
                return url.replace(_blob_host, _cdn_host)
            return url

        feedbacks = []
        for r in rows:
            clip_url = r.clip_url
            if clip_url and "blob.core.windows.net" in (clip_url or ""):
                try:
                    clip_url = generate_read_sas_from_url(clip_url, expires_hours=2)
                    clip_url = _to_cdn(clip_url)
                except Exception:
                    pass  # keep original URL
            # Build source_url as fallback when clip_url is not available
            source_url = None
            if not clip_url:
                source_url = _build_source_url(
                    getattr(r, 'compressed_blob_url', None),
                    r.user_email or '',
                    r.video_id,
                )
                source_url = _to_cdn(source_url)
            feedbacks.append({
                "video_id": r.video_id,
                "phase_index": r.phase_index,
                "time_start": r.time_start,
                "time_end": r.time_end,
                "summary": r.phase_description,
                "user_rating": r.user_rating,
                "user_comment": r.user_comment,
                "rated_at": str(r.rated_at) if r.rated_at else None,
                "importance_score": r.importance_score,
                "video_name": r.original_filename,
                "user_id": r.user_id,
                "user_email": r.user_email,
                "download_count": r.download_count,
                "clip_url": clip_url,
                "clip_id": str(r.clip_id) if r.clip_id else None,
                "source_url": source_url,
                "generation_source": r.generation_source,
                "clip_duration_sec": r.clip_duration_sec,
                "rated_by_reviewer_id": r.rated_by_reviewer_id,
                "reviewer_name": r.reviewer_name,
                "video_uploaded_at": str(r.video_uploaded_at) if r.video_uploaded_at else None,
            })

        total_pages = max(1, -(-total_filtered // per_page))  # ceil division

        return {
            "summary": {
                "total_feedbacks": sr.rated_count,
                "total_phases": sr.total,
                "total_unrated": sr.unrated_count,
                "average_rating": round(float(sr.avg_rating), 2),
                "rating_distribution": {
                    1: sr.r1, 2: sr.r2, 3: sr.r3, 4: sr.r4, 5: sr.r5,
                },
                "with_comments": sr.with_comments,
                "with_clip_count": sr.with_clip_count,
                "without_clip_count": sr.total - sr.with_clip_count,
                "downloaded_clips": dl.downloaded_clips if dl else 0,
                "total_downloads": dl.total_downloads if dl else 0,
            },
            "feedbacks": feedbacks,
            "pagination": {
                "page": page,
                "per_page": per_page,
                "total_filtered": total_filtered,
                "total_pages": total_pages,
            },
        }
    except Exception as e:
        logger.exception(f"Failed to fetch feedbacks: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch feedbacks: {e}")


@router.put("/feedbacks/{video_id}/phases/{phase_index}/rating")
async def admin_rate_phase(
    video_id: str,
    phase_index: int,
    request_body: dict,
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
):
    """
    Admin endpoint to rate a phase directly from the admin dashboard.
    Body: { "rating": 1-5, "comment": "optional text" }
    """
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")

    try:
        rating = request_body.get("rating")
        comment = request_body.get("comment", "")
        reviewer_id = request_body.get("reviewer_id")  # optional: set by frontend when reviewer is logged in

        if rating is None or not isinstance(rating, int) or rating < 1 or rating > 5:
            raise HTTPException(status_code=400, detail="rating must be an integer between 1 and 5")

        importance_score = (rating - 1) / 4.0  # 1->0.0, 2->0.25, 3->0.5, 4->0.75, 5->1.0

        sql_update = text("""
            UPDATE video_phases
            SET user_rating = :rating,
                user_comment = :comment,
                importance_score = :importance_score,
                rated_at = NOW(),
                updated_at = NOW(),
                rated_by_reviewer_id = COALESCE(:reviewer_id, rated_by_reviewer_id)
            WHERE video_id = :video_id AND phase_index = :phase_index
        """)
        result = await db.execute(sql_update, {
            "rating": rating,
            "comment": comment,
            "importance_score": importance_score,
            "video_id": video_id,
            "phase_index": phase_index,
            "reviewer_id": reviewer_id,
        })
        await db.commit()

        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="Phase not found")

        # If reviewer_id provided, update their session clips_reviewed count
        if reviewer_id:
            await db.execute(text("""
                UPDATE review_sessions
                SET clips_reviewed = clips_reviewed + 1,
                    last_heartbeat = NOW()
                WHERE reviewer_id = :reviewer_id AND ended_at IS NULL
            """), {"reviewer_id": reviewer_id})
            await db.commit()

        logger.info(f"[Admin] Phase rated: video={video_id}, phase={phase_index}, rating={rating}, reviewer_id={reviewer_id}")

        return {
            "success": True,
            "video_id": video_id,
            "phase_index": phase_index,
            "rating": rating,
            "comment": comment,
            "importance_score": importance_score,
            "reviewer_id": reviewer_id,
        }
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        logger.exception(f"Failed to rate phase: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to rate phase: {e}")


@router.get("/stuck-videos")
async def get_stuck_videos(
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
):
    """List videos that are stuck in processing (not DONE/ERROR, older than 30 min)."""
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")

    try:
        sql = text("""
            SELECT v.id, v.original_filename, v.status, v.step_progress,
                   v.upload_type, v.created_at, v.updated_at,
                   u.email as user_email
            FROM videos v
            LEFT JOIN users u ON v.user_id = u.id
            WHERE v.status NOT IN ('DONE', 'ERROR')
            ORDER BY v.created_at DESC
            LIMIT 50
        """)
        result = await db.execute(sql)
        rows = result.fetchall()

        videos = []
        for r in rows:
            videos.append({
                "id": str(r.id),
                "filename": r.original_filename,
                "status": r.status,
                "step_progress": r.step_progress,
                "upload_type": r.upload_type,
                "created_at": str(r.created_at) if r.created_at else None,
                "updated_at": str(r.updated_at) if r.updated_at else None,
                "user_email": r.user_email,
            })

        return {"count": len(videos), "videos": videos}
    except Exception as e:
        logger.exception(f"Failed to fetch stuck videos: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ──────────────────────────────────────────────────────────────────────
# Video Processing / Learning Log endpoints
# ──────────────────────────────────────────────────────────────────────

@router.get("/videos")
async def get_video_list(
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
    status_filter: Optional[str] = None,
    upload_type_filter: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
):
    """List all videos with processing status, phase count, sales_moment count,
    human label stats, and dataset inclusion status."""
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")

    try:
        # Build WHERE clause
        conditions = []
        params = {"lim": limit, "off": offset}
        if status_filter:
            conditions.append("v.status = :sf")
            params["sf"] = status_filter
        if upload_type_filter:
            conditions.append("v.upload_type = :uf")
            params["uf"] = upload_type_filter
        where_clause = ("WHERE " + " AND ".join(conditions)) if conditions else ""

        sql = text(f"""
            SELECT
                v.id,
                v.original_filename,
                v.upload_type,
                v.status,
                v.step_progress,
                v.created_at,
                v.updated_at,
                u.email AS user_email,
                COALESCE(ph.phase_count, 0) AS phase_count,
                COALESCE(sm.moment_count, 0) AS moment_count,
                COALESCE(sm.csv_moment_count, 0) AS csv_moment_count,
                COALESCE(sm.screen_moment_count, 0) AS screen_moment_count,
                COALESCE(hl.rating_count, 0) AS rating_count,
                COALESCE(hl.tag_count, 0) AS tag_count,
                COALESCE(hl.comment_count, 0) AS comment_count,
                vps.frames_extracted,
                vps.audio_extracted,
                vps.speech_done,
                vps.vision_done
            FROM videos v
            LEFT JOIN users u ON v.user_id = u.id
            LEFT JOIN (
                SELECT video_id, COUNT(*) AS phase_count
                FROM video_phases
                GROUP BY video_id
            ) ph ON CAST(ph.video_id AS UUID) = v.id
            LEFT JOIN (
                SELECT video_id,
                       COUNT(*) AS moment_count,
                       COUNT(CASE WHEN source = 'csv' THEN 1 END) AS csv_moment_count,
                       COUNT(CASE WHEN source = 'screen' THEN 1 END) AS screen_moment_count
                FROM video_sales_moments
                GROUP BY video_id
            ) sm ON CAST(sm.video_id AS UUID) = v.id
            LEFT JOIN (
                SELECT video_id,
                       COUNT(CASE WHEN user_rating IS NOT NULL THEN 1 END) AS rating_count,
                       COUNT(CASE WHEN human_sales_tags IS NOT NULL AND human_sales_tags != '[]' THEN 1 END) AS tag_count,
                       COUNT(CASE WHEN user_comment IS NOT NULL AND user_comment != '' THEN 1 END) AS comment_count
                FROM video_phases
                GROUP BY video_id
            ) hl ON CAST(hl.video_id AS UUID) = v.id
            LEFT JOIN video_processing_state vps ON CAST(vps.video_id AS UUID) = v.id
            {where_clause}
            ORDER BY v.created_at DESC
            LIMIT :lim OFFSET :off
        """)

        result = await db.execute(sql, params)
        rows = result.fetchall()

        # Total count
        count_sql = text(f"SELECT COUNT(*) FROM videos v {where_clause}")
        total = (await db.execute(count_sql, params)).scalar() or 0

        videos = []
        for r in rows:
            # Determine dataset status
            ds_status = "excluded"
            ds_reason = None
            if r.status == "DONE":
                if r.moment_count > 0:
                    ds_status = "included"
                else:
                    ds_status = "excluded"
                    ds_reason = "no_sales_moments"
            elif r.status == "ERROR":
                ds_status = "excluded"
                ds_reason = "processing_error"
            else:
                ds_status = "pending"
                ds_reason = "still_processing"

            videos.append({
                "id": str(r.id),
                "filename": r.original_filename,
                "upload_type": r.upload_type or "screen_recording",
                "status": r.status,
                "step_progress": r.step_progress,
                "created_at": str(r.created_at) if r.created_at else None,
                "updated_at": str(r.updated_at) if r.updated_at else None,
                "user_email": r.user_email,
                "phase_count": r.phase_count,
                "moment_count": r.moment_count,
                "moment_sources": {
                    "csv": r.csv_moment_count,
                    "screen": r.screen_moment_count,
                } if r.moment_count > 0 else None,
                "rating_count": r.rating_count,
                "tag_count": r.tag_count,
                "comment_count": r.comment_count,
                "dataset_status": ds_status,
                "dataset_excluded_reason": ds_reason,
                "processing_state": {
                    "frames_extracted": r.frames_extracted if r.frames_extracted is not None else False,
                    "audio_extracted": r.audio_extracted if r.audio_extracted is not None else False,
                    "speech_done": r.speech_done if r.speech_done is not None else False,
                    "vision_done": r.vision_done if r.vision_done is not None else False,
                } if r.frames_extracted is not None else None,
            })

        return {
            "total": total,
            "limit": limit,
            "offset": offset,
            "videos": videos,
        }
    except Exception as e:
        logger.exception(f"Failed to fetch video list: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/videos/{video_id}")
async def get_video_detail(
    video_id: str,
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
):
    """Get detailed processing and learning log for a specific video."""
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")

    try:
        # ── A. Basic info ──
        video_sql = text("""
            SELECT v.id, v.original_filename, v.upload_type, v.status,
                   v.step_progress, v.created_at, v.updated_at,
                   v.excel_product_blob_url, v.excel_trend_blob_url,
                   v.compressed_blob_url, v.top_products,
                   v.time_offset_seconds,
                   v.queue_enqueued_at, v.worker_claimed_at,
                   v.worker_instance_id, v.dequeue_count,
                   v.enqueue_status, v.enqueue_error,
                   u.email AS user_email
            FROM videos v
            LEFT JOIN users u ON v.user_id = u.id
            WHERE v.id = :vid
        """)
        result = await db.execute(video_sql, {"vid": video_id})
        video = result.fetchone()
        if not video:
            raise HTTPException(status_code=404, detail="Video not found")

        # ── B. Processing state ──
        state_sql = text("""
            SELECT * FROM video_processing_state WHERE video_id = :vid
        """)
        state_result = await db.execute(state_sql, {"vid": video_id})
        state = state_result.fetchone()

        # ── C. Phases summary ──
        phases_sql = text("""
            SELECT
                COUNT(*) AS total_phases,
                COUNT(CASE WHEN user_rating IS NOT NULL THEN 1 END) AS rated_phases,
                COUNT(CASE WHEN human_sales_tags IS NOT NULL AND human_sales_tags != '[]' THEN 1 END) AS tagged_phases,
                COUNT(CASE WHEN user_comment IS NOT NULL AND user_comment != '' THEN 1 END) AS commented_phases,
                AVG(user_rating) AS avg_rating,
                MIN(time_start) AS min_time,
                MAX(time_end) AS max_time
            FROM video_phases
            WHERE video_id = :vid
        """)
        phases_result = await db.execute(phases_sql, {"vid": video_id})
        phases_summary = phases_result.fetchone()

        # ── D. Sales moments breakdown ──
        # Try with source column first, fallback without it
        try:
            moments_sql = text("""
                SELECT
                    COALESCE(source, 'csv') AS source,
                    moment_type,
                    moment_type_detail,
                    COUNT(*) AS count,
                    AVG(confidence) AS avg_confidence
                FROM video_sales_moments
                WHERE video_id = :vid
                GROUP BY source, moment_type, moment_type_detail
                ORDER BY source, count DESC
            """)
            moments_result = await db.execute(moments_sql, {"vid": video_id})
            moments_rows = moments_result.fetchall()

            moments_total = sum(r.count for r in moments_rows)
            moments_by_source = {}
            for r in moments_rows:
                src = r.source or "csv"
                if src not in moments_by_source:
                    moments_by_source[src] = []
                moments_by_source[src].append({
                    "moment_type": r.moment_type,
                    "moment_type_detail": r.moment_type_detail,
                    "count": r.count,
                    "avg_confidence": round(float(r.avg_confidence), 3) if r.avg_confidence else None,
                })
        except Exception:
            await db.rollback()
            # Fallback: source/moment_type_detail/confidence columns not yet migrated
            moments_sql = text("""
                SELECT
                    moment_type,
                    COUNT(*) AS count
                FROM video_sales_moments
                WHERE video_id = :vid
                GROUP BY moment_type
                ORDER BY count DESC
            """)
            moments_result = await db.execute(moments_sql, {"vid": video_id})
            moments_rows = moments_result.fetchall()
            moments_total = sum(r.count for r in moments_rows)
            moments_by_source = {"csv": [
                {"moment_type": r.moment_type, "moment_type_detail": None, "count": r.count, "avg_confidence": None}
                for r in moments_rows
            ]}

        # ── E. Reports check ──
        try:
            reports_sql = text("""
                SELECT COUNT(*) AS report_count
                FROM reports
                WHERE video_id = :vid
            """)
            reports_result = await db.execute(reports_sql, {"vid": video_id})
            report_count = reports_result.scalar() or 0
        except Exception:
            await db.rollback()
            report_count = 0  # table may not exist

        # ── F. Transcript check ──
        transcript_sql = text("""
            SELECT COUNT(*) AS segment_count
            FROM video_speech_segments
            WHERE video_id = :vid
        """)
        try:
            transcript_result = await db.execute(transcript_sql, {"vid": video_id})
            transcript_count = transcript_result.scalar() or 0
        except Exception:
            await db.rollback()
            transcript_count = -1  # table may not exist

        # ── G. Build pipeline steps status ──
        # Derive step completion from video status
        status = video.status or ""
        step_order = [
            "STEP_0_EXTRACT_FRAMES",
            "STEP_1_DETECT_PHASES",
            "STEP_2_EXTRACT_METRICS",
            "STEP_3_TRANSCRIBE_AUDIO",
            "STEP_4_IMAGE_CAPTION",
            "STEP_5_BUILD_PHASE_UNITS",
            "STEP_6_BUILD_PHASE_DESCRIPTION",
            "STEP_7_GROUPING",
            "STEP_8_UPDATE_BEST_PHASE",
            "STEP_9_BUILD_VIDEO_STRUCTURE_FEATURES",
            "STEP_10_ASSIGN_VIDEO_STRUCTURE_GROUP",
            "STEP_11_UPDATE_VIDEO_STRUCTURE_GROUP_STATS",
            "STEP_12_UPDATE_VIDEO_STRUCTURE_BEST",
            "STEP_12_5_PRODUCT_DETECTION",
            "STEP_13_BUILD_REPORTS",
            "STEP_14_FINALIZE",
        ]

        step_labels = {
            "STEP_0_EXTRACT_FRAMES": "フレーム抽出",
            "STEP_1_DETECT_PHASES": "フェーズ検出",
            "STEP_2_EXTRACT_METRICS": "メトリクス抽出",
            "STEP_3_TRANSCRIBE_AUDIO": "音声文字起こし",
            "STEP_4_IMAGE_CAPTION": "画像キャプション",
            "STEP_5_BUILD_PHASE_UNITS": "フェーズ構築 (CSV/Screen統合含む)",
            "STEP_6_BUILD_PHASE_DESCRIPTION": "AI要約生成",
            "STEP_7_GROUPING": "グルーピング",
            "STEP_8_UPDATE_BEST_PHASE": "ベストフェーズ選定",
            "STEP_9_BUILD_VIDEO_STRUCTURE_FEATURES": "動画構造特徴量",
            "STEP_10_ASSIGN_VIDEO_STRUCTURE_GROUP": "構造グループ割当",
            "STEP_11_UPDATE_VIDEO_STRUCTURE_GROUP_STATS": "グループ統計更新",
            "STEP_12_UPDATE_VIDEO_STRUCTURE_BEST": "構造ベスト更新",
            "STEP_12_5_PRODUCT_DETECTION": "商品検出",
            "STEP_13_BUILD_REPORTS": "レポート生成",
            "STEP_14_FINALIZE": "最終処理",
        }

        if status == "DONE":
            current_step_idx = len(step_order)  # all done
        elif status == "ERROR":
            # Find last known step from status pattern
            current_step_idx = -1  # unknown
        elif status in step_order:
            current_step_idx = step_order.index(status)
        else:
            current_step_idx = -1

        pipeline_steps = []
        for i, step_name in enumerate(step_order):
            if status == "DONE":
                step_status = "success"
            elif status == "ERROR" and current_step_idx == -1:
                # Can't determine which step failed
                step_status = "unknown"
            elif i < current_step_idx:
                step_status = "success"
            elif i == current_step_idx:
                step_status = "running" if status not in ("DONE", "ERROR") else "failed"
            else:
                step_status = "pending"

            pipeline_steps.append({
                "step_name": step_name,
                "label": step_labels.get(step_name, step_name),
                "status": step_status,
            })

        # ── H. Duration ──
        duration_sec = None
        if phases_summary and phases_summary.max_time:
            duration_sec = round(float(phases_summary.max_time), 1)

        # ── I. Dataset status ──
        ds_status = "excluded"
        ds_reason = None
        if status == "DONE":
            if moments_total > 0:
                ds_status = "included"
            else:
                ds_status = "excluded"
                ds_reason = "no_sales_moments"
        elif status == "ERROR":
            ds_status = "excluded"
            ds_reason = "processing_error"
        else:
            ds_status = "pending"
            ds_reason = "still_processing"

        return {
            "basic_info": {
                "video_id": str(video.id),
                "filename": video.original_filename,
                "upload_type": video.upload_type or "screen_recording",
                "status": video.status,
                "step_progress": video.step_progress,
                "duration_sec": duration_sec,
                "created_at": str(video.created_at) if video.created_at else None,
                "updated_at": str(video.updated_at) if video.updated_at else None,
                "user_email": video.user_email,
                "has_excel_product": bool(video.excel_product_blob_url),
                "has_excel_trend": bool(video.excel_trend_blob_url),
                "has_compressed": bool(video.compressed_blob_url),
                "compressed_blob_url": video.compressed_blob_url,
                "top_products": video.top_products,
                "time_offset_seconds": video.time_offset_seconds,
            },
            "queue_info": {
                "enqueued_at": str(video.queue_enqueued_at) if video.queue_enqueued_at else None,
                "worker_claimed_at": str(video.worker_claimed_at) if video.worker_claimed_at else None,
                "worker_instance_id": video.worker_instance_id,
                "dequeue_count": video.dequeue_count,
                "enqueue_status": video.enqueue_status,
                "enqueue_error": video.enqueue_error,
            },
            "processing_state": {
                "frames_extracted": state.frames_extracted if state else None,
                "audio_extracted": state.audio_extracted if state else None,
                "speech_done": state.speech_done if state else None,
                "vision_done": state.vision_done if state else None,
                "updated_at": str(state.updated_at) if state and state.updated_at else None,
            } if state else None,
            "pipeline_steps": pipeline_steps,
            "phases": {
                "total": phases_summary.total_phases if phases_summary else 0,
                "duration_sec": duration_sec,
                "rated": phases_summary.rated_phases if phases_summary else 0,
                "tagged": phases_summary.tagged_phases if phases_summary else 0,
                "commented": phases_summary.commented_phases if phases_summary else 0,
                "avg_rating": round(float(phases_summary.avg_rating), 2) if phases_summary and phases_summary.avg_rating else None,
                "reviewers": None,  # requires reviewer_name column migration
            },
            "sales_moments": {
                "total": moments_total,
                "by_source": moments_by_source,
            },
            "reports": {
                "count": report_count,
            },
            "transcript": {
                "segment_count": transcript_count,
            },
            "human_labels": {
                "rated_phases": phases_summary.rated_phases if phases_summary else 0,
                "tagged_phases": phases_summary.tagged_phases if phases_summary else 0,
                "commented_phases": phases_summary.commented_phases if phases_summary else 0,
                "avg_rating": round(float(phases_summary.avg_rating), 2) if phases_summary and phases_summary.avg_rating else None,
                "reviewers": None,  # requires reviewer_name column migration
            },
            "dataset": {
                "status": ds_status,
                "excluded_reason": ds_reason,
            },
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Failed to fetch video detail: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/retry-video/{video_id}")
async def retry_video(
    video_id: str,
    from_step: Optional[str] = Query(None, description="Resume from specific step, e.g. STEP_12_5_PRODUCT_DETECTION"),
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None),
):
    """Re-enqueue a stuck video for processing.
    Supports both standard videos and LiveBoost (live_boost) videos.
    For standard videos: generates a fresh SAS URL and pushes a new job.
    For LiveBoost videos: creates/resets a LiveAnalysisJob and enqueues live_analysis.
    Optional from_step: resume from a specific pipeline step instead of STEP_0."""
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        # Get video info (including upload_type and excel URLs)
        sql = text("""
            SELECT v.id, v.original_filename, v.status, v.user_id,
                   v.upload_type, u.email as user_email,
                   v.excel_product_blob_url, v.excel_trend_blob_url
            FROM videos v
            LEFT JOIN users u ON v.user_id = u.id
            WHERE v.id = :vid
        """)
        result = await db.execute(sql, {"vid": video_id})
        row = result.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Video not found")

        upload_type = row.upload_type or ""

        # ── LiveBoost (live_boost) videos: use live_analysis pipeline ──
        if upload_type == "live_boost":
            return await _retry_live_boost_admin(
                db=db,
                video_id=str(row.id),
                user_id=row.user_id,
                user_email=row.user_email,
            )

        # ── Standard videos: use standard pipeline ──
        # Generate fresh SAS URL for video blob
        from app.services.storage_service import generate_download_sas, generate_read_sas_from_url
        download_url, expiry = await generate_download_sas(
            email=row.user_email,
            video_id=str(row.id),
            filename=row.original_filename,
            expires_in_minutes=1440,  # 24 hours
        )

        # ★ Refresh SAS tokens for Excel blob URLs (they expire after upload)
        _excel_updates = {}
        if row.excel_product_blob_url:
            fresh_product_url = generate_read_sas_from_url(row.excel_product_blob_url, expires_hours=24)
            if fresh_product_url:
                _excel_updates["excel_product_blob_url"] = fresh_product_url
                logger.info("[retry-video] Refreshed SAS for excel_product_blob_url")
        if row.excel_trend_blob_url:
            fresh_trend_url = generate_read_sas_from_url(row.excel_trend_blob_url, expires_hours=24)
            if fresh_trend_url:
                _excel_updates["excel_trend_blob_url"] = fresh_trend_url
                logger.info("[retry-video] Refreshed SAS for excel_trend_blob_url")
        if _excel_updates:
            set_clauses = ", ".join(f"{k} = :{k}" for k in _excel_updates)
            _excel_updates["vid"] = video_id
            await db.execute(
                text(f"UPDATE videos SET {set_clauses} WHERE id = :vid"),
                _excel_updates,
            )
            logger.info("[retry-video] Updated %d Excel blob URLs with fresh SAS", len(_excel_updates) - 1)

        # Reset status — use from_step if provided, otherwise STEP_0
        resume_status = 'STEP_0_EXTRACT_FRAMES'
        if from_step:
            # Validate from_step is a known step status
            valid_steps = [
                'STEP_0_EXTRACT_FRAMES', 'STEP_1_DETECT_PHASES',
                'STEP_2_EXTRACT_METRICS', 'STEP_3_TRANSCRIBE_AUDIO',
                'STEP_4_IMAGE_CAPTION', 'STEP_5_BUILD_PHASE_UNITS',
                'STEP_6_BUILD_PHASE_DESCRIPTION', 'STEP_7_GROUPING',
                'STEP_8_UPDATE_BEST_PHASE', 'STEP_9_BUILD_VIDEO_STRUCTURE_FEATURES',
                'STEP_10_ASSIGN_VIDEO_STRUCTURE_GROUP',
                'STEP_11_UPDATE_VIDEO_STRUCTURE_GROUP_STATS',
                'STEP_12_UPDATE_VIDEO_STRUCTURE_BEST',
                'STEP_12_5_PRODUCT_DETECTION',
                'STEP_13_BUILD_REPORTS', 'STEP_14_FINALIZE',
            ]
            if from_step in valid_steps:
                resume_status = from_step
            else:
                raise HTTPException(status_code=400, detail=f"Invalid from_step: {from_step}. Valid: {valid_steps}")
        await db.execute(
            text("UPDATE videos SET status = :status, step_progress = 0, worker_claimed_at = NULL, dequeue_count = 0, updated_at = NOW() WHERE id = :vid"),
            {"vid": video_id, "status": resume_status},
        )
        await db.commit()

        # Enqueue job
        from app.services.queue_service import enqueue_job
        await enqueue_job({
            "video_id": str(row.id),
            "blob_url": download_url,
            "original_filename": row.original_filename,
        })

        return {
            "status": "ok",
            "video_id": video_id,
            "message": f"Re-enqueued with fresh SAS URL (expires {expiry}), resume_from={resume_status}",
            "resume_from": resume_status,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Failed to retry video: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/batch-cleanup-zombies")
async def batch_cleanup_zombies(
    max_age_hours: int = 6,
    dry_run: bool = False,
    db: AsyncSession = Depends(get_db),
):
    """
    Batch-reset zombie videos stuck at STEP_0 with 0% progress for >max_age_hours.
    These videos clog the worker queue and prevent new videos from being processed.
    
    Args:
        max_age_hours: Minimum hours since last update to consider zombie (default: 6)
        dry_run: If True, only list zombies without resetting them
    """
    from datetime import datetime, timedelta
    try:
        zombie_threshold = datetime.utcnow() - timedelta(hours=max_age_hours)
        
        result = await db.execute(
            text("""
                SELECT v.id, v.original_filename, v.status, v.step_progress,
                       v.updated_at, v.dequeue_count, v.worker_claimed_at,
                       u.email as user_email
                FROM videos v
                LEFT JOIN users u ON v.user_id = u.id
                WHERE v.status = 'STEP_0_EXTRACT_FRAMES'
                  AND COALESCE(v.step_progress, 0) = 0
                  AND v.updated_at < :threshold
                ORDER BY v.updated_at ASC
                LIMIT 100
            """),
            {"threshold": zombie_threshold},
        )
        zombies = result.fetchall()
        
        if dry_run:
            return {
                "dry_run": True,
                "zombie_count": len(zombies),
                "zombies": [
                    {
                        "id": str(z.id),
                        "filename": z.original_filename,
                        "progress": z.step_progress or 0,
                        "updated_at": str(z.updated_at),
                        "dequeue_count": z.dequeue_count or 0,
                        "user_email": z.user_email,
                    }
                    for z in zombies
                ],
            }
        
        cleaned = 0
        for z in zombies:
            vid = str(z.id)
            try:
                _upd = await db.execute(
                    text("""
                        UPDATE videos
                        SET status = 'ERROR',
                            error_message = :err_msg,
                            last_error_code = 'ZOMBIE_CLEANUP',
                            last_error_message = :err_msg,
                            updated_at = NOW()
                        WHERE id = :vid
                          AND status = 'STEP_0_EXTRACT_FRAMES'
                    """),
                    {
                        "vid": vid,
                        "err_msg": f"Admin zombie cleanup: stuck at STEP_0 with 0% progress "
                                   f"for >{max_age_hours}h. Use retry to re-process.",
                    },
                )
                await db.commit()
                if _upd.rowcount > 0:
                    cleaned += 1
                    # Record error log
                    try:
                        await db.execute(
                            text("""
                                INSERT INTO video_error_logs
                                    (video_id, error_code, error_step, error_message, source)
                                VALUES
                                    (:vid, 'ZOMBIE_CLEANUP', 'STEP_0_EXTRACT_FRAMES',
                                     :msg, 'admin')
                            """),
                            {
                                "vid": vid,
                                "msg": f"Admin zombie cleanup: {z.original_filename} "
                                       f"stuck for >{max_age_hours}h",
                            },
                        )
                        await db.commit()
                    except Exception:
                        try:
                            await db.rollback()
                        except Exception:
                            pass
            except Exception as e:
                logger.warning(f"Failed to cleanup zombie {vid}: {e}")
                try:
                    await db.rollback()
                except Exception:
                    pass
        
        return {
            "dry_run": False,
            "zombie_count": len(zombies),
            "cleaned": cleaned,
            "message": f"Cleaned {cleaned}/{len(zombies)} zombie videos",
        }
    except Exception as e:
        logger.exception(f"Batch zombie cleanup failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


async def _retry_live_boost_admin(
    db: AsyncSession,
    video_id: str,
    user_id: int,
    user_email: str,
) -> dict:
    """Admin retry for LiveBoost videos — creates/resets LiveAnalysisJob and enqueues."""
    import uuid as uuid_module
    from datetime import datetime, timezone
    from sqlalchemy import select as sa_select, update as sa_update

    try:
        from app.models.orm.live_analysis_job import LiveAnalysisJob
        from app.services.queue_service import enqueue_job

        # Check for existing job
        result = await db.execute(
            sa_select(LiveAnalysisJob).where(
                LiveAnalysisJob.video_id == video_id,
            ).order_by(LiveAnalysisJob.created_at.desc())
        )
        existing_job = result.scalars().first()

        if existing_job:
            # BUILD 36: Clean up duplicate jobs
            try:
                from sqlalchemy import delete as sa_delete
                await db.execute(
                    sa_delete(LiveAnalysisJob).where(
                        LiveAnalysisJob.video_id == video_id,
                        LiveAnalysisJob.id != existing_job.id,
                    )
                )
            except Exception as cleanup_err:
                logger.warning(f"[admin/retry] Failed to cleanup duplicates: {cleanup_err}")
            existing_job.status = "pending"
            existing_job.current_step = None
            existing_job.progress = 0
            existing_job.error_message = None
            existing_job.started_at = None
            existing_job.completed_at = None
            existing_job.results = None
            job = existing_job
            total_chunks = existing_job.total_chunks
            stream_source = existing_job.stream_source or "tiktok_live"
        else:
            job = LiveAnalysisJob(
                id=uuid_module.uuid4(),
                video_id=video_id,
                user_id=user_id,
                stream_source="tiktok_live",
                status="pending",
                progress=0,
            )
            db.add(job)
            total_chunks = None
            stream_source = "tiktok_live"

        # Reset video status
        await db.execute(
            text("""
                UPDATE videos
                SET status = 'STEP_0_EXTRACT_FRAMES',
                    step_progress = 0,
                    error_message = NULL,
                    updated_at = now()
                WHERE id = :vid
            """),
            {"vid": video_id},
        )
        await db.commit()
        await db.refresh(job)

        # Verify chunks exist in blob storage before enqueuing
        from app.services.storage_service import check_blob_exists
        chunk_exists = check_blob_exists(
            email=user_email,
            video_id=video_id,
            filename="chunks/chunk_0000.mp4",
        )
        if not chunk_exists:
            logger.warning(
                f"[admin/retry-video/live_boost] No chunks in blob storage for "
                f"video={video_id} email={user_email}"
            )
            raise HTTPException(
                status_code=422,
                detail=(
                    f"Cannot retry: no chunks found in blob storage. "
                    f"Expected blob: {user_email}/{video_id}/chunks/chunk_0000.mp4. "
                    f"iOS app may have failed to upload chunks."
                ),
            )

        # Enqueue live_analysis job
        enqueue_result = await enqueue_job({
            "job_type": "live_analysis",
            "job_id": str(job.id),
            "video_id": video_id,
            "user_id": user_id,
            "stream_source": stream_source,
            "total_chunks": total_chunks,
            "email": user_email or "",
        })

        if enqueue_result.success:
            await db.execute(
                sa_update(LiveAnalysisJob)
                .where(LiveAnalysisJob.id == job.id)
                .values(
                    queue_message_id=enqueue_result.message_id,
                    queue_enqueued_at=enqueue_result.enqueued_at,
                    started_at=datetime.now(timezone.utc),
                )
            )
            await db.commit()
            logger.info(
                f"[admin/retry-video/live_boost] Enqueued OK job={job.id} video={video_id}"
            )
        else:
            logger.error(
                f"[admin/retry-video/live_boost] Enqueue FAILED: {enqueue_result.error}"
            )
            raise HTTPException(
                status_code=500,
                detail=f"LiveBoost enqueue failed: {enqueue_result.error}",
            )

        return {
            "status": "ok",
            "video_id": video_id,
            "job_id": str(job.id),
            "message": f"LiveBoost analysis re-enqueued (job={job.id})",
        }
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        logger.exception(f"[admin/retry-video/live_boost] Failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))



# ── Upload Health Check ──────────────────────────────────────────────────
@router.get("/upload-health")
async def get_upload_health(
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
):
    """Upload pipeline health metrics for the admin dashboard.

    Returns success/failure rates, average processing times, stuck uploads,
    and recent upload history.
    """
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")

    try:
        # ── Overall counts ──
        total_uploads = await _q(db, "SELECT COUNT(*) FROM videos")
        done_count = await _q(db, "SELECT COUNT(*) FROM videos WHERE status = 'DONE'")
        error_count = await _q(db, "SELECT COUNT(*) FROM videos WHERE status = 'ERROR'")
        processing_count = await _q(
            db,
            "SELECT COUNT(*) FROM videos WHERE status NOT IN ('DONE', 'ERROR', 'NEW', 'uploaded')",
        )
        queued_count = await _q(
            db,
            "SELECT COUNT(*) FROM videos WHERE status IN ('uploaded', 'NEW')",
        )

        success_rate = round(done_count / total_uploads * 100, 1) if total_uploads > 0 else 0.0
        error_rate = round(error_count / total_uploads * 100, 1) if total_uploads > 0 else 0.0

        # ── Last 24h ──
        uploads_24h = await _q(
            db,
            "SELECT COUNT(*) FROM videos WHERE created_at >= NOW() - INTERVAL '24 hours'",
        )
        done_24h = await _q(
            db,
            "SELECT COUNT(*) FROM videos WHERE status = 'DONE' AND created_at >= NOW() - INTERVAL '24 hours'",
        )
        error_24h = await _q(
            db,
            "SELECT COUNT(*) FROM videos WHERE status = 'ERROR' AND created_at >= NOW() - INTERVAL '24 hours'",
        )

        # ── Last 7 days ──
        uploads_7d = await _q(
            db,
            "SELECT COUNT(*) FROM videos WHERE created_at >= NOW() - INTERVAL '7 days'",
        )
        done_7d = await _q(
            db,
            "SELECT COUNT(*) FROM videos WHERE status = 'DONE' AND created_at >= NOW() - INTERVAL '7 days'",
        )
        error_7d = await _q(
            db,
            "SELECT COUNT(*) FROM videos WHERE status = 'ERROR' AND created_at >= NOW() - INTERVAL '7 days'",
        )

        # ── Stuck videos (processing > 2 hours) ──
        stuck_count = await _q(
            db,
            """
            SELECT COUNT(*) FROM videos
            WHERE status NOT IN ('DONE', 'ERROR', 'NEW', 'uploaded')
              AND created_at < NOW() - INTERVAL '2 hours'
            """,
        )

        # ── Recent uploads (last 20) ──
        try:
            recent_result = await db.execute(text("""
                SELECT
                    v.id,
                    v.original_filename,
                    v.status,
                    v.upload_type,
                    v.created_at,
                    u.email as user_email
                FROM videos v
                LEFT JOIN users u ON v.user_id = u.id
                ORDER BY v.created_at DESC
                LIMIT 20
            """))
            recent_rows = recent_result.fetchall()
            recent_uploads = [
                {
                    "video_id": str(row.id),
                    "filename": row.original_filename,
                    "status": row.status,
                    "upload_type": row.upload_type,
                    "created_at": str(row.created_at) if row.created_at else None,
                    "user_email": row.user_email,
                }
                for row in recent_rows
            ]
        except Exception as e:
            logger.warning(f"Failed to fetch recent uploads: {e}")
            try:
                await db.rollback()
            except Exception as _rb_err:
                logger.debug(f"Rollback cleanup failed: {_rb_err}")
            recent_uploads = []

        # ── Status distribution ──
        try:
            status_result = await db.execute(text("""
                SELECT status, COUNT(*) as cnt
                FROM videos
                GROUP BY status
                ORDER BY cnt DESC
            """))
            status_rows = status_result.fetchall()
            status_distribution = {row.status: row.cnt for row in status_rows}
        except Exception as e:
            logger.warning(f"Failed to fetch status distribution: {e}")
            try:
                await db.rollback()
            except Exception as _rb_err:
                logger.debug(f"Rollback cleanup failed: {_rb_err}")
            status_distribution = {}

        # ── Error breakdown (last 7 days) ──
        try:
            error_result = await db.execute(text("""
                SELECT
                    v.id,
                    v.original_filename,
                    v.status,
                    v.created_at,
                    u.email as user_email
                FROM videos v
                LEFT JOIN users u ON v.user_id = u.id
                WHERE v.status = 'ERROR'
                  AND v.created_at >= NOW() - INTERVAL '7 days'
                ORDER BY v.created_at DESC
                LIMIT 10
            """))
            error_rows = error_result.fetchall()
            recent_errors = [
                {
                    "video_id": str(row.id),
                    "filename": row.original_filename,
                    "created_at": str(row.created_at) if row.created_at else None,
                    "user_email": row.user_email,
                }
                for row in error_rows
            ]
        except Exception as e:
            logger.warning(f"Failed to fetch error breakdown: {e}")
            try:
                await db.rollback()
            except Exception as _rb_err:
                logger.debug(f"Rollback cleanup failed: {_rb_err}")
            recent_errors = []

        # ── Enqueue statistics ──
        enqueue_ok_count = await _q(
            db, "SELECT COUNT(*) FROM videos WHERE enqueue_status = 'OK'"
        )
        enqueue_failed_count = await _q(
            db, "SELECT COUNT(*) FROM videos WHERE enqueue_status = 'FAILED'"
        )
        enqueue_ok_24h = await _q(
            db,
            "SELECT COUNT(*) FROM videos WHERE enqueue_status = 'OK'"
            " AND created_at >= NOW() - INTERVAL '24 hours'",
        )
        enqueue_failed_24h = await _q(
            db,
            "SELECT COUNT(*) FROM videos WHERE enqueue_status = 'FAILED'"
            " AND created_at >= NOW() - INTERVAL '24 hours'",
        )
        enqueue_total = enqueue_ok_count + enqueue_failed_count
        enqueue_rate_pct = (
            round(enqueue_ok_count / enqueue_total * 100, 1) if enqueue_total > 0 else None
        )

        # ── Retry candidates (enqueue FAILED, not yet DONE/ERROR) ──
        try:
            retry_result = await db.execute(text("""
                SELECT
                    v.id,
                    v.original_filename,
                    v.status,
                    v.enqueue_status,
                    v.enqueue_error,
                    v.created_at,
                    u.email as user_email
                FROM videos v
                LEFT JOIN users u ON v.user_id = u.id
                WHERE v.enqueue_status = 'FAILED'
                  AND v.status NOT IN ('DONE', 'ERROR')
                ORDER BY v.created_at DESC
                LIMIT 10
            """))
            retry_rows = retry_result.fetchall()
            retry_candidates = [
                {
                    "video_id": str(row.id),
                    "filename": row.original_filename,
                    "status": row.status,
                    "enqueue_error": row.enqueue_error,
                    "created_at": str(row.created_at) if row.created_at else None,
                    "user_email": row.user_email,
                }
                for row in retry_rows
            ]
        except Exception as e:
            logger.warning(f"Failed to fetch retry candidates: {e}")
            try:
                await db.rollback()
            except Exception as _rb_err:
                logger.debug(f"Rollback cleanup failed: {_rb_err}")
            retry_candidates = []

        # ── Pipeline stage distribution ──
        uploaded_waiting = await _q(
            db, "SELECT COUNT(*) FROM videos WHERE status = 'uploaded'"
        )
        pipeline_stages = {
            "uploaded_waiting": uploaded_waiting,
            "processing": processing_count,
            "done": done_count,
            "error": error_count,
            "enqueue_failed": enqueue_failed_count,
            "stuck_gt_2h": stuck_count,
        }

        # ── Recent stage events (from upload_event_log) ──
        recent_stage_events = []
        try:
            stage_result = await db.execute(text("""
                SELECT
                    video_id, upload_id, user_id, stage, status,
                    duration_ms, error_message, error_type, created_at
                FROM upload_event_log
                WHERE status = 'error'
                ORDER BY created_at DESC
                LIMIT 20
            """))
            stage_rows = stage_result.fetchall()
            recent_stage_events = [
                {
                    "video_id": str(row.video_id) if row.video_id else None,
                    "upload_id": str(row.upload_id) if row.upload_id else None,
                    "stage": row.stage,
                    "status": row.status,
                    "duration_ms": row.duration_ms,
                    "error_message": row.error_message,
                    "error_type": row.error_type,
                    "created_at": str(row.created_at) if row.created_at else None,
                }
                for row in stage_rows
            ]
        except Exception as e:
            logger.warning(f"Failed to fetch stage events (table may not exist): {e}")
            try:
                await db.rollback()
            except Exception as _rb_err:
                logger.debug(f"Rollback cleanup failed: {_rb_err}")

        # ── Videos with upload_error_stage ──
        failed_stage_videos = []
        try:
            fs_result = await db.execute(text("""
                SELECT
                    v.id, v.original_filename, v.status,
                    v.upload_last_stage, v.upload_error_stage,
                    v.upload_error_message, v.created_at,
                    u.email as user_email
                FROM videos v
                LEFT JOIN users u ON v.user_id = u.id
                WHERE v.upload_error_stage IS NOT NULL
                ORDER BY v.created_at DESC
                LIMIT 10
            """))
            fs_rows = fs_result.fetchall()
            failed_stage_videos = [
                {
                    "video_id": str(row.id),
                    "filename": row.original_filename,
                    "status": row.status,
                    "last_stage": row.upload_last_stage,
                    "error_stage": row.upload_error_stage,
                    "error_message": (row.upload_error_message or "")[:200],
                    "created_at": str(row.created_at) if row.created_at else None,
                    "user_email": row.user_email,
                }
                for row in fs_rows
            ]
        except Exception as e:
            logger.warning(f"Failed to fetch failed stage videos (columns may not exist): {e}")
            try:
                await db.rollback()
            except Exception as _rb_err:
                logger.debug(f"Rollback cleanup failed: {_rb_err}")

        return {
            "overall": {
                "total_uploads": total_uploads,
                "done": done_count,
                "error": error_count,
                "processing": processing_count,
                "queued": queued_count,
                "success_rate_pct": success_rate,
                "error_rate_pct": error_rate,
            },
            "last_24h": {
                "uploads": uploads_24h,
                "done": done_24h,
                "error": error_24h,
            },
            "last_7d": {
                "uploads": uploads_7d,
                "done": done_7d,
                "error": error_7d,
            },
            "enqueue_stats": {
                "total_ok": enqueue_ok_count,
                "total_failed": enqueue_failed_count,
                "ok_last_24h": enqueue_ok_24h,
                "failed_last_24h": enqueue_failed_24h,
                "enqueue_success_rate_pct": enqueue_rate_pct,
            },
            "pipeline_stages": pipeline_stages,
            "retry_candidates": retry_candidates,
            "stuck_videos": stuck_count,
            "status_distribution": status_distribution,
            "recent_uploads": recent_uploads,
            "recent_errors": recent_errors,
            "recent_stage_events": recent_stage_events,
            "failed_stage_videos": failed_stage_videos,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Failed to get upload health: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── Recompute Phase Metrics (v2 — service-based) ─────────────────────────────

@router.post("/recompute-phase-metrics/{video_id}")
async def recompute_phase_metrics(
    video_id: str,
    dry_run: bool = True,
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None),
):
    """
    既存動画の Derived Data（phase metrics）を最新ロジックで再計算する。

    Raw Data / Human Data は一切変更しない。

    Parameters
    ----------
    video_id : str
        対象動画の UUID
    dry_run : bool
        True（デフォルト）の場合、計算結果を返すが DB は更新しない。
        False の場合、DB を更新する。
    """
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")

    # Ensure migration tables exist
    try:
        await db.execute(text("""
            CREATE TABLE IF NOT EXISTS phase_metrics_recalc_log (
                id BIGSERIAL PRIMARY KEY,
                video_id VARCHAR(255) NOT NULL,
                triggered_by VARCHAR(255),
                mode VARCHAR(20) NOT NULL DEFAULT 'dry-run',
                status VARCHAR(20) NOT NULL DEFAULT 'pending',
                logic_version INTEGER NOT NULL DEFAULT 1,
                before_json JSONB, after_json JSONB,
                diff_json JSONB, logs_json JSONB,
                error_message TEXT, duration_ms INTEGER,
                created_at TIMESTAMPTZ NOT NULL DEFAULT now()
            )
        """))
        await db.execute(text(
            "ALTER TABLE video_phases ADD COLUMN IF NOT EXISTS "
            "phase_metrics_version_applied INTEGER DEFAULT NULL"
        ))
        await db.execute(text(
            "ALTER TABLE videos ADD COLUMN IF NOT EXISTS "
            "phase_metrics_version_applied INTEGER DEFAULT NULL"
        ))
        await db.execute(text(
            "ALTER TABLE videos ADD COLUMN IF NOT EXISTS "
            "last_recalculated_at TIMESTAMPTZ DEFAULT NULL"
        ))
        await db.commit()
    except Exception as mig_err:
        logger.warning(f"Migration check: {mig_err}")
        try:
            await db.rollback()
        except Exception as _rb_err:
            logger.debug(f"Rollback cleanup failed: {_rb_err}")

    try:
        from app.services.phase_metrics_recalculator import recalculate_phase_metrics as _recalc
    except ImportError as e:
        raise HTTPException(
            status_code=500,
            detail=f"Cannot import recalculator service: {e}",
        )

    result = await _recalc(
        video_id=video_id,
        db=db,
        dry_run=dry_run,
        triggered_by=f"admin:{x_admin_key}",
    )

    if result["status"] == "error":
        error_logs = [l for l in result.get("logs", []) if "ERROR" in l]
        detail = error_logs[0] if error_logs else "Recalculation failed"
        raise HTTPException(status_code=400, detail=detail)

    return result


@router.get("/recalc-log/{video_id}")
async def get_recalc_log(
    video_id: str,
    limit: int = 10,
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None),
):
    """
    動画の再計算履歴を取得する。
    """
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        r = await db.execute(text("""
            SELECT id, triggered_by, mode, status, logic_version,
                   diff_json, error_message, duration_ms, created_at
            FROM phase_metrics_recalc_log
            WHERE video_id = :vid
            ORDER BY created_at DESC
            LIMIT :lim
        """), {"vid": video_id, "lim": limit})
        rows = r.fetchall()
    except Exception:
        return {"logs": [], "message": "recalc_log table may not exist yet"}

    logs = []
    for row in rows:
        import json as _json
        diff = None
        try:
            diff = _json.loads(row[5]) if row[5] else None
        except Exception:
            diff = row[5]
        logs.append({
            "id":             row[0],
            "triggered_by":   row[1],
            "mode":           row[2],
            "status":         row[3],
            "logic_version":  row[4],
            "diff":           diff,
            "error_message":  row[6],
            "duration_ms":    row[7],
            "created_at":     row[8].isoformat() if row[8] else None,
        })

    return {"video_id": video_id, "logs": logs}


@router.post("/recalc-all")
async def recalc_all_videos(
    dry_run: bool = True,
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None),
):
    """
    全動画の phase metrics を一括再計算する。
    """
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")

    # Get all eligible videos
    r = await db.execute(text("""
        SELECT v.id
        FROM videos v
        WHERE v.status IN ('completed', 'DONE')
          AND v.upload_type = 'clean_video'
          AND v.excel_trend_blob_url IS NOT NULL
          AND LENGTH(v.excel_trend_blob_url) > 5
          AND v.id IN (SELECT DISTINCT video_id FROM video_phases)
        ORDER BY v.created_at DESC
    """))
    video_ids = [str(row[0]) for row in r.fetchall()]

    if not video_ids:
        return {"status": "ok", "message": "No eligible videos found", "results": []}

    try:
        from app.services.phase_metrics_recalculator import recalculate_phase_metrics as _recalc
    except ImportError as e:
        raise HTTPException(status_code=500, detail=f"Cannot import recalculator: {e}")

    results = []
    for vid in video_ids:
        try:
            result = await _recalc(
                video_id=vid,
                db=db,
                dry_run=dry_run,
                triggered_by=f"admin:recalc-all",
            )
            results.append({
                "video_id": vid,
                "status": result["status"],
                "phases_updated": result.get("phases_updated", 0),
                "diff_summary": {
                    "phases_changed": result.get("diff", {}).get("phases_changed", 0),
                    "gmv_delta": result.get("diff", {}).get("gmv_delta", 0),
                },
            })
        except Exception as e:
            results.append({"video_id": vid, "status": "error", "error": str(e)})

    return {
        "status": "ok",
        "dry_run": dry_run,
        "total_videos": len(video_ids),
        "success": sum(1 for r in results if r["status"] == "success"),
        "errors": sum(1 for r in results if r["status"] == "error"),
        "results": results,
    }


# ──────────────────────────────────────────────────────────────────────
# Frontend Diagnostics endpoints
# ──────────────────────────────────────────────────────────────────────

@router.post("/frontend-diagnostics")
async def report_frontend_error(
    payload: dict,
    db: AsyncSession = Depends(get_db),
):
    """
    Frontend からセクションエラーを受信して DB に保存する。
    認証不要（エラー報告はログイン失敗時にも送れる必要がある）。

    Payload:
        video_id      - 動画ID
        section_name  - セクション名 (e.g., "MomentClips")
        endpoint      - APIエンドポイント
        error_type    - エラータイプ (auth/not_found/timeout/server/network/parse/unknown)
        error_message - エラーメッセージ
        http_status   - HTTPステータスコード
        request_id    - X-Request-Id
        page_url      - ページURL
        user_agent    - ブラウザUA
    """
    try:
        # Auto-create table if not exists (same pattern as phase_metrics_recalc_log)
        await db.execute(text("""
            CREATE TABLE IF NOT EXISTS frontend_diagnostics (
                id BIGSERIAL PRIMARY KEY,
                video_id VARCHAR(255) NOT NULL DEFAULT '',
                section_name VARCHAR(100) NOT NULL DEFAULT 'unknown',
                endpoint VARCHAR(500) DEFAULT '',
                error_type VARCHAR(50) NOT NULL DEFAULT 'unknown',
                error_message TEXT,
                http_status INTEGER,
                request_id VARCHAR(100) DEFAULT '',
                page_url VARCHAR(1000) DEFAULT '',
                user_agent VARCHAR(500) DEFAULT '',
                created_at TIMESTAMPTZ NOT NULL DEFAULT now()
            )
        """))
        await db.execute(text("""
            CREATE INDEX IF NOT EXISTS ix_fd_video_id ON frontend_diagnostics (video_id)
        """))
        await db.execute(text("""
            CREATE INDEX IF NOT EXISTS ix_fd_section_name ON frontend_diagnostics (section_name)
        """))
        await db.execute(text("""
            CREATE INDEX IF NOT EXISTS ix_fd_error_type ON frontend_diagnostics (error_type)
        """))
        await db.execute(text("""
            CREATE INDEX IF NOT EXISTS ix_fd_created_at ON frontend_diagnostics (created_at)
        """))

        sql = text("""
            INSERT INTO frontend_diagnostics
                (video_id, section_name, endpoint, error_type, error_message,
                 http_status, request_id, page_url, user_agent)
            VALUES
                (:video_id, :section_name, :endpoint, :error_type, :error_message,
                 :http_status, :request_id, :page_url, :user_agent)
        """)
        await db.execute(sql, {
            "video_id": payload.get("video_id", ""),
            "section_name": payload.get("section_name", "unknown"),
            "endpoint": payload.get("endpoint", ""),
            "error_type": payload.get("error_type", "unknown"),
            "error_message": str(payload.get("error_message", ""))[:2000],
            "http_status": payload.get("http_status"),
            "request_id": payload.get("request_id", ""),
            "page_url": str(payload.get("page_url", ""))[:1000],
            "user_agent": str(payload.get("user_agent", ""))[:500],
        })
        await db.commit()
        return {"status": "ok"}
    except Exception as e:
        logger.warning(f"Failed to save frontend diagnostic: {e}")
        try:
            await db.rollback()
        except Exception as _rb_err:
            logger.debug(f"Rollback cleanup failed: {_rb_err}")
        # エラー報告の保存失敗はフロントに影響させない
        return {"status": "ok", "note": "save_failed"}


@router.get("/frontend-diagnostics")
async def get_frontend_diagnostics(
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
    video_id: Optional[str] = None,
    section_name: Optional[str] = None,
    error_type: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
):
    """
    Admin 用: Frontend エラーログを取得する。
    フィルタ: video_id, section_name, error_type
    """
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")

    try:
        # Ensure table exists
        await db.execute(text("""
            CREATE TABLE IF NOT EXISTS frontend_diagnostics (
                id BIGSERIAL PRIMARY KEY,
                video_id VARCHAR(255) NOT NULL DEFAULT '',
                section_name VARCHAR(100) NOT NULL DEFAULT 'unknown',
                endpoint VARCHAR(500) DEFAULT '',
                error_type VARCHAR(50) NOT NULL DEFAULT 'unknown',
                error_message TEXT,
                http_status INTEGER,
                request_id VARCHAR(100) DEFAULT '',
                page_url VARCHAR(1000) DEFAULT '',
                user_agent VARCHAR(500) DEFAULT '',
                created_at TIMESTAMPTZ NOT NULL DEFAULT now()
            )
        """))

        conditions = []
        params = {"lim": limit, "off": offset}

        if video_id:
            conditions.append("video_id = :vid")
            params["vid"] = video_id
        if section_name:
            conditions.append("section_name = :sn")
            params["sn"] = section_name
        if error_type:
            conditions.append("error_type = :et")
            params["et"] = error_type

        where_clause = ("WHERE " + " AND ".join(conditions)) if conditions else ""

        # 集計
        count_sql = text(f"SELECT COUNT(*) FROM frontend_diagnostics {where_clause}")
        total = (await db.execute(count_sql, params)).scalar() or 0

        # エラーログ一覧
        sql = text(f"""
            SELECT id, video_id, section_name, endpoint, error_type,
                   error_message, http_status, request_id, page_url,
                   user_agent, created_at
            FROM frontend_diagnostics
            {where_clause}
            ORDER BY created_at DESC
            LIMIT :lim OFFSET :off
        """)
        result = await db.execute(sql, params)
        rows = result.fetchall()

        errors = []
        for r in rows:
            errors.append({
                "id": r.id,
                "video_id": r.video_id,
                "section_name": r.section_name,
                "endpoint": r.endpoint,
                "error_type": r.error_type,
                "error_message": r.error_message,
                "http_status": r.http_status,
                "request_id": r.request_id,
                "page_url": r.page_url,
                "created_at": str(r.created_at) if r.created_at else None,
            })

        # セクション別集計
        summary_sql = text("""
            SELECT section_name, error_type, COUNT(*) as cnt
            FROM frontend_diagnostics
            WHERE created_at >= NOW() - INTERVAL '24 hours'
            GROUP BY section_name, error_type
            ORDER BY cnt DESC
        """)
        summary_result = await db.execute(summary_sql)
        summary_rows = summary_result.fetchall()

        section_summary = {}
        for sr in summary_rows:
            sn = sr.section_name
            if sn not in section_summary:
                section_summary[sn] = {"total": 0, "by_type": {}}
            section_summary[sn]["total"] += sr.cnt
            section_summary[sn]["by_type"][sr.error_type] = sr.cnt

        return {
            "total": total,
            "errors": errors,
            "section_summary_24h": section_summary,
        }
    except Exception as e:
        logger.exception(f"Failed to fetch frontend diagnostics: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/frontend-diagnostics/summary")
async def get_frontend_diagnostics_summary(
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
    hours: int = 24,
):
    """
    Admin 用: Frontend エラーのサマリーを取得する。
    - セクション別エラー件数
    - エラータイプ別件数
    - 直近のエラー傾向
    """
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")

    try:
        # Ensure table exists
        await db.execute(text("""
            CREATE TABLE IF NOT EXISTS frontend_diagnostics (
                id BIGSERIAL PRIMARY KEY,
                video_id VARCHAR(255) NOT NULL DEFAULT '',
                section_name VARCHAR(100) NOT NULL DEFAULT 'unknown',
                endpoint VARCHAR(500) DEFAULT '',
                error_type VARCHAR(50) NOT NULL DEFAULT 'unknown',
                error_message TEXT,
                http_status INTEGER,
                request_id VARCHAR(100) DEFAULT '',
                page_url VARCHAR(1000) DEFAULT '',
                user_agent VARCHAR(500) DEFAULT '',
                created_at TIMESTAMPTZ NOT NULL DEFAULT now()
            )
        """))

        interval = f"{hours} hours"

        # セクション別
        by_section_sql = text(f"""
            SELECT section_name, COUNT(*) as cnt
            FROM frontend_diagnostics
            WHERE created_at >= NOW() - INTERVAL '{interval}'
            GROUP BY section_name
            ORDER BY cnt DESC
        """)
        by_section = await db.execute(by_section_sql)
        section_counts = {r.section_name: r.cnt for r in by_section.fetchall()}

        # エラータイプ別
        by_type_sql = text(f"""
            SELECT error_type, COUNT(*) as cnt
            FROM frontend_diagnostics
            WHERE created_at >= NOW() - INTERVAL '{interval}'
            GROUP BY error_type
            ORDER BY cnt DESC
        """)
        by_type = await db.execute(by_type_sql)
        type_counts = {r.error_type: r.cnt for r in by_type.fetchall()}

        # 総件数
        total_sql = text(f"""
            SELECT COUNT(*) FROM frontend_diagnostics
            WHERE created_at >= NOW() - INTERVAL '{interval}'
        """)
        total = (await db.execute(total_sql)).scalar() or 0

        # 直近10件
        recent_sql = text(f"""
            SELECT video_id, section_name, error_type, request_id, created_at
            FROM frontend_diagnostics
            WHERE created_at >= NOW() - INTERVAL '{interval}'
            ORDER BY created_at DESC
            LIMIT 10
        """)
        recent_result = await db.execute(recent_sql)
        recent = [
            {
                "video_id": r.video_id,
                "section_name": r.section_name,
                "error_type": r.error_type,
                "request_id": r.request_id,
                "created_at": str(r.created_at) if r.created_at else None,
            }
            for r in recent_result.fetchall()
        ]

        return {
            "period_hours": hours,
            "total_errors": total,
            "by_section": section_counts,
            "by_error_type": type_counts,
            "recent_errors": recent,
        }
    except Exception as e:
        logger.exception(f"Failed to fetch diagnostics summary: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ─── CSV Validation Log ───

@router.post("/csv-validation-log")
async def log_csv_validation(
    request: Request,
    db: AsyncSession = Depends(get_db),
):
    """
    CSV Date/Time Validation Gate の判定結果とユーザーの選択をログ保存する。
    テーブルが存在しない場合は自動作成する。
    """
    try:
        body = await request.json()

        # テーブル自動作成 (PostgreSQL互換)
        create_sql = text("""
            CREATE TABLE IF NOT EXISTS csv_validation_logs (
                id BIGSERIAL PRIMARY KEY,
                verdict VARCHAR(20),
                decision VARCHAR(20),
                video_filename VARCHAR(500),
                trend_filename VARCHAR(500),
                product_filename VARCHAR(500),
                checks JSONB,
                user_email VARCHAR(255),
                created_at TIMESTAMPTZ DEFAULT NOW()
            )
        """)
        await db.execute(create_sql)
        for idx_sql in [
            "CREATE INDEX IF NOT EXISTS idx_csv_val_created ON csv_validation_logs (created_at)",
            "CREATE INDEX IF NOT EXISTS idx_csv_val_verdict ON csv_validation_logs (verdict)",
            "CREATE INDEX IF NOT EXISTS idx_csv_val_decision ON csv_validation_logs (decision)",
        ]:
            await db.execute(text(idx_sql))

        insert_sql = text("""
            INSERT INTO csv_validation_logs
                (verdict, decision, video_filename, trend_filename, product_filename, checks, user_email)
            VALUES
                (:verdict, :decision, :video_filename, :trend_filename, :product_filename, :checks, :user_email)
        """)
        import json as json_mod
        await db.execute(insert_sql, {
            "verdict": body.get("verdict"),
            "decision": body.get("decision"),
            "video_filename": body.get("video_filename", "")[:500],
            "trend_filename": body.get("trend_filename", "")[:500],
            "product_filename": body.get("product_filename", "")[:500],
            "checks": json_mod.dumps(body.get("checks", []), ensure_ascii=False),
            "user_email": body.get("user_email", "")[:255],
        })
        await db.commit()

        return {"status": "ok"}
    except Exception as e:
        logger.warning(f"Failed to log CSV validation: {e}")
        return {"status": "error", "detail": str(e)}


@router.get("/csv-validation-logs")
async def get_csv_validation_logs(
    request: Request,
    db: AsyncSession = Depends(get_db),
    limit: int = 50,
    offset: int = 0,
    verdict: str = None,
    decision: str = None,
):
    """
    CSV Validation ログ一覧を取得する（Admin用）。
    """
    try:
        # テーブル自動作成 (PostgreSQL互換)
        create_sql = text("""
            CREATE TABLE IF NOT EXISTS csv_validation_logs (
                id BIGSERIAL PRIMARY KEY,
                verdict VARCHAR(20),
                decision VARCHAR(20),
                video_filename VARCHAR(500),
                trend_filename VARCHAR(500),
                product_filename VARCHAR(500),
                checks JSONB,
                user_email VARCHAR(255),
                created_at TIMESTAMPTZ DEFAULT NOW()
            )
        """)
        await db.execute(create_sql)
        for idx_sql in [
            "CREATE INDEX IF NOT EXISTS idx_csv_val_created ON csv_validation_logs (created_at)",
            "CREATE INDEX IF NOT EXISTS idx_csv_val_verdict ON csv_validation_logs (verdict)",
            "CREATE INDEX IF NOT EXISTS idx_csv_val_decision ON csv_validation_logs (decision)",
        ]:
            await db.execute(text(idx_sql))

        where_clauses = ["1=1"]
        params = {"limit_val": limit, "offset_val": offset}

        if verdict:
            where_clauses.append("verdict = :verdict")
            params["verdict"] = verdict
        if decision:
            where_clauses.append("decision = :decision")
            params["decision"] = decision

        where_str = " AND ".join(where_clauses)

        sql = text(f"""
            SELECT id, verdict, decision, video_filename, trend_filename, product_filename,
                   checks, user_email, created_at
            FROM csv_validation_logs
            WHERE {where_str}
            ORDER BY created_at DESC
            LIMIT :limit_val OFFSET :offset_val
        """)
        result = await db.execute(sql, params)
        rows = result.fetchall()

        return {
            "logs": [
                {
                    "id": r.id,
                    "verdict": r.verdict,
                    "decision": r.decision,
                    "video_filename": r.video_filename,
                    "trend_filename": r.trend_filename,
                    "product_filename": r.product_filename,
                    "checks": r.checks,
                    "user_email": r.user_email,
                    "created_at": str(r.created_at) if r.created_at else None,
                }
                for r in rows
            ],
            "limit": limit,
            "offset": offset,
        }
    except Exception as e:
        logger.exception(f"Failed to fetch CSV validation logs: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ─── Recalculate CSV Metrics from Excel ───

@router.post("/recalc-csv-metrics/{video_id}")
async def recalc_csv_metrics(
    video_id: str,
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None),
):
    """
    ExcelトレンドデータからCSVメトリクスを再計算し、video_phasesテーブルに保存する。
    ワーカーVMに依存せず、バックエンド側で直接計算を行う。

    按分ロジック:
    - CSVの各30分スロットを、そのスロット内のフェーズに時間比例で按分
    - 加算型メトリクス（GMV, orders等）: 按分
    - スナップショット型メトリクス（viewers, likes等）: 最大値
    """
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")

    import json as _json
    import httpx
    import tempfile
    import os
    import re
    from datetime import datetime, timedelta

    try:
        from app.services.storage_service import generate_read_sas_from_url
    except ImportError as e:
        raise HTTPException(status_code=500, detail=f"Cannot import storage_service: {e}")

    # Step 1: Get video info
    result = await db.execute(text("""
        SELECT v.excel_trend_blob_url, v.upload_type, v.time_offset_seconds,
               u.email
        FROM videos v
        LEFT JOIN users u ON v.user_id = u.id
        WHERE v.id = :vid
    """), {"vid": video_id})
    video_row = result.fetchone()
    if not video_row:
        raise HTTPException(status_code=404, detail="Video not found")

    trend_blob_url = video_row[0]
    upload_type = video_row[1]
    time_offset_seconds = float(video_row[2] or 0)

    if not trend_blob_url:
        raise HTTPException(status_code=400, detail="No trend Excel URL for this video")

    # Step 2: Get all phases
    phases_result = await db.execute(text("""
        SELECT phase_index, time_start, time_end
        FROM video_phases
        WHERE video_id = :vid
        ORDER BY phase_index ASC
    """), {"vid": video_id})
    phases = [{"phase_index": r[0], "time_start": float(r[1] or 0), "time_end": float(r[2] or 0)}
              for r in phases_result.fetchall()]

    if not phases:
        raise HTTPException(status_code=400, detail="No phases found for this video")

    # Step 3: Download and parse trend Excel
    async def _parse_excel(blob_url: str) -> list:
        sas_url = generate_read_sas_from_url(blob_url, expires_hours=1)
        if not sas_url:
            return []
        import openpyxl
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.get(sas_url)
            if resp.status_code != 200:
                return []
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
                f.write(resp.content)
                tmp_path = f.name
            try:
                wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
                ws = wb.active
                items = []
                if ws:
                    rows_data = list(ws.iter_rows(values_only=True))
                    if len(rows_data) >= 2:
                        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows_data[0])]
                        for data_row in rows_data[1:]:
                            if all(v is None for v in data_row):
                                continue
                            item = {}
                            for i, val in enumerate(data_row):
                                if i < len(headers):
                                    if val is None:
                                        item[headers[i]] = None
                                    elif isinstance(val, (int, float)):
                                        item[headers[i]] = val
                                    elif hasattr(val, 'hour') and hasattr(val, 'minute'):
                                        # datetime.time object → format as HH:MM
                                        item[headers[i]] = f"{val.hour:02d}:{val.minute:02d}"
                                    else:
                                        item[headers[i]] = str(val)
                            items.append(item)
                wb.close()
                return items
            finally:
                os.unlink(tmp_path)

    trends = await _parse_excel(trend_blob_url)
    if not trends:
        raise HTTPException(status_code=400, detail="Failed to parse trend Excel or no data")

    # Step 4: Detect column keys
    sample = trends[0]
    logs = [f"Trend entries: {len(trends)}", f"Phase count: {len(phases)}"]
    logs.append(f"Column headers: {list(sample.keys())}")

    # KPI aliases for column detection
    KPI_ALIASES = {
        "time": ["時間", "time", "timestamp", "时间", "시간"],
        "gmv": ["GMV", "gmv", "売上", "sales", "revenue", "成交金额", "매출"],
        "order_count": ["注文", "order", "orders", "SKU注文数", "订单", "주문"],
        "viewer_count": ["視聴者", "viewer", "viewers", "视聴者", "观众", "시청자"],
        "like_count": ["いいね", "like", "likes", "点赞", "좋아요"],
        "comment_count": ["コメント", "comment", "comments", "评论", "댓글"],
        "share_count": ["シェア", "share", "shares", "分享", "공유"],
        "new_followers": ["フォロワー", "follower", "followers", "新規フォロワー", "粉丝"],
        "product_clicks": ["商品クリック", "product_click", "clicks", "点击"],
        "ctor": ["CTOR", "ctor", "conversion"],
        "gpm": ["GPM", "gpm", "視聴GPM"],
    }

    def _find_key(sample_dict, aliases):
        for alias in aliases:
            for k in sample_dict.keys():
                if alias.lower() in k.lower():
                    return k
        return None

    def _safe_float(val):
        if val is None:
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    def _parse_time_to_seconds(val):
        if val is None:
            return None
        if hasattr(val, 'hour') and hasattr(val, 'minute'):
            return val.hour * 3600 + val.minute * 60 + getattr(val, 'second', 0)
        val_str = str(val).strip()
        try:
            return float(val_str)
        except (ValueError, TypeError):
            pass
        parts = val_str.split(":")
        try:
            if len(parts) == 2:
                h, m = int(parts[0]), int(parts[1])
                if h < 24:
                    return h * 3600 + m * 60
                else:
                    return h * 60 + m
            elif len(parts) == 3:
                h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
                return h * 3600 + m * 60 + s
        except (ValueError, TypeError):
            pass
        return None

    time_key = _find_key(sample, KPI_ALIASES["time"])
    gmv_key = _find_key(sample, KPI_ALIASES["gmv"])
    order_key = _find_key(sample, KPI_ALIASES["order_count"])
    viewer_key = _find_key(sample, KPI_ALIASES["viewer_count"])
    like_key = _find_key(sample, KPI_ALIASES["like_count"])
    comment_key = _find_key(sample, KPI_ALIASES["comment_count"])
    share_key = _find_key(sample, KPI_ALIASES["share_count"])
    follower_key = _find_key(sample, KPI_ALIASES["new_followers"])
    click_key = _find_key(sample, KPI_ALIASES["product_clicks"])
    conv_key = _find_key(sample, KPI_ALIASES["ctor"])
    gpm_key = _find_key(sample, KPI_ALIASES["gpm"])

    logs.append(f"Detected keys: time={time_key}, gmv={gmv_key}, order={order_key}, "
                f"viewer={viewer_key}, like={like_key}, comment={comment_key}, "
                f"share={share_key}, follower={follower_key}, click={click_key}, "
                f"conv={conv_key}, gpm={gpm_key}")

    if not time_key:
        raise HTTPException(status_code=400, detail=f"Cannot detect time column. Headers: {list(sample.keys())}")

    # Step 5: Build timed entries
    timed_entries = []
    for entry in trends:
        t_sec = _parse_time_to_seconds(entry.get(time_key))
        if t_sec is not None:
            timed_entries.append({"time_sec": t_sec, "entry": entry})
    timed_entries.sort(key=lambda x: x["time_sec"])

    if not timed_entries:
        raise HTTPException(status_code=400, detail="No valid time entries found in trend data")

    csv_first_sec = timed_entries[0]["time_sec"]
    logs.append(f"Timed entries: {len(timed_entries)}")
    logs.append(f"CSV first sec: {csv_first_sec}, time_offset: {time_offset_seconds}")
    logs.append(f"CSV times: {[te['time_sec'] for te in timed_entries]}")

    # Step 6: Build CSV slots
    csv_slots = []
    for i, te in enumerate(timed_entries):
        slot_start = te["time_sec"]
        if i + 1 < len(timed_entries):
            slot_end = timed_entries[i + 1]["time_sec"]
        else:
            video_end_abs = csv_first_sec + time_offset_seconds + phases[-1]["time_end"]
            slot_end = max(slot_start + 1800, video_end_abs)
        csv_slots.append({"start": slot_start, "end": slot_end, "entry": te["entry"]})

    slot_strs = [f"{s['start']:.0f}-{s['end']:.0f}" for s in csv_slots]
    logs.append(f"CSV slots: {slot_strs}")

    # Step 7: Calculate metrics for each phase
    updates = []
    for p in phases:
        start_sec = p["time_start"]
        end_sec = p["time_end"]

        phase_abs_start = csv_first_sec + time_offset_seconds + start_sec
        phase_abs_end = csv_first_sec + time_offset_seconds + end_sec

        phase_gmv = 0.0
        phase_orders = 0.0
        phase_viewers = 0
        phase_likes = 0
        phase_comments = 0.0
        phase_shares = 0.0
        phase_followers = 0.0
        phase_clicks = 0.0
        phase_conv = 0.0
        phase_gpm = 0.0
        match_count = 0

        for slot in csv_slots:
            overlap_start = max(phase_abs_start, slot["start"])
            overlap_end = min(phase_abs_end, slot["end"])
            overlap = max(0, overlap_end - overlap_start)
            if overlap <= 0:
                continue

            slot_dur = max(slot["end"] - slot["start"], 1)
            ratio = overlap / slot_dur
            e = slot["entry"]
            match_count += 1

            if gmv_key:
                phase_gmv += (_safe_float(e.get(gmv_key)) or 0) * ratio
            if order_key:
                phase_orders += (_safe_float(e.get(order_key)) or 0) * ratio
            if comment_key:
                phase_comments += (_safe_float(e.get(comment_key)) or 0) * ratio
            if share_key:
                phase_shares += (_safe_float(e.get(share_key)) or 0) * ratio
            if follower_key:
                phase_followers += (_safe_float(e.get(follower_key)) or 0) * ratio
            if click_key:
                phase_clicks += (_safe_float(e.get(click_key)) or 0) * ratio

            if viewer_key:
                phase_viewers = max(phase_viewers, int(_safe_float(e.get(viewer_key)) or 0))
            if like_key:
                phase_likes = max(phase_likes, int(_safe_float(e.get(like_key)) or 0))
            if conv_key:
                phase_conv = max(phase_conv, _safe_float(e.get(conv_key)) or 0)
            if gpm_key:
                phase_gpm = max(phase_gpm, _safe_float(e.get(gpm_key)) or 0)

        phase_orders = int(round(phase_orders))
        phase_comments = int(round(phase_comments))
        phase_shares = int(round(phase_shares))
        phase_followers = int(round(phase_followers))
        phase_clicks = int(round(phase_clicks))

        updates.append({
            "phase_index": p["phase_index"],
            "gmv": round(phase_gmv, 2),
            "order_count": phase_orders,
            "viewer_count": phase_viewers,
            "like_count": phase_likes,
            "comment_count": phase_comments,
            "share_count": phase_shares,
            "new_followers": phase_followers,
            "product_clicks": phase_clicks,
            "conversion_rate": round(phase_conv, 6),
            "gpm": round(phase_gpm, 2),
            "match_count": match_count,
        })

    # Log sample
    if updates:
        logs.append(f"Phase 1: gmv={updates[0]['gmv']}, viewers={updates[0]['viewer_count']}, matches={updates[0]['match_count']}")
        mid = len(updates) // 2
        logs.append(f"Phase {updates[mid]['phase_index']}: gmv={updates[mid]['gmv']}, viewers={updates[mid]['viewer_count']}, matches={updates[mid]['match_count']}")
        logs.append(f"Phase {updates[-1]['phase_index']}: gmv={updates[-1]['gmv']}, viewers={updates[-1]['viewer_count']}, matches={updates[-1]['match_count']}")

    phases_with_data = sum(1 for u in updates if u["gmv"] > 0 or u["viewer_count"] > 0)
    logs.append(f"Phases with data: {phases_with_data}/{len(updates)}")

    # Step 8: Update DB
    updated_count = 0
    for u in updates:
        try:
            await db.execute(text("""
                UPDATE video_phases
                SET gmv = :gmv, order_count = :order_count,
                    viewer_count = :viewer_count, like_count = :like_count,
                    comment_count = :comment_count, share_count = :share_count,
                    new_followers = :new_followers, product_clicks = :product_clicks,
                    conversion_rate = :conversion_rate, gpm = :gpm,
                    importance_score = :match_count,
                    updated_at = now()
                WHERE video_id = :video_id AND phase_index = :phase_index
            """), {
                "video_id": video_id,
                "phase_index": u["phase_index"],
                "gmv": u["gmv"],
                "order_count": u["order_count"],
                "viewer_count": u["viewer_count"],
                "like_count": u["like_count"],
                "comment_count": u["comment_count"],
                "share_count": u["share_count"],
                "new_followers": u["new_followers"],
                "product_clicks": u["product_clicks"],
                "conversion_rate": u["conversion_rate"],
                "gpm": u["gpm"],
                "match_count": u["match_count"],
            })
            updated_count += 1
        except Exception as e:
            logs.append(f"ERROR updating phase {u['phase_index']}: {e}")

    await db.commit()
    logs.append(f"Updated {updated_count}/{len(updates)} phases in DB")

    return {
        "status": "success",
        "video_id": video_id,
        "phases_total": len(phases),
        "phases_with_data": phases_with_data,
        "phases_updated": updated_count,
        "logs": logs,
        "sample_metrics": updates[:3] if updates else [],
    }


# ─── Force Video Status Update ───

@router.post("/force-status/{video_id}")
async def force_video_status(
    video_id: str,
    payload: dict,
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None),
):
    """
    動画のステータスを強制的に変更する。
    payload: {"status": "DONE"}
    """
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")

    new_status = payload.get("status")
    if not new_status:
        raise HTTPException(status_code=400, detail="status is required")

    await db.execute(
        text("UPDATE videos SET status = :status, step_progress = 0 WHERE id = :vid"),
        {"status": new_status, "vid": video_id},
    )
    await db.commit()

    return {"status": "ok", "video_id": video_id, "new_status": new_status}



# ═══════════════════════════════════════════════════════════════════════
# SYSTEM ERROR LOGS (video_error_logs) – 管理画面表示用
# ═══════════════════════════════════════════════════════════════════════

@router.get("/system-error-logs")
async def get_system_error_logs(
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None),
    video_id: Optional[str] = None,
    error_code: Optional[str] = None,
    source: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
):
    """video_error_logs の一覧を返す（管理画面 Diagnostics 用）"""
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        # Ensure table exists
        await db.execute(text("""
            CREATE TABLE IF NOT EXISTS video_error_logs (
                id BIGSERIAL PRIMARY KEY,
                video_id UUID NOT NULL,
                error_code VARCHAR(100) NOT NULL,
                error_step VARCHAR(100),
                error_message TEXT,
                error_detail TEXT,
                source VARCHAR(50) DEFAULT 'worker',
                created_at TIMESTAMPTZ DEFAULT NOW()
            )
        """))

        conditions = []
        params = {"limit": limit, "offset": offset}
        if video_id:
            conditions.append("CAST(vel.video_id AS TEXT) LIKE :video_id")
            params["video_id"] = f"%{video_id}%"
        if error_code:
            conditions.append("vel.error_code ILIKE :error_code")
            params["error_code"] = f"%{error_code}%"
        if source:
            conditions.append("vel.source = :source")
            params["source"] = source

        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

        count_result = await db.execute(
            text(f"SELECT COUNT(*) FROM video_error_logs vel {where}"),
            params,
        )
        total = count_result.scalar() or 0

        result = await db.execute(
            text(f"""
                SELECT vel.id, vel.video_id, vel.error_code, vel.error_step,
                       vel.error_message, vel.source, vel.created_at,
                       v.original_filename
                FROM video_error_logs vel
                LEFT JOIN videos v ON vel.video_id = v.id
                {where}
                ORDER BY vel.created_at DESC
                LIMIT :limit OFFSET :offset
            """),
            params,
        )
        rows = result.fetchall()
        errors = []
        for r in rows:
            errors.append({
                "id": r.id,
                "video_id": str(r.video_id),
                "filename": r.original_filename or "",
                "error_code": r.error_code,
                "error_step": r.error_step,
                "error_message": r.error_message,
                "source": r.source,
                "created_at": r.created_at.isoformat() if r.created_at else None,
            })

        return {"total": total, "errors": errors}
    except Exception as e:
        logger.exception(f"Failed to fetch system error logs: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/system-error-logs/summary")
async def get_system_error_logs_summary(
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None),
    hours: int = 24,
):
    """直近N時間のエラーサマリー"""
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        cutoff = f"NOW() - INTERVAL '{hours} hours'"

        total_result = await db.execute(
            text(f"SELECT COUNT(*) FROM video_error_logs WHERE created_at >= {cutoff}")
        )
        total = total_result.scalar() or 0

        by_code_result = await db.execute(
            text(f"""
                SELECT error_code, COUNT(*) as cnt
                FROM video_error_logs WHERE created_at >= {cutoff}
                GROUP BY error_code ORDER BY cnt DESC LIMIT 20
            """)
        )
        by_code = {r.error_code: r.cnt for r in by_code_result.fetchall()}

        by_step_result = await db.execute(
            text(f"""
                SELECT error_step, COUNT(*) as cnt
                FROM video_error_logs WHERE created_at >= {cutoff}
                GROUP BY error_step ORDER BY cnt DESC LIMIT 20
            """)
        )
        by_step = {(r.error_step or "unknown"): r.cnt for r in by_step_result.fetchall()}

        by_source_result = await db.execute(
            text(f"""
                SELECT source, COUNT(*) as cnt
                FROM video_error_logs WHERE created_at >= {cutoff}
                GROUP BY source ORDER BY cnt DESC
            """)
        )
        by_source = {(r.source or "unknown"): r.cnt for r in by_source_result.fetchall()}

        return {
            "total_errors": total,
            "period_hours": hours,
            "by_error_code": by_code,
            "by_step": by_step,
            "by_source": by_source,
        }
    except Exception as e:
        logger.exception(f"Failed to fetch error summary: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ═══════════════════════════════════════════════════════════════════════
# BUG REPORTS – 問題→原因→解決策の記録
# ═══════════════════════════════════════════════════════════════════════

@router.get("/bug-reports")
async def list_bug_reports(
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None),
    status: Optional[str] = None,
    severity: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
):
    """バグレポート一覧"""
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conditions = []
        params = {"limit": limit, "offset": offset}
        if status:
            conditions.append("status = :status")
            params["status"] = status
        if severity:
            conditions.append("severity = :severity")
            params["severity"] = severity

        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

        count_result = await db.execute(
            text(f"SELECT COUNT(*) FROM bug_reports {where}"), params
        )
        total = count_result.scalar() or 0

        result = await db.execute(
            text(f"""
                SELECT * FROM bug_reports {where}
                ORDER BY
                    CASE status WHEN 'open' THEN 0 WHEN 'investigating' THEN 1
                                WHEN 'resolved' THEN 2 ELSE 3 END,
                    created_at DESC
                LIMIT :limit OFFSET :offset
            """),
            params,
        )
        rows = result.fetchall()
        reports = [dict(r._mapping) for r in rows]
        # Serialize datetimes
        for r in reports:
            for k in ("created_at", "updated_at", "resolved_at"):
                if r.get(k):
                    r[k] = r[k].isoformat()

        return {"total": total, "reports": reports}
    except Exception as e:
        logger.exception(f"Failed to fetch bug reports: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/bug-reports")
async def create_bug_report(
    payload: dict,
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None),
):
    """バグレポート作成"""
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        result = await db.execute(
            text("""
                INSERT INTO bug_reports
                    (title, severity, status, category, symptom, root_cause,
                     solution, affected_files, related_video_ids, reported_by)
                VALUES
                    (:title, :severity, :status, :category, :symptom, :root_cause,
                     :solution, :affected_files, :related_video_ids, :reported_by)
                RETURNING id
            """),
            {
                "title": payload.get("title", "Untitled"),
                "severity": payload.get("severity", "medium"),
                "status": payload.get("status", "open"),
                "category": payload.get("category", "general"),
                "symptom": payload.get("symptom", ""),
                "root_cause": payload.get("root_cause", ""),
                "solution": payload.get("solution", ""),
                "affected_files": payload.get("affected_files", ""),
                "related_video_ids": payload.get("related_video_ids", ""),
                "reported_by": payload.get("reported_by", "system"),
            },
        )
        bug_id = result.scalar()
        await db.commit()
        return {"status": "ok", "id": bug_id}
    except Exception as e:
        await db.rollback()
        logger.exception(f"Failed to create bug report: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/bug-reports/{bug_id}")
async def update_bug_report(
    bug_id: int,
    payload: dict,
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None),
):
    """バグレポート更新（解決策の追記など）"""
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        set_clauses = []
        params = {"bug_id": bug_id}

        allowed_fields = [
            "title", "severity", "status", "category", "symptom",
            "root_cause", "solution", "affected_files", "related_video_ids",
            "resolved_by",
        ]
        for field in allowed_fields:
            if field in payload:
                set_clauses.append(f"{field} = :{field}")
                params[field] = payload[field]

        if not set_clauses:
            raise HTTPException(status_code=400, detail="No fields to update")

        # Auto-set resolved_at when status changes to resolved
        if payload.get("status") == "resolved":
            set_clauses.append("resolved_at = NOW()")

        set_clauses.append("updated_at = NOW()")

        await db.execute(
            text(f"UPDATE bug_reports SET {', '.join(set_clauses)} WHERE id = :bug_id"),
            params,
        )
        await db.commit()
        return {"status": "ok", "id": bug_id}
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        logger.exception(f"Failed to update bug report: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ═══════════════════════════════════════════════════════════════════════
# WORK LOGS – デプロイ・修正・作業の履歴
# ═══════════════════════════════════════════════════════════════════════

@router.get("/work-logs")
async def list_work_logs(
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None),
    action: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
):
    """作業ログ一覧"""
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conditions = []
        params = {"limit": limit, "offset": offset}
        if action:
            conditions.append("action = :action")
            params["action"] = action

        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

        count_result = await db.execute(
            text(f"SELECT COUNT(*) FROM work_logs {where}"), params
        )
        total = count_result.scalar() or 0

        result = await db.execute(
            text(f"""
                SELECT * FROM work_logs {where}
                ORDER BY created_at DESC
                LIMIT :limit OFFSET :offset
            """),
            params,
        )
        rows = result.fetchall()
        logs = [dict(r._mapping) for r in rows]
        for l in logs:
            if l.get("created_at"):
                l["created_at"] = l["created_at"].isoformat()

        return {"total": total, "logs": logs}
    except Exception as e:
        logger.exception(f"Failed to fetch work logs: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/work-logs")
async def create_work_log(
    payload: dict,
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None),
):
    """作業ログ作成"""
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        result = await db.execute(
            text("""
                INSERT INTO work_logs
                    (action, summary, details, files_changed,
                     commit_hash, deployed_to, author, related_bug_id)
                VALUES
                    (:action, :summary, :details, :files_changed,
                     :commit_hash, :deployed_to, :author, :related_bug_id)
                RETURNING id
            """),
            {
                "action": payload.get("action", "other"),
                "summary": payload.get("summary", ""),
                "details": payload.get("details", ""),
                "files_changed": payload.get("files_changed", ""),
                "commit_hash": payload.get("commit_hash", ""),
                "deployed_to": payload.get("deployed_to", ""),
                "author": payload.get("author", "manus-ai"),
                "related_bug_id": payload.get("related_bug_id"),
            },
        )
        log_id = result.scalar()
        await db.commit()
        return {"status": "ok", "id": log_id}
    except Exception as e:
        await db.rollback()
        logger.exception(f"Failed to create work log: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ═══════════════════════════════════════════════════════════════════════
# AI CONTEXT ENDPOINT – AI読み取り用の構造化サマリー
# ═══════════════════════════════════════════════════════════════════════

@router.get("/ai-context")
async def get_ai_context(
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None),
    scope: Optional[str] = None,
):
    """
    AI（Manus）が毎回タスク開始時に読む唯一のエンドポイント。
    プロジェクトの永続記憶 + リアルタイム状態を2層構造で返す。

    クエリパラメータ:
      - scope: フィルタリング用。"aitherhub" or "liveboost" を指定すると
               そのプロジェクトに関連する教訓のみ返す。省略時は全て返す。

    第1層（常に返す・軽量）:
      - dangers: 絶対にやってはいけないこと
      - checklist_by_file: ファイル別の変更時チェックリスト
      - checklist_by_feature: 機能別の変更時チェックリスト
      - dependencies: ファイル間の依存マップ
      - rules: システムの正常状態の定義
      - preferences: ユーザーの方針
      - feature_status: 機能の現在の状態
      - open_bugs: 未解決のバグ
      - recent_errors: 直近24hのエラーサマリー
      - recent_work: 直近の作業ログ
      - error_videos: ERRORステータスの動画
      - stuck_videos: 停滞中の動画
      - action_required: 次のManusがやるべきこと
    """
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")

    context = {}
    if scope:
        context["scope"] = scope

    # ━━ プロジェクトの永続記憶（lessons_learned）━━
    try:
        # scopeフィルタ: related_featureでフィルタリング
        # - scope=liveboost → related_feature='liveboost' のみ + 共通(NULL/空)
        # - scope=aitherhub → related_feature!='liveboost' (既存の日本語feature名も含む)
        # - scope省略 → 全件
        scope_filter = ""
        scope_params = {}
        if scope == "liveboost":
            scope_filter = "AND (related_feature = 'liveboost' OR related_feature IS NULL OR related_feature = '')"
        elif scope == "aitherhub":
            scope_filter = "AND (related_feature != 'liveboost' OR related_feature IS NULL OR related_feature = '')"

        result = await db.execute(text(f"""
            SELECT id, category, title, content, related_files, related_feature
            FROM lessons_learned
            WHERE is_active = TRUE {scope_filter}
            ORDER BY
                CASE category
                    WHEN 'danger' THEN 0 WHEN 'checklist' THEN 1
                    WHEN 'rule' THEN 2 WHEN 'dependency' THEN 3
                    WHEN 'status' THEN 4 WHEN 'preference' THEN 5
                    ELSE 6 END,
                created_at DESC
        """), scope_params)
        rows = result.fetchall()

        # danger: 絶対にやってはいけないこと
        context["dangers"] = [
            r.title for r in rows if r.category == "danger"
        ]

        # checklist: ファイル別・機能別の変更時チェック
        checklist_by_file = {}
        checklist_by_feature = {}
        for r in rows:
            if r.category != "checklist":
                continue
            # ファイル別
            if r.related_files:
                for f in r.related_files.split(","):
                    f = f.strip()
                    if f:
                        checklist_by_file.setdefault(f, []).append(r.title)
            # 機能別
            if r.related_feature:
                checklist_by_feature.setdefault(r.related_feature.strip(), []).append(r.title)
        context["checklist_by_file"] = checklist_by_file
        context["checklist_by_feature"] = checklist_by_feature

        # dependency: ファイル間の依存マップ
        dep_map = {}
        for r in rows:
            if r.category != "dependency":
                continue
            # title = 起点ファイル, content = 依存先（カンマ区切り）
            if r.title and r.content:
                dep_map[r.title] = [x.strip() for x in r.content.split(",") if x.strip()]
        context["dependencies"] = dep_map

        # rule: システムの正常状態の定義
        context["rules"] = [
            {"title": r.title, "detail": r.content[:200]} for r in rows if r.category == "rule"
        ]

        # preference: ユーザーの方針
        context["preferences"] = [
            r.title for r in rows if r.category == "preference"
        ]

        # status: 機能の現在の状態
        context["feature_status"] = [
            {"feature": r.related_feature or r.title, "status": r.content[:100]}
            for r in rows if r.category == "status"
        ]

        # lesson: 過去の失敗パターン（タイトルのみ、詳細は第2層）
        context["lessons"] = [
            {"id": r.id, "title": r.title}
            for r in rows if r.category == "lesson"
        ]

    except Exception as e:
        context["lessons_error"] = f"error: {e}"

    # ━━ リアルタイム状態 ━━

    # 1. Open bugs
    try:
        result = await db.execute(text("""
            SELECT id, title, severity, category, symptom, root_cause, status, created_at
            FROM bug_reports
            WHERE status IN ('open', 'investigating')
            ORDER BY
                CASE severity WHEN 'critical' THEN 0 WHEN 'high' THEN 1
                              WHEN 'medium' THEN 2 ELSE 3 END,
                created_at DESC
            LIMIT 20
        """))
        rows = result.fetchall()
        context["open_bugs"] = [
            {
                "id": r.id, "title": r.title, "severity": r.severity,
                "category": r.category, "symptom": (r.symptom or "")[:200],
                "root_cause": (r.root_cause or "")[:200], "status": r.status,
            }
            for r in rows
        ]
    except Exception as e:
        context["open_bugs"] = f"error: {e}"

    # 2. Recent errors (24h summary)
    try:
        total_result = await db.execute(text(
            "SELECT COUNT(*) FROM video_error_logs WHERE created_at >= NOW() - INTERVAL '24 hours'"
        ))
        total = total_result.scalar() or 0

        top_codes = await db.execute(text("""
            SELECT error_code, COUNT(*) as cnt
            FROM video_error_logs WHERE created_at >= NOW() - INTERVAL '24 hours'
            GROUP BY error_code ORDER BY cnt DESC LIMIT 10
        """))
        context["recent_errors"] = {
            "total_24h": total,
            "top_codes": {r.error_code: r.cnt for r in top_codes.fetchall()},
        }
    except Exception as e:
        context["recent_errors"] = f"error: {e}"

    # 3. Recent work logs
    try:
        result = await db.execute(text("""
            SELECT id, action, summary, commit_hash, created_at
            FROM work_logs ORDER BY created_at DESC LIMIT 10
        """))
        rows = result.fetchall()
        context["recent_work"] = [
            {
                "id": r.id, "action": r.action,
                "summary": (r.summary or "")[:200],
                "commit": r.commit_hash or "",
                "at": r.created_at.isoformat() if r.created_at else "",
            }
            for r in rows
        ]
    except Exception as e:
        context["recent_work"] = f"error: {e}"

    # 4. Error videos
    try:
        result = await db.execute(text("""
            SELECT id, original_filename, status, last_error_code, last_error_message, updated_at
            FROM videos
            WHERE status IN ('ERROR', 'error')
            ORDER BY updated_at DESC
            LIMIT 20
        """))
        rows = result.fetchall()
        context["error_videos"] = [
            {
                "id": str(r.id),
                "file": r.original_filename or "",
                "error": r.last_error_code or "",
                "msg": (r.last_error_message or "")[:200],
            }
            for r in rows
        ]
    except Exception as e:
        context["error_videos"] = f"error: {e}"

    # 5. Stuck videos (uploaded/processing for > 30 min)
    #    Also detect videos stuck in STEP_* for > 6h (stalled pipeline)
    try:
        result = await db.execute(text("""
            SELECT id, original_filename, status, updated_at, created_at,
                   EXTRACT(EPOCH FROM (NOW() - updated_at)) / 3600.0 AS stall_hours
            FROM videos
            WHERE status NOT IN ('DONE', 'COMPLETED', 'ERROR', 'error', 'deleted')
              AND updated_at < NOW() - INTERVAL '30 minutes'
            ORDER BY updated_at ASC
            LIMIT 20
        """))
        rows = result.fetchall()
        context["stuck_videos"] = [
            {
                "id": str(r.id),
                "file": r.original_filename or "",
                "status": r.status or "",
                "since": r.updated_at.isoformat() if r.updated_at else "",
                "stall_hours": round(float(r.stall_hours), 1) if r.stall_hours else 0,
            }
            for r in rows
        ]
    except Exception as e:
        context["stuck_videos"] = f"error: {e}"

    # ━━ action_required: 次のManusがやるべきこと ━━
    action_required = []
    try:
        # 教訓登録忘れ検知: 直近のwork-logの後にlessonsが登録されているか
        result = await db.execute(text("""
            SELECT w.id, w.action, w.summary, w.created_at,
                   (SELECT COUNT(*) FROM lessons_learned l
                    WHERE l.created_at > w.created_at
                      AND l.created_at < w.created_at + INTERVAL '24 hours') as lessons_after
            FROM work_logs w
            ORDER BY w.created_at DESC
            LIMIT 3
        """))
        recent_logs = result.fetchall()
        for log in recent_logs:
            if log.lessons_after == 0 and log.action not in ('read', 'review', 'check'):
                action_required.append(
                    f"作業 '{log.summary[:80]}' (ID:{log.id}) の後に教訓が登録されていません。"
                    f"バグ修正・機能追加をした場合はlessonsに登録してください。"
                )
    except Exception:
        pass  # action_requiredは必須ではないのでエラーは無視

    # 未解決バグがあれば警告
    if context.get("open_bugs") and isinstance(context["open_bugs"], list) and len(context["open_bugs"]) > 0:
        critical = [b for b in context["open_bugs"] if b.get("severity") in ("critical", "high")]
        if critical:
            action_required.append(
                f"重要度の高い未解決バグが{len(critical)}件あります。作業前に確認してください。"
            )

    # stuck_videosがあれば警告
    if context.get("stuck_videos") and isinstance(context["stuck_videos"], list) and len(context["stuck_videos"]) > 0:
        stalled_critical = [v for v in context["stuck_videos"]
                            if isinstance(v, dict) and v.get("stall_hours", 0) >= 6]
        if stalled_critical:
            action_required.append(
                f"緊急: {len(stalled_critical)}件の動画が6時間以上停止しています。"
                f" admin retry-video APIで再投入してください。"
            )
        else:
            action_required.append(
                f"停滞中の動画が{len(context['stuck_videos'])}件あります。確認が必要かもしれません。"
            )

    context["action_required"] = action_required

    return context


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Lessons Learned — プロジェクトの永続記憶
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@router.get("/lessons")
async def list_lessons(
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None),
    category: Optional[str] = None,
    is_active: Optional[bool] = True,
    related_files: Optional[str] = None,
    related_feature: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
):
    """教訓一覧（カテゴリ・ファイル・機能でフィルタ可能）"""
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conditions = []
        params = {"limit": limit, "offset": offset}

        if is_active is not None:
            conditions.append("is_active = :is_active")
            params["is_active"] = is_active
        if category:
            conditions.append("category = :category")
            params["category"] = category
        if related_files:
            conditions.append("related_files ILIKE :related_files")
            params["related_files"] = f"%{related_files}%"
        if related_feature:
            conditions.append("related_feature ILIKE :related_feature")
            params["related_feature"] = f"%{related_feature}%"

        where = ("WHERE " + " AND ".join(conditions)) if conditions else ""

        count_result = await db.execute(
            text(f"SELECT COUNT(*) FROM lessons_learned {where}"), params
        )
        total = count_result.scalar() or 0

        result = await db.execute(
            text(f"""
                SELECT * FROM lessons_learned {where}
                ORDER BY
                    CASE category
                        WHEN 'danger' THEN 0 WHEN 'checklist' THEN 1
                        WHEN 'rule' THEN 2 WHEN 'dependency' THEN 3
                        WHEN 'lesson' THEN 4 WHEN 'status' THEN 5
                        ELSE 6 END,
                    created_at DESC
                LIMIT :limit OFFSET :offset
            """),
            params,
        )
        rows = result.fetchall()
        lessons = [dict(r._mapping) for r in rows]
        for l in lessons:
            for k in ("created_at", "updated_at"):
                if l.get(k):
                    l[k] = l[k].isoformat()

        return {"total": total, "lessons": lessons}
    except Exception as e:
        logger.exception(f"Failed to fetch lessons: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/lessons")
async def create_lesson(
    payload: dict,
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None),
):
    """教訓を新規作成"""
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        result = await db.execute(
            text("""
                INSERT INTO lessons_learned
                    (category, title, content, related_files, related_feature, source_bug_id)
                VALUES
                    (:category, :title, :content, :related_files, :related_feature, :source_bug_id)
                RETURNING id
            """),
            {
                "category": payload.get("category", "lesson"),
                "title": payload.get("title", ""),
                "content": payload.get("content", ""),
                "related_files": payload.get("related_files", ""),
                "related_feature": payload.get("related_feature", ""),
                "source_bug_id": payload.get("source_bug_id"),
            },
        )
        new_id = result.scalar()
        await db.commit()
        return {"status": "ok", "id": new_id}
    except Exception as e:
        await db.rollback()
        logger.exception(f"Failed to create lesson: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/lessons/{lesson_id}")
async def update_lesson(
    lesson_id: int,
    payload: dict,
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None),
):
    """教訓を更新"""
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        set_clauses = []
        params = {"id": lesson_id}
        for field in ("category", "title", "content", "related_files", "related_feature", "is_active", "source_bug_id"):
            if field in payload:
                set_clauses.append(f"{field} = :{field}")
                params[field] = payload[field]

        if not set_clauses:
            raise HTTPException(status_code=400, detail="No fields to update")

        set_clauses.append("updated_at = NOW()")
        set_sql = ", ".join(set_clauses)

        await db.execute(
            text(f"UPDATE lessons_learned SET {set_sql} WHERE id = :id"),
            params,
        )
        await db.commit()
        return {"status": "ok"}
    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        logger.exception(f"Failed to update lesson: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/lessons/{lesson_id}")
async def deactivate_lesson(
    lesson_id: int,
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None),
):
    """教訓を無効化（DBからは削除しない）"""
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        await db.execute(
            text("UPDATE lessons_learned SET is_active = FALSE, updated_at = NOW() WHERE id = :id"),
            {"id": lesson_id},
        )
        await db.commit()
        return {"status": "ok"}
    except Exception as e:
        await db.rollback()
        logger.exception(f"Failed to deactivate lesson: {e}")
        raise HTTPException(status_code=500, detail=str(e))



# ──────────────────────────────────────────────
# List blobs for a video (debug / admin)
# ──────────────────────────────────────────────
@router.get("/list-blobs/{video_id}")
async def list_blobs_for_video(
    video_id: str,
    email: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None),
):
    """List all blobs in Azure Storage for a given video_id.
    Useful for debugging chunk upload issues."""
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")
    try:
        from azure.storage.blob import BlobServiceClient
        import os

        conn_str = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
        container_name = os.getenv("AZURE_BLOB_CONTAINER", "videos")
        if not conn_str:
            raise HTTPException(status_code=500, detail="No storage connection string")

        # Get email — use query param if provided, otherwise look up from DB
        if not email:
            result = await db.execute(
                text("SELECT u.email FROM videos v JOIN users u ON v.user_id = u.id WHERE v.id = :vid"),
                {"vid": video_id},
            )
            row = result.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Video not found")
            email = row[0]
        prefix = f"{email}/{video_id}/"

        service_client = BlobServiceClient.from_connection_string(conn_str)
        container_client = service_client.get_container_client(container_name)

        blobs = []
        for blob in container_client.list_blobs(name_starts_with=prefix):
            blobs.append({
                "name": blob.name,
                "size": blob.size,
                "last_modified": str(blob.last_modified) if blob.last_modified else None,
                "content_type": blob.content_settings.content_type if blob.content_settings else None,
            })

        return {
            "video_id": video_id,
            "email": email,
            "prefix": prefix,
            "blob_count": len(blobs),
            "blobs": blobs,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Failed to list blobs: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ── BUILD 33: Admin generate upload SAS (for testing) ──────────


@router.post("/generate-upload-sas")
async def admin_generate_upload_sas(
    payload: dict,
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
):
    """Generate a write SAS URL for uploading a blob (admin/testing use)."""
    import os
    from app.services.storage_service import generate_upload_sas

    expected_key = os.getenv("ADMIN_API_KEY", "aither:hub")
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Forbidden")

    email = payload.get("email", "")
    video_id = payload.get("video_id", "")
    filename = payload.get("filename", "")

    if not email or not video_id or not filename:
        raise HTTPException(status_code=400, detail="email, video_id, filename required")

    vid, upload_url, blob_url, expiry = await generate_upload_sas(
        email=email, video_id=video_id, filename=filename,
    )
    return {
        "video_id": vid,
        "upload_url": upload_url,
        "blob_url": blob_url,
        "expires_at": expiry.isoformat(),
    }


# ── BUILD 42: Admin generate read SAS (for debugging video playback) ──────────
@router.post("/generate-read-sas")
async def admin_generate_read_sas(
    payload: dict,
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
):
    """Generate a read-only SAS URL for downloading/viewing a blob (admin/debug use)."""
    import os
    from azure.storage.blob import generate_blob_sas, BlobSasPermissions
    expected_key = os.getenv("ADMIN_API_KEY", "aither:hub")
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Forbidden")
    blob_name = payload.get("blob_name", "")
    if not blob_name:
        raise HTTPException(status_code=400, detail="blob_name required")
    conn_str = os.getenv("AZURE_STORAGE_CONNECTION_STRING", "")
    account_name = os.getenv("AZURE_STORAGE_ACCOUNT_NAME", "")
    container_name = os.getenv("AZURE_BLOB_CONTAINER", "videos")
    account_key = ""
    for part in conn_str.split(";"):
        if part.startswith("AccountKey="):
            account_key = part.split("=", 1)[1]
            break
    if not account_key:
        raise HTTPException(status_code=500, detail="No account key found")
    from datetime import datetime, timezone, timedelta
    expiry = datetime.now(timezone.utc) + timedelta(hours=1)
    sas = generate_blob_sas(
        account_name=account_name,
        container_name=container_name,
        blob_name=blob_name,
        account_key=account_key,
        permission=BlobSasPermissions(read=True),
        expiry=expiry,
    )
    url = f"https://{account_name}.blob.core.windows.net/{container_name}/{blob_name}?{sas}"
    return {
        "blob_name": blob_name,
        "read_url": url,
        "expires_at": expiry.isoformat(),
    }



# =========================
# BULK RETRY PRODUCT DETECTION
# =========================
@router.post("/bulk-retry-product-detection")
async def bulk_retry_product_detection(
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
):
    """
    Find all DONE videos with PRODUCT_DATA_MISMATCH errors and re-enqueue them
    from STEP_12_5_PRODUCT_DETECTION.
    """
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        # Find affected videos
        sql = text("""
            SELECT DISTINCT vel.video_id, v.original_filename, v.user_id, u.email
            FROM video_error_logs vel
            JOIN videos v ON v.id = vel.video_id::uuid
            LEFT JOIN users u ON v.user_id = u.id
            WHERE vel.error_code = 'PRODUCT_DATA_MISMATCH'
              AND v.status = 'DONE'
              AND NOT EXISTS (
                  SELECT 1 FROM video_product_exposures vpe
                  WHERE vpe.video_id = vel.video_id::uuid
              )
        """)
        result = await db.execute(sql)
        rows = result.fetchall()

        if not rows:
            return {"status": "ok", "message": "No affected videos found", "count": 0}

        from app.services.storage_service import generate_download_sas
        from app.services.queue_service import enqueue_job

        enqueued = []
        failed = []
        for row in rows:
            try:
                vid = str(row.video_id)
                download_url, expiry = await generate_download_sas(
                    email=row.email,
                    video_id=vid,
                    filename=row.original_filename,
                    expires_in_minutes=1440,
                )
                await db.execute(
                    text("UPDATE videos SET status = 'STEP_12_5_PRODUCT_DETECTION', step_progress = 0 WHERE id = :vid"),
                    {"vid": vid},
                )
                await db.commit()
                await enqueue_job({
                    "video_id": vid,
                    "blob_url": download_url,
                    "original_filename": row.original_filename,
                })
                enqueued.append({"video_id": vid, "filename": row.original_filename})
            except Exception as e:
                failed.append({"video_id": str(row.video_id), "error": str(e)})

        return {
            "status": "ok",
            "enqueued": len(enqueued),
            "failed": len(failed),
            "videos": enqueued,
            "errors": failed,
        }
    except Exception as e:
        logger.exception(f"Failed to bulk retry product detection: {e}")
        raise HTTPException(status_code=500, detail=str(e))



# =========================================================
# Regenerate Product Exposures (backend-only, no worker needed)
# =========================================================

@router.post("/regenerate-product-exposures/{video_id}")
async def regenerate_product_exposures(
    video_id: str,
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
):
    """
    Regenerate product exposure timeline for a video without requiring
    the full worker pipeline. Downloads Excel data from blob storage,
    parses product list and trend data, and generates product exposures
    based on sales data timestamps.

    This is a lightweight alternative to re-running the entire video
    processing pipeline when only product exposures need to be regenerated.
    """
    import os
    import tempfile
    import httpx
    import re
    from datetime import datetime, timezone, timedelta

    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")

    try:
        # ── A. Get video info ──
        result = await db.execute(
            text("""
                SELECT v.id, v.user_id, v.excel_product_blob_url, v.excel_trend_blob_url,
                       v.time_offset_seconds, v.duration, v.top_products, v.status
                FROM videos v WHERE v.id = :vid
            """),
            {"vid": video_id},
        )
        video = result.fetchone()
        if not video:
            raise HTTPException(status_code=404, detail="Video not found")

        user_id = video.user_id
        time_offset = float(video.time_offset_seconds or 0)
        duration_sec = float(video.duration or 0)
        # Fallback: get duration from video_phases if not set
        if duration_sec <= 0:
            dur_result = await db.execute(
                text("SELECT COALESCE(MAX(time_end), 0) FROM video_phases WHERE video_id = :vid"),
                {"vid": video_id},
            )
            dur_row = dur_result.fetchone()
            if dur_row and dur_row[0]:
                duration_sec = float(dur_row[0])

        if not video.excel_product_blob_url:
            raise HTTPException(status_code=400, detail="No Excel product data URL")

        # ── B. Generate fresh SAS URLs for Excel files ──
        from app.services.storage_service import generate_read_sas_from_url
        product_url = generate_read_sas_from_url(video.excel_product_blob_url)
        trend_url = None
        if video.excel_trend_blob_url:
            trend_url = generate_read_sas_from_url(video.excel_trend_blob_url)

        if not product_url:
            raise HTTPException(status_code=500, detail="Failed to generate SAS for product Excel")

        # ── C. Download and parse Excel files ──
        import openpyxl

        def _parse_excel_from_url(url: str) -> list[dict]:
            """Download Excel from URL and parse rows into dicts."""
            resp = httpx.get(url, timeout=30, follow_redirects=True)
            resp.raise_for_status()
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(resp.content)
                tmp_path = tmp.name
            try:
                wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
                ws = wb.active
                if ws is None:
                    return []
                rows = list(ws.iter_rows(values_only=True))
                if len(rows) < 2:
                    return []
                headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
                data = []
                for row in rows[1:]:
                    if all(v is None for v in row):
                        continue
                    entry = {}
                    for i, val in enumerate(row):
                        if i < len(headers):
                            entry[headers[i]] = val
                    data.append(entry)
                wb.close()
                return data
            finally:
                os.unlink(tmp_path)

        products = _parse_excel_from_url(product_url)
        trends = _parse_excel_from_url(trend_url) if trend_url else []

        if not products:
            raise HTTPException(status_code=400, detail="No products found in Excel")

        # ── D. Build product list with name matching ──
        # Find the product name column
        PRODUCT_NAME_ALIASES = ["商品名", "product_name", "name", "商品", "品名", "アイテム名"]
        BRAND_ALIASES = ["ブランド", "brand", "brand_name", "メーカー"]
        IMAGE_ALIASES = ["画像", "image", "image_url", "商品画像"]

        def _find_key(sample: dict, aliases: list) -> str | None:
            for a in aliases:
                for k in sample.keys():
                    if a.lower() in k.lower():
                        return k
            return None

        sample_product = products[0] if products else {}
        pname_key = _find_key(sample_product, PRODUCT_NAME_ALIASES)
        brand_key = _find_key(sample_product, BRAND_ALIASES)
        image_key = _find_key(sample_product, IMAGE_ALIASES)

        product_list = []
        for p in products:
            name = str(p.get(pname_key, "")).strip() if pname_key else ""
            if not name:
                continue
            product_list.append({
                "product_name": name,
                "brand_name": str(p.get(brand_key, "")).strip() if brand_key else "",
                "image_url": str(p.get(image_key, "")).strip() if image_key else "",
            })

        if not product_list:
            raise HTTPException(status_code=400, detail="No valid products found in Excel")

        # ── E. Parse trend data for sales-based detection ──
        TIME_ALIASES = ["時間", "time", "timestamp", "配信時間", "日時"]
        ORDER_ALIASES = ["注文数", "orders", "order_count", "注文", "成約数"]
        GMV_ALIASES = ["売上", "gmv", "revenue", "売上金額", "成約金額"]
        TREND_PRODUCT_ALIASES = ["商品名", "product_name", "商品", "品名"]

        exposures = []
        audio_exposures = []
        slot_interval = 60  # default

        # ── E-1. Load audio transcription from DB (video_phases.audio_text) ──
        phase_rows = await db.execute(
            text(
                "SELECT phase_index, time_start, time_end, audio_text "
                "FROM video_phases "
                "WHERE video_id = :vid AND audio_text IS NOT NULL AND audio_text != '' "
                "ORDER BY phase_index"
            ),
            {"vid": video_id},
        )
        phases_with_audio = phase_rows.fetchall()
        logger.info("[REGEN] Loaded %d phases with audio_text from DB", len(phases_with_audio))

        # Build product keyword map for matching
        product_keywords = {}  # product_name -> [keywords]
        for pl in product_list:
            name = pl["product_name"]
            keywords = set()
            # Add full name (lowercase)
            keywords.add(name.lower())
            # Add individual words (3+ chars)
            for w in name.split():
                w_clean = w.strip().lower()
                skip_words = {'kyogoku', 'the', 'and', 'for', 'pro', '\u7528', '\u5f0f', '\u578b'}
                if len(w_clean) >= 3 and w_clean not in skip_words:
                    keywords.add(w_clean)
            # Add katakana/hiragana words
            import unicodedata
            kana_word = ""
            for ch in name:
                cat = unicodedata.category(ch)
                if cat.startswith('Lo'):  # Letter, other (CJK, kana)
                    kana_word += ch
                else:
                    if len(kana_word) >= 2:
                        keywords.add(kana_word.lower())
                    kana_word = ""
            if len(kana_word) >= 2:
                keywords.add(kana_word.lower())
            product_keywords[name] = list(keywords)

        # ── E-2. Audio-based product detection ──
        if phases_with_audio:
            for row in phases_with_audio:
                t_start = float(row[1]) if row[1] else 0
                t_end = float(row[2]) if row[2] else 0
                audio_text = str(row[3]).strip().lower()
                if not audio_text or t_end <= t_start:
                    continue

                # Search for product keywords in audio text
                for pname, keywords in product_keywords.items():
                    matched = False
                    for kw in keywords:
                        if len(kw) >= 3 and kw in audio_text:
                            matched = True
                            break
                    if matched:
                        audio_exposures.append({
                            "product_name": pname,
                            "brand_name": next((pl.get("brand_name", "") for pl in product_list if pl["product_name"] == pname), ""),
                            "product_image_url": next((pl.get("image_url", "") for pl in product_list if pl["product_name"] == pname), ""),
                            "time_start": t_start,
                            "time_end": t_end,
                            "confidence": 0.80,
                            "source": "audio",
                        })
            logger.info("[REGEN] Audio detection: %d exposures", len(audio_exposures))

        # ── E-3. Sales-based product detection from trend data ──
        def _parse_time_to_seconds(val) -> float | None:
            """Parse various time formats to seconds."""
            if val is None:
                return None
            s = str(val).strip()
            if not s:
                return None
            m = re.match(r'^(\d{1,2}):(\d{2}):(\d{2})$', s)
            if m:
                return int(m.group(1)) * 3600 + int(m.group(2)) * 60 + int(m.group(3))
            m = re.match(r'^(\d{1,2}):(\d{2})$', s)
            if m:
                return int(m.group(1)) * 3600 + int(m.group(2)) * 60
            try:
                return float(s)
            except ValueError:
                return None

        sales_exposures = []
        if trends:
            sample_trend = trends[0]
            time_key = _find_key(sample_trend, TIME_ALIASES)
            order_key = _find_key(sample_trend, ORDER_ALIASES)
            gmv_key = _find_key(sample_trend, GMV_ALIASES)
            trend_product_key = _find_key(sample_trend, TREND_PRODUCT_ALIASES)

            if time_key:
                slot_times = []
                for entry in trends:
                    t = _parse_time_to_seconds(entry.get(time_key))
                    if t is not None:
                        slot_times.append(t)
                slot_times.sort()

                if len(slot_times) >= 2:
                    intervals = [slot_times[i+1] - slot_times[i] for i in range(len(slot_times)-1)]
                    intervals = [iv for iv in intervals if iv > 0]
                    if intervals:
                        from collections import Counter
                        interval_counts = Counter(int(iv) for iv in intervals)
                        slot_interval = interval_counts.most_common(1)[0][0]
                        if slot_interval < 10:
                            slot_interval = 60

                for entry in trends:
                    t_sec = _parse_time_to_seconds(entry.get(time_key))
                    if t_sec is None:
                        continue

                    video_time = t_sec - time_offset
                    if video_time < 0:
                        continue

                    orders = 0
                    if order_key:
                        try:
                            orders = int(float(entry.get(order_key, 0) or 0))
                        except (ValueError, TypeError):
                            orders = 0

                    gmv = 0
                    if gmv_key:
                        try:
                            gmv = float(entry.get(gmv_key, 0) or 0)
                        except (ValueError, TypeError):
                            gmv = 0

                    if orders <= 0 and gmv <= 0:
                        continue

                    trend_product = ""
                    if trend_product_key:
                        trend_product = str(entry.get(trend_product_key, "")).strip()

                    matched = None
                    if trend_product:
                        tp_lower = trend_product.lower()
                        for pl in product_list:
                            pl_lower = pl["product_name"].lower()
                            if tp_lower in pl_lower or pl_lower in tp_lower:
                                matched = pl
                                break
                        if not matched:
                            for pl in product_list:
                                words = [w for w in pl["product_name"].split() if len(w) >= 3]
                                for w in words:
                                    if w.lower() in tp_lower:
                                        matched = pl
                                        break
                                if matched:
                                    break

                    if not matched and len(product_list) == 1:
                        matched = product_list[0]

                    if not matched:
                        continue

                    lookback = min(slot_interval * 2, 300)
                    sales_exposures.append({
                        "product_name": matched["product_name"],
                        "brand_name": matched.get("brand_name", ""),
                        "product_image_url": matched.get("image_url", ""),
                        "time_start": max(0, video_time - lookback),
                        "time_end": video_time + slot_interval,
                        "confidence": min(0.85, 0.60 + min(orders, 5) * 0.05),
                        "source": "sales",
                    })
            logger.info("[REGEN] Sales detection: %d exposures", len(sales_exposures))

        # Combine audio + sales exposures
        exposures = audio_exposures + sales_exposures

        # ── F. Merge overlapping exposures for same product ──
        if exposures:
            by_product = {}
            for exp in exposures:
                name = exp["product_name"]
                if name not in by_product:
                    by_product[name] = []
                by_product[name].append(exp)

            merged = []
            for name, exps in by_product.items():
                exps.sort(key=lambda x: x["time_start"])
                current = exps[0].copy()
                for i in range(1, len(exps)):
                    if exps[i]["time_start"] <= current["time_end"] + slot_interval:
                        current["time_end"] = max(current["time_end"], exps[i]["time_end"])
                        current["confidence"] = max(current["confidence"], exps[i]["confidence"])
                    else:
                        merged.append(current)
                        current = exps[i].copy()
                merged.append(current)
            exposures = merged

        # ── G. If no exposures at all, create from top_products or first N products ──
        if not exposures and product_list:
            # Use top_products if available, otherwise first 10 products
            import json as _json
            top_product_names = []
            if video.top_products:
                try:
                    tp = _json.loads(video.top_products) if isinstance(video.top_products, str) else video.top_products
                    if isinstance(tp, list):
                        top_product_names = tp
                except Exception:
                    pass

            # Match top_products to product_list
            target_products = []
            if top_product_names:
                for tp_name in top_product_names:
                    tp_lower = tp_name.lower()
                    for pl in product_list:
                        if pl["product_name"].lower() in tp_lower or tp_lower in pl["product_name"].lower():
                            target_products.append(pl)
                            break
            if not target_products:
                # Sort by GMV if available, otherwise take first 10
                target_products = product_list[:min(10, len(product_list))]

            if duration_sec > 0 and target_products:
                segment_duration = duration_sec / max(len(target_products), 1)
                for i, pl in enumerate(target_products):
                    exposures.append({
                        "product_name": pl["product_name"],
                        "brand_name": pl.get("brand_name", ""),
                        "product_image_url": pl.get("image_url", ""),
                        "time_start": i * segment_duration,
                        "time_end": (i + 1) * segment_duration,
                        "confidence": 0.5,
                    })

        # ── H. Ensure table exists ──
        try:
            await db.execute(text("""
                CREATE TABLE IF NOT EXISTS video_product_exposures (
                    id UUID DEFAULT gen_random_uuid() PRIMARY KEY,
                    video_id UUID NOT NULL,
                    user_id INTEGER,
                    product_name TEXT NOT NULL,
                    brand_name TEXT,
                    product_image_url TEXT,
                    time_start FLOAT NOT NULL,
                    time_end FLOAT NOT NULL,
                    confidence FLOAT DEFAULT 0.8,
                    source VARCHAR(20) DEFAULT 'ai',
                    created_at TIMESTAMPTZ DEFAULT now(),
                    updated_at TIMESTAMPTZ DEFAULT now()
                )
            """))
            await db.commit()
        except Exception:
            await db.rollback()

        # ── I. Delete existing AI-generated exposures and insert new ones ──
        await db.execute(
            text("DELETE FROM video_product_exposures WHERE video_id = :vid AND source = 'ai'"),
            {"vid": video_id},
        )

        inserted = 0
        for exp in exposures:
            await db.execute(
                text("""
                    INSERT INTO video_product_exposures
                        (video_id, user_id, product_name, brand_name,
                         product_image_url, time_start, time_end, confidence, source)
                    VALUES
                        (:video_id, :user_id, :product_name, :brand_name,
                         :product_image_url, :time_start, :time_end, :confidence, 'ai')
                """),
                {
                    "video_id": video_id,
                    "user_id": user_id,
                    "product_name": exp.get("product_name", ""),
                    "brand_name": exp.get("brand_name", ""),
                    "product_image_url": exp.get("product_image_url", ""),
                    "time_start": exp.get("time_start", 0),
                    "time_end": exp.get("time_end", 0),
                    "confidence": exp.get("confidence", 0.8),
                },
            )
            inserted += 1

        await db.commit()

        # ── J. Update video status to DONE if it was ERROR ──
        if video.status == "ERROR":
            await db.execute(
                text("UPDATE videos SET status = 'DONE' WHERE id = :vid"),
                {"vid": video_id},
            )
            await db.commit()

        return {
            "status": "ok",
            "video_id": video_id,
            "products_found": len(product_list),
            "trend_entries": len(trends),
            "audio_exposures": len(audio_exposures),
            "sales_exposures": len(sales_exposures) if 'sales_exposures' in dir() else 0,
            "exposures_generated": inserted,
            "exposures": [
                {
                    "product_name": e["product_name"],
                    "time_start": e["time_start"],
                    "time_end": e["time_end"],
                    "confidence": e["confidence"],
                }
                for e in exposures
            ],
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Failed to regenerate product exposures: {e}")
        raise HTTPException(status_code=500, detail=str(e))



# ──────────────────────────────────────────────
# Admin: Reset User Password
# ──────────────────────────────────────────────

@router.post("/reset-user-password")
async def admin_reset_user_password(
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
    email: str = Query(...),
    new_password: str = Query(...),
):
    """Admin endpoint to reset a user's password."""
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")

    from app.repository.auth_repo import get_user_by_email
    from app.utils.password import hash_password

    user = await get_user_by_email(db, email)
    if not user:
        raise HTTPException(status_code=404, detail=f"User {email} not found")

    user.hashed_password = hash_password(new_password)
    await db.commit()

    return {"success": True, "email": email, "message": "Password reset successfully"}


# ──────────────────────────────────────────────
# Admin: Winning Patterns Analysis
# ──────────────────────────────────────────────

@router.get("/videos/{video_id}/winning-patterns")
async def get_winning_patterns(
    video_id: str,
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
):
    """
    Extract winning patterns from a video's real performance data.
    """
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")
    import traceback as tb_mod
    results = {"video_id": video_id, "steps": {}}

    # Step 1: Test basic DB query
    try:
        r = await db.execute(text(
            "SELECT id, status FROM videos WHERE id = :vid"
        ), {"vid": video_id})
        row = r.first()
        results["steps"]["db_check"] = {"ok": True, "status": row[1] if row else "NOT_FOUND"}
    except Exception as e:
        results["steps"]["db_check"] = {"ok": False, "error": str(e), "tb": tb_mod.format_exc()[-500:]}
        return results

    # Step 2: Test CTA extraction
    try:
        from app.services.winning_patterns_service import extract_cta_phrases
        cta = await extract_cta_phrases(db, video_id)
        results["steps"]["cta"] = {"ok": True, "count": len(cta)}
        results["cta_phrases"] = cta
    except Exception as e:
        results["steps"]["cta"] = {"ok": False, "error": str(e), "tb": tb_mod.format_exc()[-500:]}

    # Step 3: Test product durations
    try:
        from app.services.winning_patterns_service import analyze_product_durations
        dur = await analyze_product_durations(db, video_id)
        results["steps"]["durations"] = {"ok": True, "count": len(dur)}
        results["product_durations"] = dur
    except Exception as e:
        results["steps"]["durations"] = {"ok": False, "error": str(e), "tb": tb_mod.format_exc()[-500:]}

    # Step 4: Test top phases
    try:
        from app.services.winning_patterns_service import extract_top_phases
        phases = await extract_top_phases(db, video_id, limit=10)
        results["steps"]["phases"] = {"ok": True, "count": len(phases)}
        results["top_phases"] = phases
    except Exception as e:
        results["steps"]["phases"] = {"ok": False, "error": str(e), "tb": tb_mod.format_exc()[-500:]}

    return results


@router.get("/winning-patterns/aggregate")
async def get_aggregate_patterns(
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
    limit_videos: int = Query(50, ge=1, le=200),
):
    """
    Aggregate winning patterns across all DONE videos.

    This is the core differentiator — patterns from real sales data
    across many livestreams.
    """
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")
    from app.services.winning_patterns_service import aggregate_patterns_across_videos

    try:
        patterns = await aggregate_patterns_across_videos(db, limit_videos=limit_videos)
        return patterns
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ──────────────────────────────────────────────
# Admin: Data-Driven Script Generation
# ──────────────────────────────────────────────

class DataDrivenScriptRequest(BaseModel):
    product_focus: Optional[str] = None
    tone: str = "professional_friendly"
    language: str = "ja"
    duration_minutes: int = 10
    cross_video: bool = True


@router.post("/videos/{video_id}/generate-data-script")
async def generate_data_driven_script_endpoint(
    video_id: str,
    body: DataDrivenScriptRequest,
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
):
    """
    Generate a script grounded in real performance data.

    Unlike generic AI script generators, every CTA, product description timing,
    and engagement hook in this script is backed by actual sales metrics
    from past livestreams.
    """
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")
    from app.services.winning_patterns_service import generate_data_driven_script

    try:
        result = await generate_data_driven_script(
            db=db,
            video_id=video_id,
            product_focus=body.product_focus,
            tone=body.tone,
            language=body.language,
            duration_minutes=body.duration_minutes,
            cross_video=body.cross_video,
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.exception(f"Data-driven script generation failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# Batch Retry Stuck Videos API (施策1: 一括再投入)
# =============================================================================
@router.post("/batch-retry-stuck")
async def batch_retry_stuck_videos(
    x_admin_key: str = Header(...),
    db: AsyncSession = Depends(get_db),
    max_videos: int = Query(50, ge=1, le=200),
    threshold_minutes: int = Query(120, ge=10, le=1440),
):
    """Batch retry all stuck videos in one API call.
    Detects videos stuck at STEP_* or 'uploaded' status for longer than
    threshold_minutes, generates fresh SAS URLs, and re-enqueues them.
    Also retries ERROR videos that were never enqueued.
    Returns summary of results."""
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")
    from app.services.storage_service import generate_download_sas
    from app.services.queue_service import enqueue_job
    from datetime import datetime, timezone, timedelta

    # Use naive datetime to avoid asyncpg offset-naive vs offset-aware mismatch
    # DB columns may be stored as timestamp without time zone
    threshold = datetime.utcnow() - timedelta(minutes=threshold_minutes)
    results = {"total_found": 0, "retried": 0, "failed": 0, "skipped": 0, "details": []}

    try:
        # Part 1: Stuck processing videos (STEP_* or uploaded, not updated recently)
        sql_stuck = text("""
            SELECT v.id, v.original_filename, v.status, v.user_id,
                   v.dequeue_count, v.upload_type,
                   u.email as user_email
            FROM videos v
            LEFT JOIN users u ON v.user_id = u.id
            WHERE (v.status IN ('uploaded', 'QUEUED') OR v.status LIKE 'STEP_%%')
              AND v.status != 'completed'
              AND v.updated_at < :threshold
              AND COALESCE(v.dequeue_count, 0) < 10
            ORDER BY v.updated_at ASC
            LIMIT :max_videos
        """)
        result = await db.execute(sql_stuck, {
            "threshold": threshold,
            "max_videos": max_videos,
        })
        stuck_rows = result.fetchall()

        # Part 2: Never-enqueued ERROR videos
        sql_never_enqueued = text("""
            SELECT v.id, v.original_filename, v.status, v.user_id,
                   v.dequeue_count, v.upload_type,
                   u.email as user_email
            FROM videos v
            LEFT JOIN users u ON v.user_id = u.id
            WHERE v.status = 'ERROR'
              AND (v.enqueue_status = 'FAILED' OR v.enqueue_status IS NULL)
              AND v.worker_claimed_at IS NULL
              AND COALESCE(v.dequeue_count, 0) < 10
            ORDER BY v.created_at ASC
            LIMIT :max_videos
        """)
        result2 = await db.execute(sql_never_enqueued, {"max_videos": max_videos})
        never_enqueued_rows = result2.fetchall()

        all_rows = list(stuck_rows) + list(never_enqueued_rows)
        # Deduplicate by video_id
        seen_ids = set()
        unique_rows = []
        for row in all_rows:
            vid = str(row.id)
            if vid not in seen_ids:
                seen_ids.add(vid)
                unique_rows.append(row)
        results["total_found"] = len(unique_rows)

        for row in unique_rows:
            video_id = str(row.id)
            upload_type = row.upload_type or ""
            detail = {"video_id": video_id, "filename": row.original_filename, "old_status": row.status}

            # Skip live_boost videos (different pipeline)
            if upload_type == "live_boost":
                detail["result"] = "skipped_live_boost"
                results["skipped"] += 1
                results["details"].append(detail)
                continue

            try:
                # Generate fresh SAS URL
                download_url, expiry = await generate_download_sas(
                    email=row.user_email or "unknown@unknown.com",
                    video_id=video_id,
                    filename=row.original_filename,
                    expires_in_minutes=1440,
                )
                # Reset status to STEP_0 (NEVER use 'uploaded' as fallback per danger rules)
                resume_status = row.status if row.status.startswith("STEP_") else "STEP_0_EXTRACT_FRAMES"
                await db.execute(
                    text("""
                        UPDATE videos
                        SET status = :status,
                            step_progress = 0,
                            error_message = NULL,
                            enqueue_status = NULL,
                            enqueue_error = NULL,
                            last_error_code = NULL,
                            last_error_message = NULL,
                            worker_claimed_at = NULL,
                            dequeue_count = COALESCE(dequeue_count, 0) + 1,
                            updated_at = NOW()
                        WHERE id = :vid
                    """),
                    {"vid": video_id, "status": resume_status},
                )
                await db.commit()

                # Enqueue job
                enqueue_result = await enqueue_job({
                    "video_id": video_id,
                    "blob_url": download_url,
                    "original_filename": row.original_filename,
                })
                if enqueue_result.success:
                    # Update enqueue evidence
                    try:
                        await db.execute(
                            text("""
                                UPDATE videos
                                SET enqueue_status = 'OK',
                                    queue_message_id = :msg_id,
                                    queue_enqueued_at = :enqueued_at,
                                    enqueue_error = NULL
                                WHERE id = :vid
                            """),
                            {
                                "vid": video_id,
                                "msg_id": enqueue_result.message_id,
                                "enqueued_at": enqueue_result.enqueued_at,
                            },
                        )
                        await db.commit()
                    except Exception:
                        try:
                            await db.rollback()
                        except Exception:
                            pass
                    detail["result"] = "retried_ok"
                    detail["resume_from"] = resume_status
                    results["retried"] += 1
                else:
                    detail["result"] = f"enqueue_failed: {enqueue_result.error}"
                    results["failed"] += 1
            except Exception as e:
                detail["result"] = f"error: {str(e)[:200]}"
                results["failed"] += 1
                try:
                    await db.rollback()
                except Exception:
                    pass
            results["details"].append(detail)

        # Record to video_error_logs for observability
        try:
            await db.execute(
                text("""
                    INSERT INTO video_error_logs
                        (video_id, error_code, error_step, error_message, source)
                    VALUES
                        ('00000000-0000-0000-0000-000000000000', 'BATCH_RETRY',
                         'ADMIN_API', :msg, 'admin')
                """),
                {"msg": f"Batch retry: found={results['total_found']} retried={results['retried']} failed={results['failed']} skipped={results['skipped']}"},
            )
            await db.commit()
        except Exception:
            try:
                await db.rollback()
            except Exception:
                pass

        return results
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Batch retry stuck videos failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# =============================================================================
# Monitor Health API (施策2: monitor健全性確認)
# =============================================================================
@router.get("/monitor-health")
async def get_monitor_health(
    x_admin_key: str = Header(...),
    db: AsyncSession = Depends(get_db),
):
    """Get stuck video monitor health status.
    Returns the last N health log entries so admin can verify the monitor is running."""
    if x_admin_key != f"{ADMIN_ID}:{ADMIN_PASS}":
        raise HTTPException(status_code=403, detail="Forbidden")
    try:
        result = await db.execute(
            text("""
                SELECT * FROM monitor_health_logs
                ORDER BY checked_at DESC
                LIMIT 20
            """)
        )
        rows = result.fetchall()
        logs = []
        for row in rows:
            logs.append({
                "id": row.id,
                "checked_at": str(row.checked_at),
                "stuck_found": row.stuck_found,
                "stuck_retried": row.stuck_retried,
                "never_enqueued_found": row.never_enqueued_found,
                "never_enqueued_retried": row.never_enqueued_retried,
                "deploy_interrupted_found": row.deploy_interrupted_found,
                "deploy_interrupted_retried": row.deploy_interrupted_retried,
                "errors": row.errors,
            })
        return {"monitor_health_logs": logs}
    except Exception as e:
        # Table might not exist yet
        return {"monitor_health_logs": [], "error": str(e)}


# ── Temporary: outbound IP check (for Shopee IP whitelist) ──
@router.get("/outbound-ip")
async def get_outbound_ip():
    """Fetch this server's outbound IP by calling httpbin.org."""
    import httpx
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.get("https://httpbin.org/ip")
            data = resp.json()
            return {"outbound_ip": data.get("origin", "unknown"), "source": "httpbin.org"}
    except Exception as e:
        return {"error": str(e)}


# =============================================================================
# Worker VM Health Monitoring & Auto-Restart
# =============================================================================

@router.get("/worker-vm/status")
async def get_worker_vm_status(
    db: AsyncSession = Depends(get_db),
):
    """Check Worker VM status from DB heartbeats."""
    from datetime import datetime, timezone

    try:
        result = await db.execute(text("""
            SELECT worker_id, last_heartbeat, mem_total_gb, mem_used_gb, mem_pct,
                   load_1m, load_5m, disk_total_gb, disk_free_gb, disk_pct, status
            FROM worker_heartbeats
            ORDER BY last_heartbeat DESC
            LIMIT 5
        """))
        rows = result.fetchall()

        if not rows:
            return {
                "status": "unknown",
                "message": "No heartbeat data found. Worker may not have started yet.",
                "workers": [],
            }

        now = datetime.now(timezone.utc)
        workers = []
        for row in rows:
            age_seconds = (now - row.last_heartbeat).total_seconds()
            is_alive = age_seconds < 300
            workers.append({
                "worker_id": row.worker_id,
                "last_heartbeat": row.last_heartbeat.isoformat(),
                "age_seconds": round(age_seconds, 1),
                "is_alive": is_alive,
                "status": row.status,
                "mem_total_gb": row.mem_total_gb,
                "mem_used_gb": row.mem_used_gb,
                "mem_pct": row.mem_pct,
                "load_1m": row.load_1m,
                "load_5m": row.load_5m,
                "disk_total_gb": row.disk_total_gb,
                "disk_free_gb": row.disk_free_gb,
                "disk_pct": row.disk_pct,
            })

        any_alive = any(w["is_alive"] for w in workers)
        return {
            "status": "healthy" if any_alive else "dead",
            "message": "Worker is running" if any_alive else "Worker heartbeat stale (>5min). VM may be down.",
            "workers": workers,
        }
    except Exception as e:
        return {"status": "error", "message": f"Could not check worker status: {e}", "workers": []}


@router.post("/worker-vm/restart")
async def restart_worker_vm():
    """Restart the Worker VM via Azure REST API.
    Requires AZURE_SUBSCRIPTION_ID, AZURE_RESOURCE_GROUP, AZURE_VM_NAME env vars.
    Uses DefaultAzureCredential (Managed Identity on App Service)."""
    import os

    subscription_id = os.getenv("AZURE_SUBSCRIPTION_ID")
    resource_group = os.getenv("AZURE_RESOURCE_GROUP")
    vm_name = os.getenv("AZURE_VM_NAME", "aitherhubVm")

    if not subscription_id or not resource_group:
        return {
            "status": "error",
            "message": "AZURE_SUBSCRIPTION_ID and AZURE_RESOURCE_GROUP must be set. "
                       "Also ensure the App Service has a Managed Identity with "
                       "Virtual Machine Contributor role on the VM.",
        }

    try:
        from azure.identity import DefaultAzureCredential
        import httpx

        credential = DefaultAzureCredential()
        token = credential.get_token("https://management.azure.com/.default")

        url = (
            f"https://management.azure.com/subscriptions/{subscription_id}"
            f"/resourceGroups/{resource_group}"
            f"/providers/Microsoft.Compute/virtualMachines/{vm_name}"
            f"/restart?api-version=2024-07-01"
        )

        async with httpx.AsyncClient(timeout=60) as client:
            resp = await client.post(url, headers={"Authorization": f"Bearer {token.token}"})

        if resp.status_code in (200, 202):
            return {
                "status": "ok",
                "message": f"VM '{vm_name}' restart initiated (HTTP {resp.status_code}).",
            }
        else:
            return {"status": "error", "message": f"Azure API returned HTTP {resp.status_code}: {resp.text}"}
    except Exception as e:
        return {"status": "error", "message": f"Failed to restart VM: {e}"}


@router.get("/worker-vm/auto-check")
async def worker_vm_auto_check(
    db: AsyncSession = Depends(get_db),
):
    """Automated health check - call from cron/scheduler.
    If worker heartbeat stale >10 min, attempts VM restart."""
    import os
    from datetime import datetime, timezone

    try:
        result = await db.execute(text("""
            SELECT last_heartbeat FROM worker_heartbeats
            ORDER BY last_heartbeat DESC LIMIT 1
        """))
        row = result.fetchone()

        if not row:
            return {"action": "none", "reason": "No heartbeat data"}

        now = datetime.now(timezone.utc)
        age_seconds = (now - row.last_heartbeat).total_seconds()

        if age_seconds < 600:
            return {
                "action": "none",
                "reason": f"Worker alive (heartbeat {age_seconds:.0f}s ago)",
                "last_heartbeat": row.last_heartbeat.isoformat(),
            }

        # Worker is dead - attempt restart
        subscription_id = os.getenv("AZURE_SUBSCRIPTION_ID")
        resource_group = os.getenv("AZURE_RESOURCE_GROUP")
        vm_name = os.getenv("AZURE_VM_NAME", "aitherhubVm")

        if not subscription_id or not resource_group:
            return {
                "action": "alert_only",
                "reason": f"Worker dead (heartbeat {age_seconds:.0f}s ago). Cannot auto-restart: missing Azure config.",
            }

        try:
            from azure.identity import DefaultAzureCredential
            import httpx

            credential = DefaultAzureCredential()
            token = credential.get_token("https://management.azure.com/.default")

            url = (
                f"https://management.azure.com/subscriptions/{subscription_id}"
                f"/resourceGroups/{resource_group}"
                f"/providers/Microsoft.Compute/virtualMachines/{vm_name}"
                f"/restart?api-version=2024-07-01"
            )

            async with httpx.AsyncClient(timeout=60) as client:
                resp = await client.post(url, headers={"Authorization": f"Bearer {token.token}"})

            if resp.status_code in (200, 202):
                await db.execute(text("""
                    INSERT INTO worker_heartbeats (worker_id, last_heartbeat, status)
                    VALUES ('auto-restart', NOW(), 'restarting')
                    ON CONFLICT (worker_id) DO UPDATE SET last_heartbeat = NOW(), status = 'restarting'
                """))
                await db.commit()
                return {"action": "restarted", "reason": f"Worker dead ({age_seconds:.0f}s). VM restart initiated."}
            else:
                return {"action": "restart_failed", "reason": f"Azure API HTTP {resp.status_code}"}
        except Exception as e:
            return {"action": "restart_failed", "reason": f"Auto-restart failed: {e}"}
    except Exception as e:
        return {"action": "error", "reason": str(e)}


# ─────────────────────────────────────────────────────────────────────────────
# Reset dead / stuck clips
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/reset-dead-clips")
async def reset_dead_clips(
    video_id: str | None = None,
    db: AsyncSession = Depends(get_db),
):
    """Reset clips that are stuck in 'dead' or 'error' status back to 'pending'.

    This allows the worker to re-process them. Optionally filter by video_id.
    Root cause of dead clips was the simple-worker duplicate OOM crash loop
    which caused clips to exceed POISON_MAX_RETRY.
    """
    try:
        if video_id:
            result = await db.execute(text("""
                UPDATE video_clips
                SET status = 'pending', progress_pct = 0, progress_step = 'queued',
                    error_message = NULL, updated_at = NOW()
                WHERE video_id = :video_id AND status IN ('dead', 'error')
                RETURNING id
            """), {"video_id": video_id})
        else:
            result = await db.execute(text("""
                UPDATE video_clips
                SET status = 'pending', progress_pct = 0, progress_step = 'queued',
                    error_message = NULL, updated_at = NOW()
                WHERE status IN ('dead', 'error')
                RETURNING id
            """))
        rows = result.fetchall()
        await db.commit()
        clip_ids = [str(r[0]) for r in rows]
        return {
            "status": "ok",
            "reset_count": len(clip_ids),
            "clip_ids": clip_ids,
            "message": f"Reset {len(clip_ids)} dead/error clip(s) to pending",
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


@router.post("/reset-dead-clips/{video_id}/requeue")
async def reset_and_requeue_dead_clips(
    video_id: str,
    db: AsyncSession = Depends(get_db),
):
    """Reset dead/error clips for a video AND re-enqueue them to the worker queue.

    This is the full recovery path: reset DB status + push new queue messages.
    """
    try:
        # 1. Reset clips in DB
        result = await db.execute(text("""
            UPDATE video_clips
            SET status = 'pending', progress_pct = 0, progress_step = 'queued',
                error_message = NULL, updated_at = NOW()
            WHERE video_id = :video_id AND status IN ('dead', 'error')
            RETURNING id, time_start, time_end
        """), {"video_id": video_id})
        rows = result.fetchall()

        if not rows:
            return {"status": "ok", "message": "No dead/error clips found for this video"}

        # 2. Get video info for SAS URL generation
        #    Note: videos table has no "blob_url" column. Use compressed_blob_url
        #    or generate a fresh download SAS from the video's storage path.
        from app.services.storage_service import generate_read_sas_from_url, generate_download_sas
        vid_result = await db.execute(text("""
            SELECT id, user_id, original_filename, compressed_blob_url FROM videos WHERE id = :video_id
        """), {"video_id": video_id})
        vid_row = vid_result.fetchone()

        # 3. Generate fresh SAS URL for the video
        fresh_blob_url = None
        if vid_row:
            # Try compressed_blob_url first (faster for clip generation)
            compressed = getattr(vid_row, 'compressed_blob_url', None)
            if compressed:
                try:
                    from app.services.storage_service import ACCOUNT_NAME, CONTAINER_NAME
                    full_url = f"https://{ACCOUNT_NAME}.blob.core.windows.net/{CONTAINER_NAME}/{compressed}"
                    sas_url = generate_read_sas_from_url(full_url)
                    if sas_url:
                        fresh_blob_url = sas_url
                        logger.info(f"Using compressed video SAS for requeue of video {video_id}")
                except Exception as e:
                    logger.warning(f"Compressed SAS failed for {video_id}: {e}")

            # Fallback: generate download SAS from original filename
            if not fresh_blob_url and vid_row.original_filename:
                try:
                    # Get user email for storage path
                    user_result = await db.execute(text(
                        "SELECT email FROM users WHERE id = :uid"
                    ), {"uid": str(vid_row.user_id)})
                    user_row = user_result.fetchone()
                    if user_row:
                        dl_url, _ = await generate_download_sas(
                            email=user_row.email,
                            video_id=video_id,
                            filename=vid_row.original_filename,
                            expires_in_minutes=1440,
                        )
                        if dl_url:
                            fresh_blob_url = dl_url
                            logger.info(f"Generated download SAS for requeue of video {video_id}")
                except Exception as e:
                    logger.warning(f"Download SAS generation failed for {video_id}: {e}")

        if not fresh_blob_url:
            # Last resort: try to use blob_url from the first clip's job_payload
            for row in rows:
                try:
                    clip_payload = row.job_payload if hasattr(row, 'job_payload') else None
                    if not clip_payload:
                        # Re-fetch with job_payload
                        cp_result = await db.execute(text(
                            "SELECT job_payload FROM video_clips WHERE id = :cid"
                        ), {"cid": str(row.id)})
                        cp_row = cp_result.fetchone()
                        if cp_row and cp_row.job_payload:
                            clip_payload = cp_row.job_payload
                    if clip_payload:
                        if isinstance(clip_payload, str):
                            clip_payload = json.loads(clip_payload)
                        old_url = clip_payload.get("blob_url", "")
                        if old_url:
                            base_url = old_url.split("?")[0]
                            new_sas = generate_read_sas_from_url(base_url)
                            if new_sas:
                                fresh_blob_url = new_sas
                                logger.info(f"Regenerated SAS from clip job_payload for video {video_id}")
                                break
                except Exception:
                    continue

        if not fresh_blob_url:
            await db.commit()
            return {
                "status": "partial",
                "message": f"Reset {len(rows)} clips in DB but could not requeue (no blob_url available)",
                "reset_count": len(rows),
            }

        # 4. Enqueue to Azure Storage Queue
        import os
        from azure.storage.queue import QueueClient

        conn_str = os.getenv("AZURE_STORAGE_CONNECTION_STRING", "")
        queue_name = os.getenv("AZURE_QUEUE_NAME", "aitherhub-jobs")
        queue_client = QueueClient.from_connection_string(conn_str, queue_name)

        enqueued = 0
        for row in rows:
            import json as _json
            msg = _json.dumps({
                "type": "generate_clip",
                "video_id": video_id,
                "clip_id": str(row.id),
                "blob_url": fresh_blob_url,
                "time_start": float(row.time_start) if row.time_start else 0,
                "time_end": float(row.time_end) if row.time_end else 0,
            })
            queue_client.send_message(msg, visibility_timeout=0)
            enqueued += 1

        await db.commit()
        return {
            "status": "ok",
            "reset_count": len(rows),
            "enqueued_count": enqueued,
            "message": f"Reset and requeued {enqueued} clip(s)",
        }
    except Exception as e:
        import traceback
        return {"status": "error", "message": str(e), "trace": traceback.format_exc()}


# ═══════════════════════════════════════════════════════════════════════
# Admin Cross-Brand Analytics + ML Insights
# ═══════════════════════════════════════════════════════════════════════

from datetime import datetime, timedelta


async def _qa(db: AsyncSession, sql: str, params: dict = None, default=None):
    """Run a query returning all rows with rollback on failure."""
    try:
        r = await db.execute(text(sql), params or {})
        return r.fetchall()
    except Exception as e:
        logger.warning(f"Admin analytics query error: {e}")
        try:
            await db.rollback()
        except Exception:
            pass
        return default if default is not None else []


@router.get("/analytics/overview")
async def admin_analytics_overview(
    days: int = Query(30, ge=1, le=365),
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
):
    """Cross-brand KPI overview with period comparison."""
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")

    try:
        now = datetime.utcnow()
        period_start = now - timedelta(days=days)
        prev_start = period_start - timedelta(days=days)

        # Current period KPIs
        kpi_sql = """
        SELECT
            COUNT(*) FILTER (WHERE event_type = 'video_play') as plays,
            COUNT(DISTINCT session_id) FILTER (WHERE event_type = 'video_play') as unique_viewers,
            COUNT(*) FILTER (WHERE event_type = 'cta_click') as cta_clicks,
            COUNT(*) FILTER (WHERE event_type = 'add_to_cart') as add_to_cart,
            COUNT(*) FILTER (WHERE event_type = 'purchase_click') as purchases,
            COUNT(*) FILTER (WHERE event_type = 'conversion') as conversions,
            COUNT(*) FILTER (WHERE event_type = 'video_progress'
                AND extra_data IS NOT NULL AND extra_data->>'milestone' = '100') as completions,
            COUNT(*) FILTER (WHERE event_type = 'video_replay') as replays
        FROM widget_tracking_events
        WHERE created_at >= :start
        """
        current = await _qa(db, kpi_sql, {"start": period_start})
        prev = await _qa(db, kpi_sql, {"start": prev_start})

        c = current[0] if current else None
        p = prev[0] if prev else None

        def safe(row, idx):
            return int(row[idx]) if row and row[idx] else 0

        plays = safe(c, 0)
        unique_viewers = safe(c, 1)
        cta_clicks = safe(c, 2)
        add_to_cart = safe(c, 3)
        purchases = safe(c, 4)
        conversions = safe(c, 5)
        completions = safe(c, 6)
        replays = safe(c, 7)

        prev_plays = safe(p, 0)
        prev_cta = safe(p, 2)
        prev_purchases = safe(p, 4)
        prev_completions = safe(p, 6)

        completion_rate = round(completions / plays * 100, 1) if plays > 0 else 0
        ctr = round(cta_clicks / plays * 100, 1) if plays > 0 else 0
        cvr = round(conversions / plays * 100, 2) if plays > 0 else 0

        prev_completion_rate = round(prev_completions / prev_plays * 100, 1) if prev_plays > 0 else 0
        prev_ctr = round(prev_cta / prev_plays * 100, 1) if prev_plays > 0 else 0
        prev_cvr = round(prev_purchases / prev_plays * 100, 2) if prev_plays > 0 else 0

        def delta(cur, prev_val):
            if prev_val == 0:
                return None
            return round((cur - prev_val) / prev_val * 100, 1)

        # Total active clips and brands
        active_clips = await _q(db, """
            SELECT COUNT(DISTINCT wca.clip_id)
            FROM widget_clip_assignments wca WHERE wca.is_active = true
        """)
        active_brands = await _q(db, """
            SELECT COUNT(DISTINCT wca.client_id)
            FROM widget_clip_assignments wca WHERE wca.is_active = true
        """)

        return {
            "period_days": days,
            "kpi": {
                "plays": plays, "plays_delta": delta(plays, prev_plays),
                "unique_viewers": unique_viewers,
                "completion_rate": completion_rate, "completion_rate_delta": delta(completion_rate, prev_completion_rate),
                "ctr": ctr, "ctr_delta": delta(ctr, prev_ctr),
                "cvr": cvr, "cvr_delta": delta(cvr, prev_cvr),
                "cta_clicks": cta_clicks,
                "add_to_cart": add_to_cart,
                "purchases": purchases,
                "conversions": conversions,
                "completions": completions,
                "replays": replays,
            },
            "active_clips": active_clips,
            "active_brands": active_brands,
        }
    except Exception as e:
        logger.error(f"Admin analytics overview error: {e}")
        import traceback
        return {"error": str(e), "trace": traceback.format_exc()}


@router.get("/analytics/brand-comparison")
async def admin_analytics_brand_comparison(
    days: int = Query(30, ge=1, le=365),
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
):
    """Compare performance across all brands."""
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")

    try:
        period_start = datetime.utcnow() - timedelta(days=days)

        rows = await _qa(db, """
        WITH brand_events AS (
            SELECT
                wca.client_id,
                wte.event_type,
                wte.clip_id,
                wte.extra_data
            FROM widget_tracking_events wte
            JOIN widget_clip_assignments wca ON wca.clip_id = wte.clip_id AND wca.is_active = true
            WHERE wte.created_at >= :start
        )
        SELECT
            be.client_id,
            COALESCE(wc.name, be.client_id) as brand_name,
            COUNT(*) FILTER (WHERE be.event_type = 'video_play') as plays,
            COUNT(*) FILTER (WHERE be.event_type = 'cta_click') as cta_clicks,
            COUNT(*) FILTER (WHERE be.event_type = 'add_to_cart') as cart,
            COUNT(*) FILTER (WHERE be.event_type = 'purchase_click') as purchases,
            COUNT(*) FILTER (WHERE be.event_type = 'conversion') as conversions,
            COUNT(*) FILTER (WHERE be.event_type = 'video_progress'
                AND be.extra_data IS NOT NULL AND be.extra_data->>'milestone' = '100') as completions,
            COUNT(*) FILTER (WHERE be.event_type = 'video_replay') as replays,
            COUNT(DISTINCT be.clip_id) FILTER (WHERE be.event_type = 'video_play') as active_clips
        FROM brand_events be
        LEFT JOIN widget_clients wc ON wc.client_id = be.client_id
        GROUP BY be.client_id, wc.name
        ORDER BY plays DESC
        """, {"start": period_start})

        brands = []
        for r in rows:
            plays = int(r[2]) if r[2] else 0
            cta = int(r[3]) if r[3] else 0
            completions = int(r[7]) if r[7] else 0
            conversions = int(r[6]) if r[6] else 0
            brands.append({
                "client_id": r[0],
                "brand_name": r[1] or r[0],
                "plays": plays,
                "cta_clicks": cta,
                "add_to_cart": int(r[4]) if r[4] else 0,
                "purchases": int(r[5]) if r[5] else 0,
                "conversions": conversions,
                "completions": completions,
                "replays": int(r[8]) if r[8] else 0,
                "active_clips": int(r[9]) if r[9] else 0,
                "completion_rate": round(completions / plays * 100, 1) if plays > 0 else 0,
                "ctr": round(cta / plays * 100, 1) if plays > 0 else 0,
                "cvr": round(conversions / plays * 100, 2) if plays > 0 else 0,
            })

        return {"brands": brands, "period_days": days}
    except Exception as e:
        logger.error(f"Admin brand comparison error: {e}")
        return {"brands": [], "error": str(e)}


@router.get("/analytics/top-clips")
async def admin_analytics_top_clips(
    days: int = Query(30, ge=1, le=365),
    limit: int = Query(20, ge=1, le=100),
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
):
    """Top performing clips across all brands by engagement score."""
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")

    try:
        period_start = datetime.utcnow() - timedelta(days=days)

        rows = await _qa(db, """
        WITH clip_metrics AS (
            SELECT
                wte.clip_id,
                COUNT(*) FILTER (WHERE wte.event_type = 'video_play') as plays,
                COUNT(*) FILTER (WHERE wte.event_type = 'cta_click') as cta_clicks,
                COUNT(*) FILTER (WHERE wte.event_type = 'conversion') as conversions,
                COUNT(*) FILTER (WHERE wte.event_type = 'video_progress'
                    AND wte.extra_data IS NOT NULL AND wte.extra_data->>'milestone' = '100') as completions,
                COUNT(*) FILTER (WHERE wte.event_type = 'video_replay') as replays
            FROM widget_tracking_events wte
            WHERE wte.created_at >= :start
            GROUP BY wte.clip_id
            HAVING COUNT(*) FILTER (WHERE wte.event_type = 'video_play') > 0
        )
        SELECT
            cm.clip_id,
            cm.plays,
            cm.cta_clicks,
            cm.conversions,
            cm.completions,
            cm.replays,
            vc.clip_url,
            vc.product_name,
            vc.duration,
            COALESCE(wc.name, wca.client_id, 'Unknown') as brand_name,
            wca.client_id,
            -- Engagement score: weighted combination
            (cm.completions::float / NULLIF(cm.plays, 0) * 40
             + cm.cta_clicks::float / NULLIF(cm.plays, 0) * 30
             + cm.conversions::float / NULLIF(cm.plays, 0) * 20
             + cm.replays::float / NULLIF(cm.plays, 0) * 10) as engagement_score
        FROM clip_metrics cm
        LEFT JOIN video_clips vc ON vc.clip_id = cm.clip_id
        LEFT JOIN widget_clip_assignments wca ON wca.clip_id = cm.clip_id AND wca.is_active = true
        LEFT JOIN widget_clients wc ON wc.client_id = wca.client_id
        ORDER BY engagement_score DESC NULLS LAST
        LIMIT :lim
        """, {"start": period_start, "lim": limit})

        clips = []
        for r in rows:
            plays = int(r[1]) if r[1] else 0
            completions = int(r[4]) if r[4] else 0
            cta = int(r[2]) if r[2] else 0
            conversions = int(r[3]) if r[3] else 0
            clips.append({
                "clip_id": r[0],
                "plays": plays,
                "cta_clicks": cta,
                "conversions": conversions,
                "completions": completions,
                "replays": int(r[5]) if r[5] else 0,
                "clip_url": r[6],
                "product_name": r[7],
                "duration": float(r[8]) if r[8] else 0,
                "brand_name": r[9],
                "client_id": r[10],
                "completion_rate": round(completions / plays * 100, 1) if plays > 0 else 0,
                "ctr": round(cta / plays * 100, 1) if plays > 0 else 0,
                "cvr": round(conversions / plays * 100, 2) if plays > 0 else 0,
                "engagement_score": round(float(r[11]), 2) if r[11] else 0,
            })

        return {"clips": clips, "period_days": days}
    except Exception as e:
        logger.error(f"Admin top clips error: {e}")
        return {"clips": [], "error": str(e)}


@router.get("/analytics/funnel")
async def admin_analytics_funnel(
    days: int = Query(30, ge=1, le=365),
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
):
    """Cross-brand funnel analysis."""
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")

    try:
        period_start = datetime.utcnow() - timedelta(days=days)

        row = await _qa(db, """
        SELECT
            COUNT(*) FILTER (WHERE event_type = 'video_play') as plays,
            COUNT(*) FILTER (WHERE event_type = 'video_progress'
                AND extra_data IS NOT NULL AND extra_data->>'milestone' IN ('50', '75', '100')) as deep_views,
            COUNT(*) FILTER (WHERE event_type = 'cta_click') as cta_clicks,
            COUNT(*) FILTER (WHERE event_type = 'add_to_cart') as add_to_cart,
            COUNT(*) FILTER (WHERE event_type = 'purchase_click') as purchase_clicks,
            COUNT(*) FILTER (WHERE event_type = 'conversion') as conversions
        FROM widget_tracking_events
        WHERE created_at >= :start
        """, {"start": period_start})

        r = row[0] if row else None
        plays = int(r[0]) if r and r[0] else 0

        stages = [
            {"name": "動画再生", "count": plays, "rate": 100},
            {"name": "深い視聴 (50%+)", "count": int(r[1]) if r and r[1] else 0,
             "rate": round(int(r[1]) / plays * 100, 1) if plays > 0 and r and r[1] else 0},
            {"name": "商品クリック", "count": int(r[2]) if r and r[2] else 0,
             "rate": round(int(r[2]) / plays * 100, 1) if plays > 0 and r and r[2] else 0},
            {"name": "カート追加", "count": int(r[3]) if r and r[3] else 0,
             "rate": round(int(r[3]) / plays * 100, 1) if plays > 0 and r and r[3] else 0},
            {"name": "購入クリック", "count": int(r[4]) if r and r[4] else 0,
             "rate": round(int(r[4]) / plays * 100, 1) if plays > 0 and r and r[4] else 0},
            {"name": "コンバージョン", "count": int(r[5]) if r and r[5] else 0,
             "rate": round(int(r[5]) / plays * 100, 1) if plays > 0 and r and r[5] else 0},
        ]

        return {"stages": stages, "period_days": days}
    except Exception as e:
        logger.error(f"Admin funnel error: {e}")
        return {"stages": [], "error": str(e)}


@router.get("/analytics/daily")
async def admin_analytics_daily(
    days: int = Query(30, ge=1, le=365),
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
):
    """Daily trend data for charts."""
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")

    try:
        period_start = datetime.utcnow() - timedelta(days=days)

        rows = await _qa(db, """
        SELECT
            DATE(created_at) as day,
            COUNT(*) FILTER (WHERE event_type = 'video_play') as plays,
            COUNT(*) FILTER (WHERE event_type = 'cta_click') as cta_clicks,
            COUNT(*) FILTER (WHERE event_type = 'conversion') as conversions,
            COUNT(*) FILTER (WHERE event_type = 'video_progress'
                AND extra_data IS NOT NULL AND extra_data->>'milestone' = '100') as completions,
            COUNT(DISTINCT session_id) FILTER (WHERE event_type = 'video_play') as unique_viewers
        FROM widget_tracking_events
        WHERE created_at >= :start
        GROUP BY DATE(created_at)
        ORDER BY day
        """, {"start": period_start})

        daily = []
        for r in rows:
            daily.append({
                "date": str(r[0]),
                "plays": int(r[1]) if r[1] else 0,
                "cta_clicks": int(r[2]) if r[2] else 0,
                "conversions": int(r[3]) if r[3] else 0,
                "completions": int(r[4]) if r[4] else 0,
                "unique_viewers": int(r[5]) if r[5] else 0,
            })

        return {"daily": daily, "period_days": days}
    except Exception as e:
        logger.error(f"Admin daily error: {e}")
        return {"daily": [], "error": str(e)}


@router.get("/analytics/ml-insights")
async def admin_analytics_ml_insights(
    days: int = Query(30, ge=1, le=365),
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
):
    """ML-powered insights: tag effectiveness, winning patterns, duration analysis."""
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")

    try:
        period_start = datetime.utcnow() - timedelta(days=days)

        # 1. Tag effectiveness analysis
        tag_rows = await _qa(db, """
        WITH clip_tags AS (
            SELECT vc.clip_id, unnest(vc.sales_psychology_tags) as tag
            FROM video_clips vc
            WHERE vc.sales_psychology_tags IS NOT NULL
              AND array_length(vc.sales_psychology_tags, 1) > 0
        ),
        clip_perf AS (
            SELECT
                wte.clip_id,
                COUNT(*) FILTER (WHERE wte.event_type = 'video_play') as plays,
                COUNT(*) FILTER (WHERE wte.event_type = 'cta_click') as cta_clicks,
                COUNT(*) FILTER (WHERE wte.event_type = 'conversion') as conversions,
                COUNT(*) FILTER (WHERE wte.event_type = 'video_progress'
                    AND wte.extra_data IS NOT NULL AND wte.extra_data->>'milestone' = '100') as completions
            FROM widget_tracking_events wte
            WHERE wte.created_at >= :start
            GROUP BY wte.clip_id
        )
        SELECT
            ct.tag,
            COUNT(DISTINCT ct.clip_id) as clip_count,
            SUM(cp.plays) as total_plays,
            CASE WHEN SUM(cp.plays) > 0
                THEN ROUND(SUM(cp.completions)::numeric / SUM(cp.plays) * 100, 1)
                ELSE 0 END as avg_completion_rate,
            CASE WHEN SUM(cp.plays) > 0
                THEN ROUND(SUM(cp.cta_clicks)::numeric / SUM(cp.plays) * 100, 1)
                ELSE 0 END as avg_ctr,
            CASE WHEN SUM(cp.plays) > 0
                THEN ROUND(SUM(cp.conversions)::numeric / SUM(cp.plays) * 100, 2)
                ELSE 0 END as avg_cvr
        FROM clip_tags ct
        JOIN clip_perf cp ON cp.clip_id = ct.clip_id
        GROUP BY ct.tag
        HAVING SUM(cp.plays) > 0
        ORDER BY avg_ctr DESC
        """, {"start": period_start})

        tag_effectiveness = []
        for r in tag_rows:
            tag_effectiveness.append({
                "tag": r[0],
                "clip_count": int(r[1]) if r[1] else 0,
                "total_plays": int(r[2]) if r[2] else 0,
                "completion_rate": float(r[3]) if r[3] else 0,
                "ctr": float(r[4]) if r[4] else 0,
                "cvr": float(r[5]) if r[5] else 0,
            })

        # 2. Duration analysis (which clip lengths perform best)
        duration_rows = await _qa(db, """
        WITH clip_dur AS (
            SELECT
                vc.clip_id,
                CASE
                    WHEN vc.duration < 10 THEN '0-10s'
                    WHEN vc.duration < 20 THEN '10-20s'
                    WHEN vc.duration < 30 THEN '20-30s'
                    WHEN vc.duration < 45 THEN '30-45s'
                    WHEN vc.duration < 60 THEN '45-60s'
                    ELSE '60s+'
                END as duration_bucket
            FROM video_clips vc
            WHERE vc.duration IS NOT NULL AND vc.duration > 0
        ),
        clip_perf AS (
            SELECT
                wte.clip_id,
                COUNT(*) FILTER (WHERE wte.event_type = 'video_play') as plays,
                COUNT(*) FILTER (WHERE wte.event_type = 'cta_click') as cta_clicks,
                COUNT(*) FILTER (WHERE wte.event_type = 'conversion') as conversions,
                COUNT(*) FILTER (WHERE wte.event_type = 'video_progress'
                    AND wte.extra_data IS NOT NULL AND wte.extra_data->>'milestone' = '100') as completions
            FROM widget_tracking_events wte
            WHERE wte.created_at >= :start
            GROUP BY wte.clip_id
        )
        SELECT
            cd.duration_bucket,
            COUNT(DISTINCT cd.clip_id) as clip_count,
            SUM(cp.plays) as total_plays,
            CASE WHEN SUM(cp.plays) > 0
                THEN ROUND(SUM(cp.completions)::numeric / SUM(cp.plays) * 100, 1)
                ELSE 0 END as avg_completion_rate,
            CASE WHEN SUM(cp.plays) > 0
                THEN ROUND(SUM(cp.cta_clicks)::numeric / SUM(cp.plays) * 100, 1)
                ELSE 0 END as avg_ctr,
            CASE WHEN SUM(cp.plays) > 0
                THEN ROUND(SUM(cp.conversions)::numeric / SUM(cp.plays) * 100, 2)
                ELSE 0 END as avg_cvr
        FROM clip_dur cd
        JOIN clip_perf cp ON cp.clip_id = cd.clip_id
        GROUP BY cd.duration_bucket
        ORDER BY
            CASE cd.duration_bucket
                WHEN '0-10s' THEN 1
                WHEN '10-20s' THEN 2
                WHEN '20-30s' THEN 3
                WHEN '30-45s' THEN 4
                WHEN '45-60s' THEN 5
                ELSE 6
            END
        """, {"start": period_start})

        duration_analysis = []
        for r in duration_rows:
            duration_analysis.append({
                "bucket": r[0],
                "clip_count": int(r[1]) if r[1] else 0,
                "total_plays": int(r[2]) if r[2] else 0,
                "completion_rate": float(r[3]) if r[3] else 0,
                "ctr": float(r[4]) if r[4] else 0,
                "cvr": float(r[5]) if r[5] else 0,
            })

        # 3. Sold vs Unsold pattern analysis
        sold_rows = await _qa(db, """
        WITH clip_perf AS (
            SELECT
                wte.clip_id,
                COUNT(*) FILTER (WHERE wte.event_type = 'video_play') as plays,
                COUNT(*) FILTER (WHERE wte.event_type = 'cta_click') as cta_clicks,
                COUNT(*) FILTER (WHERE wte.event_type = 'video_progress'
                    AND wte.extra_data IS NOT NULL AND wte.extra_data->>'milestone' = '100') as completions
            FROM widget_tracking_events wte
            WHERE wte.created_at >= :start
            GROUP BY wte.clip_id
        )
        SELECT
            COALESCE(vc.is_sold, false) as is_sold,
            COUNT(DISTINCT vc.clip_id) as clip_count,
            AVG(vc.duration) as avg_duration,
            AVG(array_length(vc.sales_psychology_tags, 1)) as avg_tag_count,
            CASE WHEN SUM(cp.plays) > 0
                THEN ROUND(SUM(cp.completions)::numeric / SUM(cp.plays) * 100, 1)
                ELSE 0 END as avg_completion_rate,
            CASE WHEN SUM(cp.plays) > 0
                THEN ROUND(SUM(cp.cta_clicks)::numeric / SUM(cp.plays) * 100, 1)
                ELSE 0 END as avg_ctr
        FROM video_clips vc
        LEFT JOIN clip_perf cp ON cp.clip_id = vc.clip_id
        GROUP BY COALESCE(vc.is_sold, false)
        """, {"start": period_start})

        sold_analysis = []
        for r in sold_rows:
            sold_analysis.append({
                "is_sold": bool(r[0]),
                "clip_count": int(r[1]) if r[1] else 0,
                "avg_duration": round(float(r[2]), 1) if r[2] else 0,
                "avg_tag_count": round(float(r[3]), 1) if r[3] else 0,
                "completion_rate": float(r[4]) if r[4] else 0,
                "ctr": float(r[5]) if r[5] else 0,
            })

        # 4. Top winning tag combinations
        combo_rows = await _qa(db, """
        WITH clip_perf AS (
            SELECT
                wte.clip_id,
                COUNT(*) FILTER (WHERE wte.event_type = 'video_play') as plays,
                COUNT(*) FILTER (WHERE wte.event_type = 'cta_click') as cta_clicks,
                COUNT(*) FILTER (WHERE wte.event_type = 'conversion') as conversions
            FROM widget_tracking_events wte
            WHERE wte.created_at >= :start
            GROUP BY wte.clip_id
            HAVING COUNT(*) FILTER (WHERE wte.event_type = 'video_play') >= 3
        )
        SELECT
            array_to_string(vc.sales_psychology_tags, ' + ') as tag_combo,
            COUNT(*) as clip_count,
            SUM(cp.plays) as total_plays,
            CASE WHEN SUM(cp.plays) > 0
                THEN ROUND(SUM(cp.cta_clicks)::numeric / SUM(cp.plays) * 100, 1)
                ELSE 0 END as avg_ctr,
            CASE WHEN SUM(cp.plays) > 0
                THEN ROUND(SUM(cp.conversions)::numeric / SUM(cp.plays) * 100, 2)
                ELSE 0 END as avg_cvr
        FROM video_clips vc
        JOIN clip_perf cp ON cp.clip_id = vc.clip_id
        WHERE vc.sales_psychology_tags IS NOT NULL
          AND array_length(vc.sales_psychology_tags, 1) >= 2
        GROUP BY vc.sales_psychology_tags
        HAVING COUNT(*) >= 2
        ORDER BY avg_ctr DESC
        LIMIT 10
        """, {"start": period_start})

        winning_combos = []
        for r in combo_rows:
            winning_combos.append({
                "tags": r[0],
                "clip_count": int(r[1]) if r[1] else 0,
                "total_plays": int(r[2]) if r[2] else 0,
                "ctr": float(r[3]) if r[3] else 0,
                "cvr": float(r[4]) if r[4] else 0,
            })

        return {
            "tag_effectiveness": tag_effectiveness,
            "duration_analysis": duration_analysis,
            "sold_vs_unsold": sold_analysis,
            "winning_combos": winning_combos,
            "period_days": days,
        }
    except Exception as e:
        logger.error(f"Admin ML insights error: {e}")
        import traceback
        return {"error": str(e), "trace": traceback.format_exc()}


@router.get("/analytics/hourly-heatmap")
async def admin_analytics_hourly_heatmap(
    days: int = Query(30, ge=1, le=365),
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
):
    """Hourly activity heatmap (day of week x hour)."""
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")

    try:
        period_start = datetime.utcnow() - timedelta(days=days)

        rows = await _qa(db, """
        SELECT
            EXTRACT(DOW FROM created_at + INTERVAL '9 hours') as dow,
            EXTRACT(HOUR FROM created_at + INTERVAL '9 hours') as hour,
            COUNT(*) FILTER (WHERE event_type = 'video_play') as plays,
            COUNT(*) FILTER (WHERE event_type = 'cta_click') as clicks
        FROM widget_tracking_events
        WHERE created_at >= :start
        GROUP BY dow, hour
        ORDER BY dow, hour
        """, {"start": period_start})

        heatmap = []
        for r in rows:
            heatmap.append({
                "dow": int(r[0]) if r[0] is not None else 0,
                "hour": int(r[1]) if r[1] is not None else 0,
                "plays": int(r[2]) if r[2] else 0,
                "clicks": int(r[3]) if r[3] else 0,
            })

        return {"heatmap": heatmap, "period_days": days}
    except Exception as e:
        logger.error(f"Admin heatmap error: {e}")
        return {"heatmap": [], "error": str(e)}


# ── Backfill transcript_text from captions ─────────────────────────
@router.post("/backfill-transcript-text")
async def backfill_transcript_text(
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
):
    """
    One-shot backfill: for clips that have captions JSONB but NULL transcript_text,
    build transcript_text by joining all caption texts.
    """
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")

    try:
        sql = text("""
            UPDATE video_clips
            SET transcript_text = sub.built_text,
                updated_at = NOW()
            FROM (
                SELECT id,
                       LEFT(
                           string_agg(elem->>'text', ' ' ORDER BY (elem->>'start')::float),
                           500
                       ) AS built_text
                FROM video_clips,
                     jsonb_array_elements(captions) AS elem
                WHERE (transcript_text IS NULL OR transcript_text = '')
                  AND captions IS NOT NULL
                  AND jsonb_array_length(captions) > 0
                GROUP BY id
            ) sub
            WHERE video_clips.id = sub.id
              AND sub.built_text IS NOT NULL
              AND sub.built_text != ''
        """)
        result = await db.execute(sql)
        count = result.rowcount
        await db.commit()
        logger.info(f"[BACKFILL] Updated transcript_text for {count} clips from captions")
        return {"updated": count, "message": f"Backfilled transcript_text for {count} clips"}
    except Exception as e:
        logger.error(f"[BACKFILL] transcript_text backfill failed: {e}")
        try:
            await db.rollback()
        except Exception:
            pass
        raise HTTPException(status_code=500, detail=str(e))



@router.get("/users-list")
async def get_users_list(
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
    db: AsyncSession = Depends(get_db),
):
    """Get list of all users with their video counts and last upload date."""
    expected_key = f"{ADMIN_ID}:{ADMIN_PASS}"
    if x_admin_key != expected_key:
        raise HTTPException(status_code=403, detail="Invalid admin credentials")

    result = await db.execute(
        text("""
            SELECT
                u.id,
                u.email,
                u.display_name,
                u.created_at,
                COUNT(v.id) as video_count,
                MAX(v.created_at) as last_upload
            FROM users u
            LEFT JOIN videos v ON v.user_id = u.id
            GROUP BY u.id, u.email, u.display_name, u.created_at
            ORDER BY video_count DESC, u.created_at DESC
        """)
    )
    rows = result.fetchall()
    users = []
    for r in rows:
        users.append({
            "id": str(r.id),
            "email": r.email or "",
            "name": r.display_name or r.email or "",
            "video_count": r.video_count,
            "last_upload": r.last_upload.isoformat() if r.last_upload else None,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        })
    return {"users": users, "total": len(users)}
