from typing import List
import json
import os
import uuid as uuid_module
import asyncio
from datetime import datetime, timedelta, timezone

from typing import Optional
from fastapi import APIRouter, HTTPException, Depends, BackgroundTasks, Request
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text, select
from loguru import logger

from app.schema.video_schema import (
    RenameVideoRequest,
    RenameVideoResponse,
    DeleteVideoResponse,
    VideoResponse,
    LiveCaptureRequest,
    LiveCaptureResponse,
    LiveCheckResponse,
)
from app.services.video_service import VideoService
from app.repository.video_repository import VideoRepository
from app.core.dependencies import get_db, get_current_user
from app.utils.video_progress import calculate_progress, get_status_message
from app.core.container import Container
from app.models.orm.video import Video

router = APIRouter(
    prefix="/videos",
    tags=["videos"],
)

# Initialize service (could be injected via DI container)
video_service = VideoService()


import os as _os

_BLOB_HOST = _os.getenv("AZURE_BLOB_HOST", "https://aitherhub.blob.core.windows.net")
_CDN_HOST = _os.getenv("CDN_HOST", "https://cdn.aitherhub.com")


def _replace_blob_url_to_cdn(url: str) -> str:
    """Replace blob storage domain with CDN domain if applicable."""
    if url and isinstance(url, str):
        return url.replace(_BLOB_HOST, _CDN_HOST)
    return url





@router.get("/user/{user_id}", response_model=List[VideoResponse])
async def get_videos_by_user(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user),
):
    """
    Return list of videos for the given `user_id`.

    This endpoint requires authentication and only allows a user to fetch their own videos.
    """
    try:
        # Enforce that a user can only access their own videos
        if current_user and current_user.get("id") != user_id:
            raise HTTPException(status_code=403, detail="Forbidden")

        video_repo = VideoRepository(lambda: db)
        videos = await video_repo.get_videos_by_user(user_id=user_id)

        return [VideoResponse.from_orm(v) for v in videos]
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to fetch videos: {exc}")




@router.get("/user/{user_id}/with-clips")
async def get_videos_by_user_with_clips(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user),
):
    """
    Return list of videos for the given `user_id` with clip counts.
    This is used by the sidebar to show clip availability indicators.
    """
    try:
        if current_user and current_user.get("id") != user_id:
            raise HTTPException(status_code=403, detail="Forbidden")

        # Get videos with clip counts + sales/duration summary + memo count in a single query
        sql = text("""
            SELECT v.id, v.original_filename, v.status,
                   v.upload_type, v.created_at, v.updated_at,
                   COALESCE(c.clip_count, 0) as clip_count,
                   COALESCE(c.completed_count, 0) as completed_clip_count,
                   p.total_gmv,
                   p.max_time_end,
                   COALESCE(m.memo_count, 0) as memo_count,
                   v.top_products as top_products_json
            FROM videos v
            LEFT JOIN (
                SELECT video_id,
                       COUNT(DISTINCT phase_index) as clip_count,
                       COUNT(DISTINCT CASE WHEN status = 'completed' THEN phase_index END) as completed_count
                FROM video_clips
                GROUP BY video_id
            ) c ON v.id = c.video_id
            LEFT JOIN (
                SELECT video_id,
                       SUM(COALESCE(gmv, 0)) as total_gmv,
                       MAX(time_end) as max_time_end
                FROM video_phases
                GROUP BY video_id
            ) p ON v.id = p.video_id
            LEFT JOIN (
                SELECT video_id,
                       COUNT(*) as memo_count
                FROM video_phases
                WHERE (user_comment IS NOT NULL AND user_comment != '')
                   OR (user_rating IS NOT NULL AND user_rating > 0)
                GROUP BY video_id
            ) m ON v.id = m.video_id
            WHERE (v.user_id = :user_id OR v.user_id IS NULL)
            ORDER BY v.created_at DESC
        """)
        result = await db.execute(sql, {"user_id": user_id})
        rows = result.fetchall()

        import json as _json

        videos = []
        for row in rows:
            vid = str(row.id)
            # Parse cached top_products from videos table
            top_prods = []
            if row.top_products_json:
                try:
                    top_prods = _json.loads(row.top_products_json)
                except (ValueError, TypeError):
                    top_prods = []
            videos.append({
                "id": vid,
                "original_filename": row.original_filename,
                "status": row.status,
                "upload_type": row.upload_type,
                "created_at": row.created_at.isoformat() if row.created_at else None,
                "updated_at": row.updated_at.isoformat() if row.updated_at else None,
                "clip_count": row.clip_count,
                "completed_clip_count": row.completed_clip_count,
                "total_gmv": float(row.total_gmv) if row.total_gmv and float(row.total_gmv) > 0 else None,
                "stream_duration": float(row.max_time_end) if row.max_time_end else None,
                "memo_count": row.memo_count,
                "top_products": top_prods,
            })

        return videos

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(f"Failed to fetch videos with clips: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch videos: {exc}")


@router.delete("/{video_id}", response_model=DeleteVideoResponse)
async def delete_video(
    video_id: str,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user),
):
    """Delete a video and its related data (only owner can delete)."""
    try:
        user_id = current_user["id"]
        video_repo = VideoRepository(lambda: db)

        # Delete ALL related records first to avoid FK constraint violations
        # Order matters: delete child tables before parent tables
        # Use safe_delete to skip tables that may not exist yet
        tables_to_delete = [
            # Level 3: grandchild tables (FK to child tables)
            "DELETE FROM speech_segments WHERE audio_chunk_id IN (SELECT id FROM audio_chunks WHERE video_id = :vid)",
            "DELETE FROM frame_analysis_results WHERE frame_id IN (SELECT id FROM video_frames WHERE video_id = :vid)",
            # Level 2: child tables with video_id FK
            "DELETE FROM video_frames WHERE video_id = :vid",
            "DELETE FROM video_product_exposures WHERE video_id = :vid",
            "DELETE FROM video_clips WHERE video_id = :vid",
            "DELETE FROM audio_chunks WHERE video_id = :vid",
            "DELETE FROM chats WHERE video_id = :vid",
            "DELETE FROM group_best_phases WHERE video_id = :vid",
            "DELETE FROM phase_insights WHERE video_id = :vid",
            "DELETE FROM video_phases WHERE video_id = :vid",
            "DELETE FROM video_insights WHERE video_id = :vid",
            "DELETE FROM processing_jobs WHERE video_id = :vid",
            "DELETE FROM reports WHERE video_id = :vid",
            "DELETE FROM video_processing_state WHERE video_id = :vid",
            # Structure tables
            "DELETE FROM video_structure_group_best_videos WHERE video_id = :vid",
            "DELETE FROM video_structure_group_members WHERE video_id = :vid",
            "DELETE FROM video_structure_features WHERE video_id = :vid",
        ]

        for sql in tables_to_delete:
            try:
                await db.execute(text(sql), {"vid": video_id})
            except Exception as table_err:
                # Skip if table doesn't exist (e.g., migration not yet applied)
                logger.warning(f"Skipping delete for non-existent table: {table_err}")
                await db.rollback()
                # Re-start transaction for next delete
                continue

        await db.commit()

        # Delete the video record
        deleted = await video_repo.delete_video(video_id=video_id, user_id=user_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Video not found or not owned by user")

        return DeleteVideoResponse(id=video_id, message="Video deleted successfully")
    except HTTPException:
        raise
    except Exception as exc:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete video: {exc}")


@router.patch("/{video_id}/rename", response_model=RenameVideoResponse)
async def rename_video(
    video_id: str,
    payload: RenameVideoRequest,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user),
):
    """Rename a video (only owner can rename)."""
    try:
        user_id = current_user["id"]
        video_repo = VideoRepository(lambda: db)
        video = await video_repo.rename_video(
            video_id=video_id, user_id=user_id, new_name=payload.name
        )
        if not video:
            raise HTTPException(status_code=404, detail="Video not found or not owned by user")

        return RenameVideoResponse(
            id=str(video.id),
            original_filename=video.original_filename,
            message="Video renamed successfully",
        )
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to rename video: {exc}")


@router.get("/{video_id}/status/stream")
async def stream_video_status(
    video_id: str,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user),
):
    """
    Stream video processing status updates via Server-Sent Events (SSE).

    This endpoint provides real-time status updates for video processing.
    It polls the database every 2 seconds and sends status changes to the client.
    The stream automatically closes when processing reaches DONE or ERROR status.
    Supports long-running videos up to 4 hours with heartbeat messages every 15 seconds.
    Uses SSE comment keep-alive pings between heartbeats to prevent Azure App Service idle timeout (~230s).

    Args:
        video_id: UUID of the video to monitor
        db: Database session
        current_user: Authenticated user

    Returns:
        StreamingResponse with SSE events containing:
        - status: Current processing status
        - progress: Progress percentage (0-100)
        - message: User-friendly Japanese status message
        - updated_at: Timestamp of last update
        - heartbeat: Boolean indicating heartbeat message (sent every 15 seconds)

    Example SSE events:
        data: {"video_id": "...", "status": "STEP_3_TRANSCRIBE_AUDIO", "progress": 40, "message": "音声書き起こし中...", "updated_at": "2026-01-20T..."}
        data: {"heartbeat": true, "timestamp": "2026-01-20T...", "poll_count": 15}
    """

    async def event_generator():
        last_status = None
        last_step_progress = None
        last_processing_logs_len = 0
        poll_count = 0
        max_polls = 7200  # 4 hours max for long videos (7200 * 2 seconds = 14400 seconds = 4 hours)

        try:
            # Verify video exists and ownership
            video_repo = VideoRepository(lambda: db)
            video = await video_repo.get_video_by_id(video_id)

            if not video:
                yield f"data: {json.dumps({'error': 'Video not found'})}\n\n"
                return

            # Any authenticated user can view video status (removed owner-only restriction)

            # Stream status updates
            # ── Keep-alive strategy ──
            # Azure App Service has an idle timeout of ~230 seconds.
            # We use a multi-layer keep-alive approach:
            # 1. SSE comment pings (`: keep-alive`) every poll cycle (2s) - invisible to SSE parsers
            # 2. JSON heartbeat messages every 15 seconds - visible to frontend
            # 3. Status updates whenever status/progress changes
            # This ensures the connection is NEVER idle for more than 2 seconds.
            while poll_count < max_polls:
                try:
                    # Refresh video data
                    video = await video_repo.get_video_by_id(video_id)

                    if not video:
                        yield f"data: {json.dumps({'error': 'Video not found'})}\n\n"
                        break

                    current_status = video.status
                    current_step_progress = getattr(video, 'step_progress', None) or 0

                    # Check if processing_logs changed (compare by length for efficiency)
                    _current_pl = getattr(video, 'processing_logs', None) or []
                    _current_pl_len = len(_current_pl) if isinstance(_current_pl, list) else 0
                    _pl_changed = (_current_pl_len != last_processing_logs_len)
                    if _pl_changed:
                        last_processing_logs_len = _current_pl_len

                    # Send update if status changed OR step_progress changed OR processing_logs changed
                    if current_status != last_status or current_step_progress != last_step_progress or _pl_changed:
                        progress = calculate_progress(current_status)
                        message = get_status_message(current_status)

                        # Include processing_logs if available
                        _processing_logs = _current_pl if _current_pl_len > 0 else None

                        payload = {
                            "video_id": str(video.id),
                            "status": current_status,
                            "progress": progress,
                            "step_progress": current_step_progress,
                            "message": message,
                            "updated_at": video.updated_at.isoformat() if video.updated_at else None,
                            "created_at": video.created_at.isoformat() if video.created_at else None,
                            "server_now": datetime.utcnow().isoformat(),
                            # Enqueue & worker evidence
                            "enqueue_status": getattr(video, 'enqueue_status', None),
                            "queue_enqueued_at": video.queue_enqueued_at.isoformat() if getattr(video, 'queue_enqueued_at', None) else None,
                            "enqueue_error": getattr(video, 'enqueue_error', None),
                            "worker_claimed_at": video.worker_claimed_at.isoformat() if getattr(video, 'worker_claimed_at', None) else None,
                            "dequeue_count": getattr(video, 'dequeue_count', None),
                            # Processing logs for live display
                            "processing_logs": _processing_logs,
                        }

                        yield f"data: {json.dumps(payload)}\n\n"
                        last_status = current_status
                        last_step_progress = current_step_progress

                        logger.info(f"SSE: Video {video_id} status={current_status} step_progress={current_step_progress}%")
                    else:
                        # No status change - send SSE comment ping to prevent Azure idle timeout
                        # SSE spec: lines starting with ':' are comments, ignored by EventSource parsers
                        yield ": keep-alive\n\n"

                    # Send JSON heartbeat every 15 seconds (7-8 poll cycles * 2s)
                    # This is visible to the frontend and resets its heartbeat timer
                    if poll_count > 0 and poll_count % 7 == 0:
                        heartbeat_payload = {
                            "heartbeat": True,
                            "timestamp": datetime.utcnow().isoformat(),
                            "poll_count": poll_count
                        }
                        yield f"data: {json.dumps(heartbeat_payload)}\n\n"
                        logger.debug(f"SSE: Heartbeat sent for video {video_id} (poll {poll_count})")

                    # Stop streaming if processing complete or error
                    if current_status in ["DONE", "ERROR"]:
                        # On ERROR, fetch latest error log to include in final SSE event
                        if current_status == "ERROR":
                            try:
                                err_result = await db.execute(
                                    text("""
                                        SELECT error_code, error_step, error_message, source, created_at
                                        FROM video_error_logs
                                        WHERE video_id = :vid
                                        ORDER BY created_at DESC
                                        LIMIT 1
                                    """),
                                    {"vid": video_id},
                                )
                                err_row = err_result.fetchone()
                                if err_row:
                                    error_payload = {
                                        "video_id": str(video_id),
                                        "status": "ERROR",
                                        "latest_error": {
                                            "error_code": err_row.error_code,
                                            "error_step": err_row.error_step,
                                            "error_message": err_row.error_message,
                                            "source": err_row.source,
                                            "created_at": err_row.created_at.isoformat() if err_row.created_at else None,
                                        },
                                    }
                                    yield f"data: {json.dumps(error_payload)}\n\n"
                            except Exception as err_log_exc:
                                logger.warning(f"SSE: Failed to fetch error log for {video_id}: {err_log_exc}")
                        yield "data: [DONE]\n\n"
                        logger.info(f"SSE: Video {video_id} processing completed with status {current_status}")
                        break

                    # Poll every 2 seconds
                    await asyncio.sleep(2)
                    poll_count += 1

                except Exception as e:
                    logger.error(f"SSE poll error for video {video_id}: {e}")
                    yield f"data: {json.dumps({'error': f'Poll error: {str(e)}'})}\n\n"
                    break

            # Timeout reached
            if poll_count >= max_polls:
                logger.warning(f"SSE: Video {video_id} stream timeout after {max_polls * 2} seconds")
                yield f"data: {json.dumps({'error': 'Stream timeout'})}\n\n"

        except Exception as e:
            logger.error(f"SSE stream error for video {video_id}: {e}")
            yield f"data: {json.dumps({'error': str(e)})}\n\n"

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",  # Disable nginx buffering
            "Connection": "keep-alive",
        }
    )


async def _auto_recalc_csv_metrics(
    db: AsyncSession, video_id: str, trend_blob_url: str, time_offset_seconds: float
) -> list:
    """
    Auto-recalculate CSV metrics from Excel trend data when all metrics are zero.
    Returns list of dicts with phase_index and metric values, or empty list on failure.
    Also persists the recalculated values to DB.
    """
    import httpx
    import tempfile
    import openpyxl

    try:
        from app.services.storage_service import generate_read_sas_from_url
    except ImportError:
        return []

    # Download and parse trend Excel
    sas_url = generate_read_sas_from_url(trend_blob_url, expires_hours=1)
    if not sas_url:
        return []

    async with httpx.AsyncClient(timeout=30.0) as client:
        resp = await client.get(sas_url)
        if resp.status_code != 200:
            return []

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        f.write(resp.content)
        tmp_path = f.name

    try:
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        ws = wb.active
        trends = []
        if ws:
            rows_data = list(ws.iter_rows(values_only=True))
            if len(rows_data) >= 2:
                headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows_data[0])]
                for data_row in rows_data[1:]:
                    if all(v is None for v in data_row):
                        continue
                    item = {}
                    for i, val in enumerate(data_row):
                        if i < len(headers):
                            if val is None:
                                item[headers[i]] = None
                            elif isinstance(val, (int, float)):
                                item[headers[i]] = val
                            elif hasattr(val, 'hour') and hasattr(val, 'minute'):
                                item[headers[i]] = f"{val.hour:02d}:{val.minute:02d}"
                            else:
                                item[headers[i]] = str(val)
                    trends.append(item)
        wb.close()
    finally:
        os.unlink(tmp_path)

    if not trends:
        return []

    # Get phases
    phases_result = await db.execute(text("""
        SELECT phase_index, time_start, time_end
        FROM video_phases WHERE video_id = :vid ORDER BY phase_index ASC
    """), {"vid": video_id})
    phases = [{"phase_index": r[0], "time_start": float(r[1] or 0), "time_end": float(r[2] or 0)}
              for r in phases_result.fetchall()]
    if not phases:
        return []

    # KPI aliases
    KPI_ALIASES = {
        "time": ["\u6642\u9593", "time", "timestamp", "\u65f6\u95f4", "\uc2dc\uac04"],
        "gmv": ["GMV", "gmv", "\u58f2\u4e0a", "sales", "revenue", "\u6210\u4ea4\u91d1\u989d", "\ub9e4\ucd9c"],
        "order_count": ["\u6ce8\u6587", "order", "orders", "SKU\u6ce8\u6587\u6570", "\u8ba2\u5355", "\uc8fc\ubb38"],
        "viewer_count": ["\u8996\u8074\u8005", "viewer", "viewers", "\u89c6\u542c\u8005", "\u89c2\u4f17", "\uc2dc\uccad\uc790"],
        "like_count": ["\u3044\u3044\u306d", "like", "likes", "\u70b9\u8d5e", "\uc88b\uc544\uc694"],
        "comment_count": ["\u30b3\u30e1\u30f3\u30c8", "comment", "comments", "\u8bc4\u8bba", "\ub313\uae00"],
        "share_count": ["\u30b7\u30a7\u30a2", "share", "shares", "\u5206\u4eab", "\uacf5\uc720"],
        "new_followers": ["\u30d5\u30a9\u30ed\u30ef\u30fc", "follower", "followers", "\u65b0\u898f\u30d5\u30a9\u30ed\u30ef\u30fc", "\u7c89\u4e1d"],
        "product_clicks": ["\u5546\u54c1\u30af\u30ea\u30c3\u30af", "product_click", "clicks", "\u70b9\u51fb"],
        "ctor": ["CTOR", "ctor", "conversion"],
        "gpm": ["GPM", "gpm", "\u8996\u8074GPM"],
    }

    def _find_key(sample_dict, aliases):
        for alias in aliases:
            for k in sample_dict.keys():
                if alias.lower() in k.lower():
                    return k
        return None

    def _safe_float(val):
        if val is None:
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    def _parse_time_to_seconds(val):
        if val is None:
            return None
        if hasattr(val, 'hour') and hasattr(val, 'minute'):
            return val.hour * 3600 + val.minute * 60 + getattr(val, 'second', 0)
        val_str = str(val).strip()
        try:
            return float(val_str)
        except (ValueError, TypeError):
            pass
        parts = val_str.split(":")
        try:
            if len(parts) == 2:
                h, m = int(parts[0]), int(parts[1])
                if h < 24:
                    return h * 3600 + m * 60
                else:
                    return h * 60 + m
            elif len(parts) == 3:
                h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
                return h * 3600 + m * 60 + s
        except (ValueError, TypeError):
            pass
        return None

    sample = trends[0]
    time_key = _find_key(sample, KPI_ALIASES["time"])
    if not time_key:
        return []

    gmv_key = _find_key(sample, KPI_ALIASES["gmv"])
    order_key = _find_key(sample, KPI_ALIASES["order_count"])
    viewer_key = _find_key(sample, KPI_ALIASES["viewer_count"])
    like_key = _find_key(sample, KPI_ALIASES["like_count"])
    comment_key = _find_key(sample, KPI_ALIASES["comment_count"])
    share_key = _find_key(sample, KPI_ALIASES["share_count"])
    follower_key = _find_key(sample, KPI_ALIASES["new_followers"])
    click_key = _find_key(sample, KPI_ALIASES["product_clicks"])
    conv_key = _find_key(sample, KPI_ALIASES["ctor"])
    gpm_key = _find_key(sample, KPI_ALIASES["gpm"])

    # Build timed entries
    timed_entries = []
    for entry in trends:
        t_sec = _parse_time_to_seconds(entry.get(time_key))
        if t_sec is not None:
            timed_entries.append({"time_sec": t_sec, "entry": entry})
    timed_entries.sort(key=lambda x: x["time_sec"])
    if not timed_entries:
        return []

    csv_first_sec = timed_entries[0]["time_sec"]

    # Build CSV slots
    csv_slots = []
    for i, te in enumerate(timed_entries):
        slot_start = te["time_sec"]
        if i + 1 < len(timed_entries):
            slot_end = timed_entries[i + 1]["time_sec"]
        else:
            video_end_abs = csv_first_sec + time_offset_seconds + phases[-1]["time_end"]
            slot_end = max(slot_start + 1800, video_end_abs)
        csv_slots.append({"start": slot_start, "end": slot_end, "entry": te["entry"]})

    # Calculate metrics for each phase
    updates = []
    for p in phases:
        phase_abs_start = csv_first_sec + time_offset_seconds + p["time_start"]
        phase_abs_end = csv_first_sec + time_offset_seconds + p["time_end"]

        phase_gmv = 0.0
        phase_orders = 0.0
        phase_viewers = 0
        phase_likes = 0
        phase_comments = 0.0
        phase_shares = 0.0
        phase_followers = 0.0
        phase_clicks = 0.0
        phase_conv = 0.0
        phase_gpm = 0.0

        for slot in csv_slots:
            overlap_start = max(phase_abs_start, slot["start"])
            overlap_end = min(phase_abs_end, slot["end"])
            overlap = max(0, overlap_end - overlap_start)
            if overlap <= 0:
                continue
            slot_dur = max(slot["end"] - slot["start"], 1)
            ratio = overlap / slot_dur
            e = slot["entry"]

            if gmv_key:
                phase_gmv += (_safe_float(e.get(gmv_key)) or 0) * ratio
            if order_key:
                phase_orders += (_safe_float(e.get(order_key)) or 0) * ratio
            if comment_key:
                phase_comments += (_safe_float(e.get(comment_key)) or 0) * ratio
            if share_key:
                phase_shares += (_safe_float(e.get(share_key)) or 0) * ratio
            if follower_key:
                phase_followers += (_safe_float(e.get(follower_key)) or 0) * ratio
            if click_key:
                phase_clicks += (_safe_float(e.get(click_key)) or 0) * ratio
            if viewer_key:
                phase_viewers = max(phase_viewers, int(_safe_float(e.get(viewer_key)) or 0))
            if like_key:
                phase_likes = max(phase_likes, int(_safe_float(e.get(like_key)) or 0))
            if conv_key:
                phase_conv = max(phase_conv, _safe_float(e.get(conv_key)) or 0)
            if gpm_key:
                phase_gpm = max(phase_gpm, _safe_float(e.get(gpm_key)) or 0)

        updates.append({
            "phase_index": p["phase_index"],
            "gmv": round(phase_gmv, 2),
            "order_count": int(round(phase_orders)),
            "viewer_count": phase_viewers,
            "like_count": phase_likes,
            "comment_count": int(round(phase_comments)),
            "share_count": int(round(phase_shares)),
            "new_followers": int(round(phase_followers)),
            "product_clicks": int(round(phase_clicks)),
            "conversion_rate": round(phase_conv, 6),
            "gpm": round(phase_gpm, 2),
        })

    # Persist to DB
    for u in updates:
        try:
            await db.execute(text("""
                UPDATE video_phases
                SET gmv = :gmv, order_count = :order_count,
                    viewer_count = :viewer_count, like_count = :like_count,
                    comment_count = :comment_count, share_count = :share_count,
                    new_followers = :new_followers, product_clicks = :product_clicks,
                    conversion_rate = :conversion_rate, gpm = :gpm,
                    updated_at = now()
                WHERE video_id = :video_id AND phase_index = :phase_index
            """), {
                "video_id": video_id,
                "phase_index": u["phase_index"],
                "gmv": u["gmv"],
                "order_count": u["order_count"],
                "viewer_count": u["viewer_count"],
                "like_count": u["like_count"],
                "comment_count": u["comment_count"],
                "share_count": u["share_count"],
                "new_followers": u["new_followers"],
                "product_clicks": u["product_clicks"],
                "conversion_rate": u["conversion_rate"],
                "gpm": u["gpm"],
            })
        except Exception:
            pass
    try:
        await db.commit()
    except Exception:
        pass

    return updates


@router.get("/{video_id}")
async def get_video_detail(
    video_id: str,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_user),
):
    """
        Video detail endpoint returning report 1 data.
        Optimized: single combined query, inline SAS generation, no ORM overhead.
    """
    import time as _time
    import os as _os
    from azure.storage.blob import generate_blob_sas as _generate_blob_sas, BlobSasPermissions as _BlobSasPermissions

    try:
        _t0 = _time.monotonic()

        # ---- Step 1: Single query to get video + user email ----
        sql_video = text("""
            SELECT v.id, v.original_filename, v.status, v.user_id,
                   v.upload_type, v.excel_product_blob_url, v.excel_trend_blob_url,
                   v.compressed_blob_url, v.time_offset_seconds,
                   v.brand_client_id,
                   u.email
            FROM videos v
            JOIN users u ON v.user_id = u.id
            WHERE v.id = :video_id
        """)
        vres = await db.execute(sql_video, {"video_id": video_id})
        video_row = vres.fetchone()
        if not video_row:
            raise HTTPException(status_code=404, detail="Video not found")

        # Any authenticated user can view video details (removed owner-only restriction)

        email = video_row.email
        compressed_blob = video_row.compressed_blob_url
        _t1 = _time.monotonic()

        # ---- Step 2: Parallel fetch phase_insights + video_phases + video_insights ----
        sql_combined = text("""
            SELECT
                vp.id as phase_id, vp.phase_index, vp.phase_description,
                vp.time_start, vp.time_end,
                COALESCE(vp.gmv, 0) as gmv,
                COALESCE(vp.order_count, 0) as order_count,
                COALESCE(vp.viewer_count, 0) as viewer_count,
                COALESCE(vp.like_count, 0) as like_count,
                COALESCE(vp.comment_count, 0) as comment_count,
                COALESCE(vp.share_count, 0) as share_count,
                COALESCE(vp.new_followers, 0) as new_followers,
                COALESCE(vp.product_clicks, 0) as product_clicks,
                COALESCE(vp.conversion_rate, 0) as conversion_rate,
                COALESCE(vp.gpm, 0) as gpm,
                COALESCE(vp.importance_score, 0) as importance_score,
                vp.product_names,
                vp.user_rating,
                vp.user_comment,
                vp.sas_token,
                vp.sas_expireddate,
                vp.cta_score,
                vp.audio_features,
                vp.sales_psychology_tags,
                vp.human_sales_tags,
                pi.insight
            FROM video_phases vp
            LEFT JOIN phase_insights pi ON pi.video_id = vp.video_id AND pi.phase_index = vp.phase_index
            WHERE vp.video_id = :video_id
            ORDER BY vp.phase_index ASC
        """)

        sql_latest_insight = text("""
            SELECT title, content
            FROM video_insights
            WHERE video_id = :video_id
            ORDER BY created_at DESC
            LIMIT 1
        """)

        # Execute both queries concurrently
        # Fallback: if cta_score/audio_features columns don't exist yet, retry without them
        has_cta_columns = True
        try:
            combined_task = db.execute(sql_combined, {"video_id": video_id})
            insight_task = db.execute(sql_latest_insight, {"video_id": video_id})
            combined_res, insight_res = await asyncio.gather(combined_task, insight_task)
        except Exception:
            has_cta_columns = False
            await db.rollback()
            sql_combined_fallback = text("""
                SELECT
                    vp.id as phase_id, vp.phase_index, vp.phase_description,
                    vp.time_start, vp.time_end,
                    COALESCE(vp.gmv, 0) as gmv,
                    COALESCE(vp.order_count, 0) as order_count,
                    COALESCE(vp.viewer_count, 0) as viewer_count,
                    COALESCE(vp.like_count, 0) as like_count,
                    COALESCE(vp.comment_count, 0) as comment_count,
                    COALESCE(vp.share_count, 0) as share_count,
                    COALESCE(vp.new_followers, 0) as new_followers,
                    COALESCE(vp.product_clicks, 0) as product_clicks,
                    COALESCE(vp.conversion_rate, 0) as conversion_rate,
                    COALESCE(vp.gpm, 0) as gpm,
                    COALESCE(vp.importance_score, 0) as importance_score,
                    vp.product_names,
                    vp.user_rating,
                    vp.user_comment,
                    vp.sas_token,
                    vp.sas_expireddate,
                    NULL as cta_score,
                    NULL as audio_features,
                    NULL as sales_psychology_tags,
                    NULL as human_sales_tags,
                    pi.insight
                FROM video_phases vp
                LEFT JOIN phase_insights pi ON pi.video_id = vp.video_id AND pi.phase_index = vp.phase_index
                WHERE vp.video_id = :video_id
                ORDER BY vp.phase_index ASC
            """)
            combined_task = db.execute(sql_combined_fallback, {"video_id": video_id})
            insight_task = db.execute(sql_latest_insight, {"video_id": video_id})
            combined_res, insight_res = await asyncio.gather(combined_task, insight_task)

        combined_rows = combined_res.fetchall()
        latest_insight = insight_res.fetchone()
        _t2 = _time.monotonic()

        # ---- Step 3: Build SAS URLs inline (no async service call needed) ----
        conn_str = _os.getenv("AZURE_STORAGE_CONNECTION_STRING", "")
        account_name = _os.getenv("AZURE_STORAGE_ACCOUNT_NAME", "")
        container_name = _os.getenv("AZURE_BLOB_CONTAINER", "videos")
        account_key = ""
        for part in conn_str.split(";"):
            if part.startswith("AccountKey="):
                account_key = part.split("=", 1)[1]
                break

        now_utc = datetime.now(timezone.utc)
        now_naive = datetime.utcnow()
        sas_expiry = now_utc + timedelta(days=7)
        phases_needing_sas_update = []  # (phase_id, sas_url, expiry)

        def _make_sas_url(blob_name: str) -> str:
            """Generate SAS URL locally without any async/HTTP call."""
            sas = _generate_blob_sas(
                account_name=account_name,
                container_name=container_name,
                blob_name=blob_name,
                account_key=account_key,
                permission=_BlobSasPermissions(read=True),
                expiry=sas_expiry,
            )
            url = f"https://{account_name}.blob.core.windows.net/{container_name}/{blob_name}?{sas}"
            return _replace_blob_url_to_cdn(url)

        report1_items = []
        for r in combined_rows:
            # Check cached SAS
            video_clip_url = None
            if email and r.time_start is not None and r.time_end is not None:
                sas_token = r.sas_token
                sas_expire = r.sas_expireddate
                cache_valid = False
                if sas_token and sas_expire:
                    try:
                        if sas_expire.tzinfo is not None and sas_expire.tzinfo.utcoffset(sas_expire) is not None:
                            cache_valid = sas_expire.astimezone(timezone.utc) >= now_utc
                        else:
                            cache_valid = sas_expire >= now_naive
                    except Exception as _e:
                        logger.debug(f"Non-critical error suppressed: {_e}")

                if cache_valid:
                    video_clip_url = sas_token
                elif account_key:
                    try:
                        ts = float(r.time_start)
                        te = float(r.time_end)
                        fname = f"{ts:.1f}_{te:.1f}.mp4"
                        blob_name = f"{email}/{video_id}/reportvideo/{fname}"
                        video_clip_url = _make_sas_url(blob_name)
                        if r.phase_id:
                            phases_needing_sas_update.append((r.phase_id, video_clip_url, sas_expiry))
                    except Exception:
                        video_clip_url = None

            # Parse product_names
            product_names_list = []
            pn_raw = r.product_names
            if pn_raw:
                try:
                    product_names_list = json.loads(pn_raw) if isinstance(pn_raw, str) else pn_raw
                except (json.JSONDecodeError, TypeError):
                    product_names_list = []

            # Only include phases that have insights (matching original behavior)
            if r.insight is not None:
                # Parse audio_features JSON text
                audio_features_parsed = None
                try:
                    if r.audio_features:
                        audio_features_parsed = json.loads(r.audio_features) if isinstance(r.audio_features, str) else r.audio_features
                except (json.JSONDecodeError, TypeError) as _e:
                    logger.debug(f"JSON parse skipped: {_e}")

                # Parse sales_psychology_tags JSON text
                sales_tags_parsed = []
                try:
                    raw_tags = getattr(r, 'sales_psychology_tags', None)
                    if raw_tags:
                        sales_tags_parsed = json.loads(raw_tags) if isinstance(raw_tags, str) else raw_tags
                except (json.JSONDecodeError, TypeError) as _e:
                    logger.debug(f"JSON parse skipped: {_e}")

                # Parse human_sales_tags JSON text
                human_tags_parsed = None
                try:
                    raw_human_tags = getattr(r, 'human_sales_tags', None)
                    if raw_human_tags:
                        human_tags_parsed = json.loads(raw_human_tags) if isinstance(raw_human_tags, str) else raw_human_tags
                except (json.JSONDecodeError, TypeError) as _e:
                    logger.debug(f"JSON parse skipped: {_e}")

                report1_items.append({
                    "phase_index": int(r.phase_index),
                    "phase_description": r.phase_description,
                    "time_start": r.time_start,
                    "time_end": r.time_end,
                    "insight": r.insight,
                    "video_clip_url": video_clip_url,
                    "user_rating": r.user_rating,
                    "user_comment": r.user_comment,
                    "cta_score": getattr(r, 'cta_score', None),
                    "audio_features": audio_features_parsed,
                    "sales_psychology_tags": sales_tags_parsed,
                    "human_sales_tags": human_tags_parsed,
                    "csv_metrics": {
                        "gmv": r.gmv,
                        "order_count": r.order_count,
                        "viewer_count": r.viewer_count,
                        "like_count": r.like_count,
                        "comment_count": r.comment_count,
                        "share_count": r.share_count,
                        "new_followers": r.new_followers,
                        "product_clicks": r.product_clicks,
                        "conversion_rate": r.conversion_rate,
                        "gpm": r.gpm,
                        "importance_score": r.importance_score,
                        "product_names": product_names_list,
                    },
                })

        _t3 = _time.monotonic()

        # ---- Step 3.5: Auto-recalc CSV metrics if all zero (lazy fix) ----
        if (
            video_row.status == "DONE"
            and report1_items
            and video_row.excel_trend_blob_url
            and all(item["csv_metrics"]["viewer_count"] == 0 for item in report1_items)
        ):
            try:
                recalc_result = await _auto_recalc_csv_metrics(
                    db, video_id, video_row.excel_trend_blob_url,
                    float(getattr(video_row, 'time_offset_seconds', 0) or 0),
                )
                if recalc_result:
                    # Update report1_items in-place with recalculated metrics
                    recalc_map = {u["phase_index"]: u for u in recalc_result}
                    for item in report1_items:
                        u = recalc_map.get(item["phase_index"])
                        if u:
                            item["csv_metrics"].update({
                                "gmv": u["gmv"],
                                "order_count": u["order_count"],
                                "viewer_count": u["viewer_count"],
                                "like_count": u["like_count"],
                                "comment_count": u["comment_count"],
                                "share_count": u["share_count"],
                                "new_followers": u["new_followers"],
                                "product_clicks": u["product_clicks"],
                                "conversion_rate": u["conversion_rate"],
                                "gpm": u["gpm"],
                            })
                    logger.info(f"[AUTO-RECALC] Recalculated CSV metrics for video {video_id}")
            except Exception as _recalc_err:
                logger.warning(f"[AUTO-RECALC] Failed for video {video_id}: {_recalc_err}")

        # ---- Step 4: Batch persist new SAS tokens (fire-and-forget style) ----
        if phases_needing_sas_update:
            try:
                for pid, sas_url, exp_at in phases_needing_sas_update:
                    await db.execute(
                        text("UPDATE video_phases SET sas_token = :sas, sas_expireddate = :exp WHERE id = :id"),
                        {"sas": sas_url, "exp": exp_at, "id": pid}
                    )
                await db.commit()
            except Exception as _e:
                logger.debug(f"Non-critical error suppressed: {_e}")  # Non-critical

        # ---- Step 5: Build report3 ----
        report3 = []
        if latest_insight:
            parsed = latest_insight.content
            try:
                if isinstance(parsed, str):
                    s = parsed.lstrip()
                    if s.startswith("{") or s.startswith("["):
                        parsed = json.loads(parsed)

                if isinstance(parsed, dict) and parsed.get("video_insights") and isinstance(parsed.get("video_insights"), list):
                    for item in parsed.get("video_insights"):
                        report3.append({"title": item.get("title"), "content": item.get("content")})
                elif isinstance(parsed, list):
                    for item in parsed:
                        report3.append({"title": item.get("title"), "content": item.get("content")})
                else:
                    report3.append({"title": latest_insight.title, "content": latest_insight.content})
            except Exception:
                report3.append({"title": latest_insight.title, "content": latest_insight.content})

        # ---- Step 6: Generate preview URL (inline, no service call) ----
        # BUILD 41 FIX: compressed_blob_url has two formats depending on source:
        #   clean_video (worker): full path  "email/video_id/video_id_preview.mp4"
        #   live_boost (pipeline): relative  "assembled/VIDEO_ID_assembled.mp4"
        # Also, iOS generates UPPERCASE UUIDs (UUID().uuidString) while PostgreSQL
        # normalises to lowercase. Blob Storage paths are case-sensitive, so we
        # must reconstruct the path using the original case from the blob URL.
        preview_url = None
        if compressed_blob and email and account_key:
            try:
                # Detect if compressed_blob already contains the full path
                # (i.e. starts with email or contains 3+ path segments)
                segments = compressed_blob.split("/")
                if "@" in segments[0] or len(segments) >= 3:
                    # Full path — use as-is (clean_video / worker pipeline)
                    blob_name = compressed_blob
                else:
                    # Relative path under email/video_id/ (live_boost pipeline)
                    # The video_id folder on Blob was created by iOS with UPPERCASE UUID.
                    # Extract the original case from the filename in compressed_blob_url
                    # e.g. "assembled/8E8C6B5F-..._assembled.mp4" → "8E8C6B5F-..."
                    fname = segments[-1]  # e.g. "8E8C6B5F-..._assembled.mp4"
                    # Try to extract UUID from filename (before first underscore after UUID pattern)
                    import re
                    uuid_match = re.search(r'([0-9A-Fa-f]{8}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{12})', fname)
                    if uuid_match:
                        original_case_vid = uuid_match.group(1)
                    else:
                        # Fallback: use video_id as-is (lowercase from DB)
                        original_case_vid = video_id
                    blob_name = f"{email}/{original_case_vid}/{compressed_blob}"
                # BUILD 42: Use direct Blob URL (not CDN) for preview_url.
                # CDN may interfere with AVPlayer streaming (Range requests,
                # SAS token handling, or caching issues).
                _preview_sas = _generate_blob_sas(
                    account_name=account_name,
                    container_name=container_name,
                    blob_name=blob_name,
                    account_key=account_key,
                    permission=_BlobSasPermissions(read=True),
                    expiry=sas_expiry,
                )
                preview_url = f"https://{account_name}.blob.core.windows.net/{container_name}/{blob_name}?{_preview_sas}"
                logger.debug(f"[preview_url] blob_name={blob_name} (direct blob, no CDN)")
            except Exception as exc:
                logger.warning(f"[preview_url] Failed to generate SAS: {exc}")
                preview_url = None

        _t_end = _time.monotonic()
        _perf = {
            "video_query_ms": round((_t1-_t0)*1000),
            "combined_query_ms": round((_t2-_t1)*1000),
            "build_response_ms": round((_t3-_t2)*1000),
            "total_ms": round((_t_end-_t0)*1000),
            "phase_count": len(combined_rows),
            "sas_generated": len(phases_needing_sas_update),
        }
        logger.info(f"[PERF] {_perf}")

        return {
            "id": str(video_row.id),
            "original_filename": video_row.original_filename,
            "status": video_row.status,
            "step_progress": getattr(video_row, 'step_progress', None) or 0,
            "upload_type": video_row.upload_type,
            "excel_product_blob_url": video_row.excel_product_blob_url,
            "excel_trend_blob_url": video_row.excel_trend_blob_url,
            "compressed_blob_url": compressed_blob,
            "preview_url": preview_url,
            "brand_client_id": getattr(video_row, 'brand_client_id', None),
            "reports_1": report1_items,
            "report3": report3,
            "_perf": _perf,
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(f"Failed to fetch video detail: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch video detail: {exc}")


@router.get("/{video_id}/product-data")
async def get_video_product_data(
    video_id: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Fetch and parse the product Excel file for a video.
    Returns parsed product data as JSON.
    Uses SAS tokens to access Azure Blob Storage (public access is disabled).
    """
    try:
        import httpx
        import tempfile
        from app.services.storage_service import generate_read_sas_from_url

        # Get video's excel_product_blob_url and user email
        result = await db.execute(
            text("""
                SELECT v.excel_product_blob_url, v.excel_trend_blob_url, u.email
                FROM videos v
                JOIN users u ON v.user_id = u.id
                WHERE v.id = :vid
            """),
            {"vid": video_id},
        )
        row = result.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Video not found")

        product_blob_url = row[0]
        trend_blob_url = row[1]
        email = row[2]
        logger.info("[PRODUCT-DATA] video=%s product_url=%s trend_url=%s email=%s",
                    video_id, product_blob_url is not None, trend_blob_url is not None, email)

        response_data = {
            "products": [],
            "trends": [],
            "has_product_data": False,
            "has_trend_data": False,
        }

        # Helper: download and parse Excel file
        async def _parse_excel(blob_url: str) -> list:
            """Download Excel via SAS URL and parse rows into list of dicts."""
            sas_url = generate_read_sas_from_url(blob_url, expires_hours=1)
            logger.info("[PRODUCT-DATA] blob_url=%s sas_generated=%s", blob_url[:80] if blob_url else None, sas_url is not None)
            if not sas_url:
                logger.warning("Failed to generate SAS for Excel blob: %s", blob_url[:100] if blob_url else None)
                return []
            import openpyxl
            async with httpx.AsyncClient(timeout=30.0) as client:
                resp = await client.get(sas_url)
                if resp.status_code != 200:
                    logger.warning(f"Failed to download Excel (HTTP {resp.status_code}): {sas_url[:100]}...")
                    return []

                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
                    f.write(resp.content)
                    tmp_path = f.name

                try:
                    wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
                    ws = wb.active
                    items = []
                    if ws:
                        rows_data = list(ws.iter_rows(values_only=True))
                        if len(rows_data) >= 2:
                            headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows_data[0])]
                            for data_row in rows_data[1:]:
                                if all(v is None for v in data_row):
                                    continue
                                item = {}
                                for i, val in enumerate(data_row):
                                    if i < len(headers):
                                        if val is None:
                                            item[headers[i]] = None
                                        elif isinstance(val, (int, float)):
                                            item[headers[i]] = val
                                        else:
                                            item[headers[i]] = str(val)
                                items.append(item)
                    wb.close()
                    return items
                finally:
                    os.unlink(tmp_path)

        # Parse product Excel
        if product_blob_url:
            try:
                logger.info("[PRODUCT-DATA] Parsing product Excel: %s", product_blob_url[:100])
                products = await _parse_excel(product_blob_url)
                logger.info("[PRODUCT-DATA] Product parse result: %d items", len(products))
                response_data["products"] = products
                response_data["has_product_data"] = len(products) > 0

                # Cache top 2 products by GMV in videos table
                if products:
                    try:
                        # Detect GMV and name columns
                        gmv_key = None
                        name_key = None
                        sample = products[0]
                        for k in sample.keys():
                            kl = k.lower() if k else ""
                            if "gmv" in kl:
                                gmv_key = k
                            if "商品名" in k or "product" in kl or "name" in kl:
                                name_key = k
                        if gmv_key and name_key:
                            sorted_products = sorted(
                                products,
                                key=lambda x: float(x.get(gmv_key, 0) or 0),
                                reverse=True,
                            )
                            top2 = []
                            for p in sorted_products[:2]:
                                pname = p.get(name_key, "")
                                if pname:
                                    # Truncate long product names
                                    pname = str(pname)[:50]
                                    top2.append(pname)
                            if top2:
                                import json as _json
                                await db.execute(
                                    text("UPDATE videos SET top_products = :tp WHERE id = :vid"),
                                    {"tp": _json.dumps(top2, ensure_ascii=False), "vid": video_id},
                                )
                                await db.commit()
                                logger.info(f"Cached top_products for video {video_id}: {top2}")
                    except Exception as cache_err:
                        logger.warning(f"Failed to cache top_products: {cache_err}")
            except Exception as e:
                logger.warning(f"Failed to parse product Excel: {e}", exc_info=True)

        else:
            logger.info("[PRODUCT-DATA] No product_blob_url")

        # Parse trend Excel
        if trend_blob_url:
            try:
                logger.info("[PRODUCT-DATA] Parsing trend Excel: %s", trend_blob_url[:100])
                trends = await _parse_excel(trend_blob_url)
                logger.info("[PRODUCT-DATA] Trend parse result: %d items", len(trends))
                response_data["trends"] = trends
                response_data["has_trend_data"] = len(trends) > 0
            except Exception as e:
                logger.warning(f"Failed to parse trend Excel: {e}", exc_info=True)
        return response_data

    except HTTPException:
        raise
    except Exception as exc:
        logger.error(f"Failed to fetch product data: {exc}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to fetch product data: {exc}")



@router.put("/{video_id}/phases/{phase_index}/rating")
async def rate_phase(
    video_id: str,
    phase_index: int,
    request_body: dict,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    """
    Save a human rating (1-5) and optional comment for a specific phase.
    Also updates the quality_score in Qdrant for RAG learning.

    Body:
    {
        "rating": 1-5,
        "comment": "optional text"
    }
    """
    try:
        user_id = user.get("user_id") or user.get("id")
        rating = request_body.get("rating")
        comment = request_body.get("comment", "")
        reviewer_name = request_body.get("reviewer_name", "")

        if rating is None or not isinstance(rating, int) or rating < 1 or rating > 5:
            raise HTTPException(status_code=400, detail="rating must be an integer between 1 and 5")

        # Verify video belongs to user
        video_repo = VideoRepository(lambda: db)
        video = await video_repo.get_video_by_id(video_id)
        if not video:
            raise HTTPException(status_code=404, detail="Video not found")
        # Any authenticated user can rate phases (removed owner-only restriction)

        # Validate rating(1-5) to importance_score (0.0-1.0)
        importance_score = (rating - 1) / 4.0  # 1->0.0, 2->0.25, 3->0.5, 4->0.75, 5->1.0

        # Update video_phases with user rating, comment, and importance_score
        # Use try-except for graceful fallback if columns don't exist yet
        try:
            sql_update = text("""
                UPDATE video_phases
                SET user_rating = :rating,
                    user_comment = :comment,
                    importance_score = :importance_score,
                    rated_at = NOW(),
                    updated_at = NOW()
                WHERE video_id = :video_id AND phase_index = :phase_index
            """)
            await db.execute(sql_update, {
                "rating": rating,
                "comment": comment,
                "importance_score": importance_score,
                "video_id": video_id,
                "phase_index": phase_index,
            })
            await db.commit()
        except Exception as db_err:
            await db.rollback()
            # Fallback: try without user_rating/user_comment columns
            try:
                sql_fallback = text("""
                    UPDATE video_phases
                    SET importance_score = :importance_score,
                        updated_at = NOW()
                    WHERE video_id = :video_id AND phase_index = :phase_index
                """)
                await db.execute(sql_fallback, {
                    "importance_score": importance_score,
                    "video_id": video_id,
                    "phase_index": phase_index,
                })
                await db.commit()
            except Exception:
                await db.rollback()
                logger.warning(f"Could not update video_phases for rating: {db_err}")

        # Update Qdrant quality_score for RAG learning (in background for faster response)
        def _update_qdrant_bg(vid, pidx, r, c):
            try:
                from app.services.rag.knowledge_store import update_quality_score_with_comment
                update_quality_score_with_comment(
                    video_id=vid, phase_index=pidx, rating=r, comment=c,
                )
            except ImportError:
                try:
                    from app.services.rag.knowledge_store import update_quality_score
                    old_rating = 1 if r >= 4 else (-1 if r <= 2 else 0)
                    update_quality_score(video_id=vid, phase_index=pidx, rating=old_rating)
                except Exception as rag_err:
                    logger.warning(f"Could not update Qdrant quality_score: {rag_err}")
            except Exception as rag_err:
                logger.warning(f"Could not update Qdrant quality_score: {rag_err}")

        background_tasks.add_task(_update_qdrant_bg, video_id, phase_index, rating, comment)

        logger.info(f"Phase rated: video={video_id}, phase={phase_index}, rating={rating}, comment={comment[:50] if comment else ''}")

        return {
            "success": True,
            "video_id": video_id,
            "phase_index": phase_index,
            "rating": rating,
            "comment": comment,
            "importance_score": importance_score,
            "reviewer_name": reviewer_name,
        }

    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(f"Failed to rate phase: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to rate phase: {exc}")


# =========================================================
# Phase Comment API (save comment without requiring rating)
# =========================================================

@router.put("/{video_id}/phases/{phase_index}/comment")
async def save_phase_comment(
    video_id: str,
    phase_index: int,
    request_body: dict,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    """
    Save a comment for a specific phase (rating not required).
    Body:
    {
        "comment": "text",
        "reviewer_name": "optional"
    }
    """
    try:
        user_id = user.get("user_id") or user.get("id")
        comment = request_body.get("comment", "")
        reviewer_name = request_body.get("reviewer_name", "")

        # Verify video belongs to user
        video_repo = VideoRepository(lambda: db)
        video = await video_repo.get_video_by_id(video_id)
        if not video:
            raise HTTPException(status_code=404, detail="Video not found")
        # Any authenticated user can comment on phases (removed owner-only restriction)

        # Upsert comment
        sql_update = text("""
            UPDATE video_phases
            SET user_comment = :comment,
                updated_at = NOW()
            WHERE video_id = :video_id AND phase_index = :phase_index
        """)
        result = await db.execute(sql_update, {
            "comment": comment,
            "video_id": video_id,
            "phase_index": phase_index,
        })
        await db.commit()

        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="Phase not found")

        logger.info(f"Phase comment saved: video={video_id}, phase={phase_index}, comment={comment[:50] if comment else ''}")

        return {
            "success": True,
            "video_id": video_id,
            "phase_index": phase_index,
            "comment": comment,
        }

    except HTTPException:
        raise
    except Exception as exc:
        await db.rollback()
        logger.exception(f"Failed to save phase comment: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to save phase comment: {exc}")


# =========================================================
# Human Sales Tags API (Human-in-the-loop)
# =========================================================

ALL_SALES_TAGS = {
    # Sales psychology tags
    "HOOK", "EMPATHY", "PROBLEM", "EDUCATION", "SOLUTION",
    "DEMONSTRATION", "COMPARISON", "PROOF", "TRUST", "SOCIAL_PROOF",
    "OBJECTION_HANDLING", "URGENCY", "LIMITED_OFFER", "BONUS", "CTA",
    # Phase behavior tags
    "CHAT", "PREP", "PHONE_OP", "LONG_GREET",
    "COMMENT_READ", "SILENCE", "PRICE_SHOW",
}


@router.patch("/{video_id}/phases/{phase_index}/tags")
async def update_human_sales_tags(
    video_id: str,
    phase_index: int,
    request_body: dict,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    """
    Save human-corrected sales psychology tags for a specific phase.
    Body:
    {
        "human_sales_tags": ["HOOK", "EMPATHY", "CTA"]
    }
    """
    try:
        user_id = user.get("user_id") or user.get("id")
        tags = request_body.get("human_sales_tags")
        reviewer_name = request_body.get("reviewer_name", "")

        if tags is None or not isinstance(tags, list):
            raise HTTPException(status_code=400, detail="human_sales_tags must be a list of tag strings")

        # Validate tags
        invalid = [t for t in tags if t not in ALL_SALES_TAGS]
        if invalid:
            raise HTTPException(status_code=400, detail=f"Invalid tags: {invalid}. Valid: {sorted(ALL_SALES_TAGS)}")

        # Verify video belongs to user
        video_repo = VideoRepository(lambda: db)
        video = await video_repo.get_video_by_id(video_id)
        if not video:
            raise HTTPException(status_code=404, detail="Video not found")
        # Any authenticated user can tag phases (removed owner-only restriction)

        tags_json = json.dumps(tags)

        sql_update = text("""
            UPDATE video_phases
            SET human_sales_tags = :tags,
                updated_at = NOW()
            WHERE video_id = :video_id AND phase_index = :phase_index
        """)
        result = await db.execute(sql_update, {
            "tags": tags_json,
            "video_id": video_id,
            "phase_index": phase_index,
        })
        await db.commit()

        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="Phase not found")

        logger.info(f"Human tags saved: video={video_id}, phase={phase_index}, tags={tags}")

        return {
            "success": True,
            "video_id": video_id,
            "phase_index": phase_index,
            "human_sales_tags": tags,
        }

    except HTTPException:
        raise
    except Exception as exc:
        await db.rollback()
        logger.exception(f"Failed to save human sales tags: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to save human sales tags: {exc}")


# =========================================================
# ============================================================
# TikTok Live Capture Endpoints
# ============================================================

@router.post("/live-check", response_model=LiveCheckResponse)
async def live_check(
    payload: LiveCaptureRequest,
    current_user=Depends(get_current_user),
):
    """Check if a TikTok user is currently live."""
    from app.services.tiktok_service import TikTokLiveService

    try:
        info = await TikTokLiveService.check_and_get_info(payload.live_url)
        return LiveCheckResponse(
            is_live=info["is_live"],
            username=info.get("username"),
            room_id=info.get("room_id"),
            title=info.get("title"),
            message="LIVE" if info["is_live"] else "User is not currently live",
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as exc:
        logger.exception(f"Live check failed: {exc}")
        raise HTTPException(status_code=500, detail=f"Live check failed: {exc}")


@router.post("/live-capture", response_model=LiveCaptureResponse)
async def live_capture(
    payload: LiveCaptureRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Start capturing a TikTok live stream.
    1. Validates the URL and checks if the user is live
    2. Creates a video record in the database
    3. Enqueues a live_capture job for the worker
    """
    from app.services.tiktok_service import TikTokLiveService
    from app.services.queue_service import enqueue_job

    # Step 1: Check live status
    try:
        info = await TikTokLiveService.check_and_get_info(payload.live_url)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as exc:
        logger.exception(f"Live check failed: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to check live status: {exc}")

    if not info["is_live"]:
        raise HTTPException(
            status_code=400,
            detail=f"@{info.get('username', 'unknown')} is not currently live",
        )

    username = info["username"]
    title = info.get("title", "")

    # Step 2: Create video record
    video_id = str(uuid_module.uuid4())
    original_filename = f"tiktok_live_{username}.mp4"

    try:
        video_repo = VideoRepository(lambda: db)
        service = VideoService(video_repository=video_repo)

        video = await video_repo.create_video(
            user_id=current_user["id"],
            video_id=video_id,
            original_filename=original_filename,
            status="capturing",
            upload_type="live_capture",
        )
        await db.commit()
    except Exception as exc:
        logger.exception(f"Failed to create video record: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to create video record: {exc}")

    # Step 3: Enqueue live_capture job
    try:
        queue_payload = {
            "job_type": "live_capture",
            "video_id": video_id,
            "live_url": payload.live_url,
            "email": current_user["email"],
            "user_id": current_user["id"],
            "duration": payload.duration or 0,
            "username": username,
            "stream_title": title,
        }
        await enqueue_job(queue_payload)
    except Exception as exc:
        logger.exception(f"Failed to enqueue live capture job: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to start capture: {exc}")

    return LiveCaptureResponse(
        video_id=video_id,
        status="capturing",
        stream_title=title,
        username=username,
        message=f"Live capture started for @{username}; recording and analysis will begin automatically",
    )

# =========================================================
# =========================================================
# Retry Analysis API (user-facing)
# =========================================================

@router.post("/{video_id}/retry-analysis")
async def retry_analysis(
    video_id: str,
    db: AsyncSession = Depends(get_db),
    user=Depends(get_current_user),
):
    """
    Re-enqueue a failed video for analysis without re-uploading.
    The uploaded video asset is preserved in Blob storage.
    Only the analysis job is re-submitted.

    Supports both standard videos AND LiveBoost (live_boost) videos:
    - Standard videos: enqueues a standard video_analysis job with blob_url
    - LiveBoost videos: enqueues a live_analysis job (chunks are in blob storage)
    """
    try:
        user_id = user.get("user_id") or user.get("id")

        # Verify video exists and belongs to user
        sql = text("""
            SELECT v.id, v.original_filename, v.status, v.user_id,
                   v.upload_type, u.email as user_email
            FROM videos v
            LEFT JOIN users u ON v.user_id = u.id
            WHERE v.id = :vid
        """)
        result = await db.execute(sql, {"vid": video_id})
        row = result.fetchone()

        if not row:
            raise HTTPException(status_code=404, detail="Video not found")
        # Any authenticated user can retry analysis (removed owner-only restriction)

        # Determine analysis type: stuck QUEUED, stalled processing states
        # DONE/COMPLETED videos should NOT be retried — CSV metrics are
        # auto-recalculated on page view via _auto_recalc_csv_metrics.
        allowed_statuses = ("ERROR", "error", "uploaded", "UPLOADED", "QUEUED")
        # Also allow any STEP_* status (e.g. STEP_0_EXTRACT_FRAMES) that may be stalled
        is_stuck_step = row.status and row.status.startswith("STEP_")
        if row.status in ("DONE", "COMPLETED", "completed"):
            raise HTTPException(
                status_code=400,
                detail="この動画は解析完了済みです。データは自動的に最新化されます。",
            )
        if row.status not in allowed_statuses and not is_stuck_step:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot retry: video status is '{row.status}'. "
                       f"Retry is only available for failed or stuck videos.",
            )

        previous_status = row.status
        upload_type = row.upload_type or ""

        # ── LiveBoost (live_boost) videos: use live_analysis pipeline ──
        if upload_type == "live_boost":
            return await _retry_live_boost_analysis(
                db=db,
                video_id=video_id,
                user_id=user_id,
                user_email=row.user_email,
                previous_status=previous_status,
            )

        # ── Standard videos: use standard video_analysis pipeline ──
        # Generate fresh SAS URL for the existing blob
        from app.services.storage_service import generate_download_sas
        download_url, expiry = await generate_download_sas(
            email=row.user_email,
            video_id=str(row.id),
            filename=row.original_filename,
            expires_in_minutes=1440,  # 24 hours
        )

        # Determine resume status: keep current STEP_* status for resume,
        # only reset to 'uploaded' if status is ERROR or non-STEP
        is_step_status = previous_status and previous_status.startswith("STEP_")

        if is_step_status:
            # Keep the STEP_* status so worker can resume from this step
            resume_status = previous_status
            await db.execute(
                text("""
                    UPDATE videos
                    SET step_progress = 0,
                        error_message = NULL
                    WHERE id = :vid
                """),
                {"vid": video_id},
            )
        else:
            # ERROR or other status: try to resume from the error step
            # instead of restarting from scratch.
            # NEVER fall back to 'uploaded' — use STEP_0 as the safe minimum
            # so the worker can still leverage cached artifacts.
            resume_status = 'STEP_0_EXTRACT_FRAMES'  # safe fallback (not 'uploaded')
            try:
                err_result = await db.execute(
                    text("""
                        SELECT error_step FROM video_error_logs
                        WHERE video_id = :vid
                        ORDER BY created_at DESC LIMIT 1
                    """),
                    {"vid": video_id},
                )
                err_row = err_result.fetchone()
                if (
                    err_row
                    and err_row.error_step
                    and err_row.error_step.startswith("STEP_")
                ):
                    resume_status = err_row.error_step
                    logger.info(
                        "[retry-analysis] Resuming from error_step=%s for video %s",
                        resume_status, video_id,
                    )
                else:
                    logger.info(
                        "[retry-analysis] No valid error_step found, starting from STEP_0 for video %s",
                        video_id,
                    )
            except Exception as _e:
                logger.warning(
                    "[retry-analysis] Could not read error_step, starting from STEP_0: %s", _e
                )

            await db.execute(
                text("""
                    UPDATE videos
                    SET status = :status,
                        step_progress = 0,
                        error_message = NULL
                    WHERE id = :vid
                """),
                {"vid": video_id, "status": resume_status},
            )
        await db.commit()

        # Enqueue analysis job
        from app.services.queue_service import enqueue_job
        await enqueue_job({
            "video_id": str(row.id),
            "blob_url": download_url,
            "original_filename": row.original_filename,
        })

        logger.info(
            f"[retry-analysis] User {user_id} retried analysis for video {video_id} "
            f"(was: {previous_status}, resume_from: {resume_status})"
        )

        return {
            "success": True,
            "video_id": video_id,
            "message": f"解析を再開しました。{resume_status}から再開します。",
            "new_status": resume_status,
        }

    except HTTPException:
        raise
    except Exception as exc:
        await db.rollback()
        logger.exception(f"[retry-analysis] Failed: {exc}")
        raise HTTPException(status_code=500, detail=f"解析の再試行に失敗しました: {exc}")


async def _retry_live_boost_analysis(
    db: AsyncSession,
    video_id: str,
    user_id: int,
    user_email: str,
    previous_status: str,
) -> dict:
    """
    Retry analysis for LiveBoost (live_boost) videos.

    LiveBoost videos use a separate pipeline (live_analysis) that:
    1. Downloads chunks from blob storage
    2. Assembles them into a single video
    3. Runs STT, OCR, sales detection, clip generation

    This function:
    - Resets or creates the LiveAnalysisJob
    - Enqueues a live_analysis job to the worker queue
    - Updates the videos table status
    """
    from app.models.orm.live_analysis_job import LiveAnalysisJob
    from app.services.queue_service import enqueue_job

    try:
        # Check for existing LiveAnalysisJob
        result = await db.execute(
            select(LiveAnalysisJob).where(
                LiveAnalysisJob.video_id == video_id,
            ).order_by(LiveAnalysisJob.created_at.desc())
        )
        existing_job = result.scalars().first()

        if existing_job:
            # BUILD 36: Clean up duplicate jobs
            try:
                from sqlalchemy import delete as sa_delete
                await db.execute(
                    sa_delete(LiveAnalysisJob).where(
                        LiveAnalysisJob.video_id == video_id,
                        LiveAnalysisJob.id != existing_job.id,
                    )
                )
            except Exception:
                pass
            # Reset existing job for retry
            existing_job.status = "pending"
            existing_job.current_step = None
            existing_job.progress = 0
            existing_job.error_message = None
            existing_job.started_at = None
            existing_job.completed_at = None
            existing_job.results = None
            job = existing_job
            total_chunks = existing_job.total_chunks
            stream_source = existing_job.stream_source or "tiktok_live"
            logger.info(
                f"[retry-analysis/live_boost] Reset existing job {job.id} for video {video_id}"
            )
        else:
            # Create new LiveAnalysisJob
            job = LiveAnalysisJob(
                id=uuid_module.uuid4(),
                video_id=video_id,
                user_id=user_id,
                stream_source="tiktok_live",
                status="pending",
                progress=0,
            )
            db.add(job)
            total_chunks = None
            stream_source = "tiktok_live"
            logger.info(
                f"[retry-analysis/live_boost] Created new job for video {video_id}"
            )

        # Reset video status
        await db.execute(
            text("""
                UPDATE videos
                SET status = 'STEP_0_EXTRACT_FRAMES',
                    step_progress = 0,
                    error_message = NULL,
                    updated_at = now()
                WHERE id = :vid
            """),
            {"vid": video_id},
        )
        await db.commit()
        await db.refresh(job)

        # Enqueue live_analysis job
        queue_payload = {
            "job_type": "live_analysis",
            "job_id": str(job.id),
            "video_id": video_id,
            "user_id": user_id,
            "stream_source": stream_source,
            "total_chunks": total_chunks,
            "email": user_email or "",
        }
        enqueue_result = await enqueue_job(queue_payload)

        if enqueue_result.success:
            from sqlalchemy import update as sa_update
            await db.execute(
                sa_update(LiveAnalysisJob)
                .where(LiveAnalysisJob.id == job.id)
                .values(
                    queue_message_id=enqueue_result.message_id,
                    queue_enqueued_at=enqueue_result.enqueued_at,
                    started_at=datetime.now(timezone.utc),
                )
            )
            await db.commit()
            logger.info(
                f"[retry-analysis/live_boost] Enqueued OK job={job.id} video={video_id} "
                f"msg_id={enqueue_result.message_id}"
            )
        else:
            # Enqueue failed — mark job as failed
            await db.execute(
                sa_update(LiveAnalysisJob)
                .where(LiveAnalysisJob.id == job.id)
                .values(
                    status="failed",
                    error_message=f"Retry enqueue failed: {enqueue_result.error}",
                )
            )
            await db.execute(
                text("""
                    UPDATE videos
                    SET status = 'ERROR', updated_at = now()
                    WHERE id = :vid
                """),
                {"vid": video_id},
            )
            await db.commit()
            logger.error(
                f"[retry-analysis/live_boost] Enqueue FAILED for video {video_id}: "
                f"{enqueue_result.error}"
            )
            raise HTTPException(
                status_code=500,
                detail=f"LiveBoost解析の再投入に失敗しました: {enqueue_result.error}",
            )

        return {
            "success": True,
            "video_id": video_id,
            "message": "LiveBoost解析を再開しました。チャンクの結合から開始します。",
            "new_status": "STEP_0_EXTRACT_FRAMES",
            "job_id": str(job.id),
        }

    except HTTPException:
        raise
    except Exception as exc:
        await db.rollback()
        logger.exception(f"[retry-analysis/live_boost] Failed: {exc}")
        raise HTTPException(
            status_code=500,
            detail=f"LiveBoost解析の再試行に失敗しました: {exc}",
        )


# =========================


# ──────────────────────────────────────────────
# Data-Driven Script Generation (User-facing)
# ──────────────────────────────────────────────
from pydantic import BaseModel, Field


class UserScriptRequest(BaseModel):
    product_focus: Optional[str] = Field(None, description="Product to emphasize")
    tone: str = Field("professional_friendly", description="Script tone")
    language: str = Field("ja", description="Output language")
    duration_minutes: int = Field(10, ge=1, le=60, description="Target duration in minutes")
    cross_video: bool = Field(True, description="Include patterns from other videos")


@router.post("/{video_id}/generate-script")
async def generate_script_for_video(
    video_id: str,
    body: UserScriptRequest,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Generate a data-driven live commerce script based on real performance data.

    This analyzes the video's sales moments, product exposures, and audio transcripts
    to create a script grounded in actual sales data — not generic AI guesses.
    """
     # Verify video exists (any authenticated user can generate scripts)
    result = await db.execute(
        text("SELECT id, user_id FROM videos WHERE id = :vid"),
        {"vid": video_id},
    )
    video = result.fetchone()
    if not video:
        raise HTTPException(status_code=404, detail="Video not found")
    from app.services.winning_patterns_service import generate_data_driven_script

    try:
        result = await generate_data_driven_script(
            db=db,
            video_id=video_id,
            product_focus=body.product_focus,
            tone=body.tone,
            language=body.language,
            duration_minutes=body.duration_minutes,
            cross_video=body.cross_video,
        )
        return result
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.exception(f"Script generation failed: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{video_id}/winning-patterns")
async def get_video_winning_patterns(
    video_id: str,
    current_user: dict = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Get winning patterns (CTA phrases, product durations, top phases)
    extracted from this video's real performance data.
    """
     # Verify video exists (any authenticated user can view winning patterns)
    result = await db.execute(
        text("SELECT id FROM videos WHERE id = :vid"),
        {"vid": video_id},
    )
    video = result.fetchone()
    if not video:
        raise HTTPException(status_code=404, detail="Video not found")
    from app.services.winning_patterns_service import (
        extract_cta_phrases,
        analyze_product_durations,
        extract_top_phases,
    )

    try:
        cta_phrases = await extract_cta_phrases(db, video_id)
        product_durations = await analyze_product_durations(db, video_id)
        top_phases = await extract_top_phases(db, video_id, limit=10)

        return {
            "video_id": video_id,
            "cta_phrases": cta_phrases,
            "product_durations": product_durations,
            "top_phases": top_phases,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# Brand assignment for video + all its clips
# ============================================================

from fastapi import Header, Query as FastQuery

@router.patch("/{video_id}/brand")
async def assign_brand_to_video(
    video_id: str,
    client_id: str = FastQuery(..., description="Brand client_id to assign (empty string to unassign)"),
    db: AsyncSession = Depends(get_db),
    x_admin_key: Optional[str] = Header(None, alias="X-Admin-Key"),
):
    """
    Assign (or unassign) a brand to a video AND all its clips at once.
    - Updates videos.brand_client_id
    - Creates/updates widget_clip_assignments for every clip of this video
    """
    # Auth check (reuse same pattern as clip_db)
    _ADMIN_ID = os.getenv("ADMIN_ID", "aither")
    _ADMIN_PASS = os.getenv("ADMIN_PASS", "hub")
    valid_key = f"{_ADMIN_ID}:{_ADMIN_PASS}"
    if x_admin_key != valid_key:
        raise HTTPException(status_code=403, detail="Admin only")

    try:
        # Check video exists
        vid_row = await db.execute(
            text("SELECT id FROM videos WHERE id = :vid"),
            {"vid": video_id},
        )
        if not vid_row.first():
            raise HTTPException(status_code=404, detail="Video not found")

        effective_brand = client_id if client_id else None

        # Update video's brand_client_id
        await db.execute(
            text("UPDATE videos SET brand_client_id = :bid WHERE id = :vid"),
            {"bid": effective_brand, "vid": video_id},
        )

        # Get all clips for this video
        clips_result = await db.execute(
            text("SELECT id FROM video_clips WHERE video_id = :vid"),
            {"vid": video_id},
        )
        clip_ids = [str(row[0]) for row in clips_result.fetchall()]

        assigned_count = 0
        if effective_brand and clip_ids:
            # Check brand exists
            brand_check = await db.execute(
                text("SELECT client_id FROM widget_clients WHERE client_id = :bid AND is_active = TRUE"),
                {"bid": effective_brand},
            )
            if not brand_check.first():
                raise HTTPException(status_code=404, detail="Brand not found")

            # Get next sort order
            max_order_row = await db.execute(
                text("SELECT COALESCE(MAX(sort_order), -1) + 1 FROM widget_clip_assignments WHERE client_id = :cid"),
                {"cid": effective_brand},
            )
            next_order = max_order_row.scalar() or 0

            for clip_id in clip_ids:
                await db.execute(
                    text("""
                        INSERT INTO widget_clip_assignments (id, client_id, clip_id, sort_order, is_active, created_at)
                        VALUES (:id, :client_id, :clip_id, :sort_order, TRUE, NOW())
                        ON CONFLICT (client_id, clip_id) DO UPDATE
                        SET is_active = TRUE, sort_order = :sort_order
                    """),
                    {
                        "id": str(uuid_module.uuid4()),
                        "client_id": effective_brand,
                        "clip_id": clip_id,
                        "sort_order": next_order,
                    },
                )
                next_order += 1
                assigned_count += 1
        elif not effective_brand and clip_ids:
            # Unassign: deactivate all clip assignments for this video's clips
            # (We don't know which brand was previously assigned, so deactivate all)
            for clip_id in clip_ids:
                await db.execute(
                    text("UPDATE widget_clip_assignments SET is_active = FALSE WHERE clip_id = :cid"),
                    {"cid": clip_id},
                )

        await db.commit()

        return {
            "status": "assigned" if effective_brand else "unassigned",
            "video_id": video_id,
            "brand_client_id": effective_brand,
            "clips_affected": assigned_count if effective_brand else len(clip_ids),
        }
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception(f"Brand assign failed for video {video_id}: {exc}")
        raise HTTPException(status_code=500, detail=f"Brand assign failed: {exc}")


# ============================================================
# Include split sub-modules
# ============================================================
from app.api.v1.endpoints.video_clips import router as clips_router
from app.api.v1.endpoints.video_products import router as products_router
from app.api.v1.endpoints.video_sales import router as sales_router
from app.api.v1.endpoints.video_excel import router as excel_router

# Merge sub-routers into the main video router
for sub in [clips_router, products_router, sales_router, excel_router]:
    for route in sub.routes:
        router.routes.append(route)


