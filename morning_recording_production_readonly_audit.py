from __future__ import annotations

import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import quote

import requests
from openpyxl import load_workbook

BASE_URL = "https://lcjmall.com"
SOURCE = Path("/home/ubuntu/upload/pasted_file_06GgiQ_LCJ経営管理表_经营用账户.xlsx")
OUTPUT = Path(__file__).resolve().with_name("morning_recording_production_readonly_audit.json")
TODAY = "2026-08-27"


def trpc_json(response: requests.Response):
    response.raise_for_status()
    payload = response.json()
    if isinstance(payload, list):
        payload = payload[0]
    if "error" in payload:
        raise RuntimeError(payload["error"].get("json", {}).get("message") or "tRPC error")
    return payload["result"]["data"]["json"]


def query(session: requests.Session, procedure: str, value=None):
    encoded = quote(json.dumps({"json": value}, ensure_ascii=False, separators=(",", ":")))
    return trpc_json(session.get(f"{BASE_URL}/api/trpc/{procedure}?input={encoded}", timeout=60))


def login() -> requests.Session:
    workbook = load_workbook(SOURCE, data_only=False, read_only=False)
    sheet = workbook["经营用账户"]
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if "lcj系统登录网站" in str(values[0] or "").strip().lower():
            email = str(values[2] or "").strip()
            password = str(values[3] or "").strip()
            session = requests.Session()
            result = trpc_json(session.post(
                f"{BASE_URL}/api/trpc/auth.login",
                json={"json": {"email": email, "password": password}},
                timeout=30,
            ))
            if not result.get("success"):
                raise RuntimeError("Production login failed")
            return session
    raise RuntimeError("LCJ login credential row not found")


def summarize_records(records: list[dict]) -> dict:
    durations = [int(record.get("durationSeconds") or 0) for record in records]
    return {
        "count": len(records),
        "statusCounts": dict(Counter(str(record.get("status") or "unknown") for record in records)),
        "validCount": sum(1 for record in records if record.get("isValid") is True),
        "invalidCount": sum(1 for record in records if record.get("isValid") is False),
        "tooShortCount": sum(1 for record in records if record.get("invalidReason") == "too_short"),
        "durationsSeconds": sorted(durations),
        "minimumDurationSecondsSeen": sorted({int(record.get("minimumDurationSeconds") or 0) for record in records}),
    }


session = login()
settings = query(session, "morningMeeting.getTeamMeetingSettings")
today = query(session, "morningMeeting.getTodayDailyRecordings", {"date": TODAY})
principles_history = query(session, "morningMeeting.getSeparatedHistory", {
    "type": "principles", "limit": 100, "offset": 0, "dateFrom": TODAY, "dateTo": TODAY,
})
team_history = query(session, "morningMeeting.getSeparatedHistory", {
    "type": "team", "limit": 100, "offset": 0, "dateFrom": TODAY, "dateTo": TODAY,
})
legacy_history = query(session, "morningMeeting.getSeparatedHistory", {
    "type": "legacy", "limit": 100, "offset": 0, "dateFrom": TODAY, "dateTo": TODAY,
})

principles = principles_history.get("records") or []
teams = team_history.get("records") or []
members = today.get("members") or []
member_team_by_key = {str(member.get("targetKey") or ""): member.get("teamCode") for member in members}
report = {
    "checkedAt": datetime.now(timezone.utc).isoformat(),
    "date": TODAY,
    "productionWrites": 0,
    "settings": settings,
    "today": {
        "minimumDurationSeconds": today.get("minimumDurationSeconds"),
        "memberCount": len(members),
        "principlesCompletedCount": sum(1 for member in members if member.get("principlesCompleted")),
        "principlesInvalidCount": sum(1 for member in members if member.get("principlesInvalidReason")),
        "teamAttendanceCount": sum(1 for member in members if member.get("attendedTeamMeeting")),
        "teamMeetings": {
            code: None if record is None else {
                "id": record.get("id"),
                "status": record.get("status"),
                "durationSeconds": record.get("durationSeconds"),
                "isValid": record.get("isValid"),
                "invalidReason": record.get("invalidReason"),
                "participantCount": record.get("participantCount"),
                "createdAt": record.get("createdAt"),
            }
            for code, record in (today.get("teamMeetings") or {}).items()
        },
    },
    "principlesHistory": summarize_records(principles),
    "teamHistory": summarize_records(teams),
    "teamRecords": [
        {
            "id": record.get("id"),
            "date": record.get("date"),
            "recordingKind": record.get("recordingKind"),
            "teamCode": record.get("teamCode"),
            "dailyKey": record.get("dailyKey"),
            "durationSeconds": record.get("durationSeconds"),
            "status": record.get("status"),
            "isValid": record.get("isValid"),
            "participantCount": record.get("participantCount"),
            "participantTeamCounts": dict(Counter(
                str(member_team_by_key.get(str(participant.get("targetKey") or "")) or "unknown")
                for participant in (record.get("participantSnapshot") or [])
            )),
            "createdAt": record.get("createdAt"),
        }
        for record in teams
    ],
    "legacyHistory": {"count": len(legacy_history.get("records") or [])},
    "shortPrinciples": [
        {
            "id": record.get("id"),
            "durationSeconds": record.get("durationSeconds"),
            "status": record.get("status"),
            "isValid": record.get("isValid"),
            "invalidReason": record.get("invalidReason"),
            "createdAt": record.get("createdAt"),
        }
        for record in principles
        if int(record.get("durationSeconds") or 0) < int(settings.get("minimumDurationSeconds") or 0)
    ],
}
OUTPUT.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
print(json.dumps(report, ensure_ascii=False, indent=2))
